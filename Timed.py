# Timed.py - Professional Time Tracking Module for Engineers
import os
import json
import csv
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import datetime, timedelta
import calendar
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.dates as mdates
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfutils
from reportlab.pdfgen import canvas
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import time
from cryptography.fernet import Fernet
import base64
import platform
import subprocess


# Constants after the imports
BILLABLE_TYPES = ["Billable", "Non-Billable", "Internal", "Administrative"]


class TimeTracker:
    def __init__(self, parent, tool_dir, appdata_dir, username, theme, log_action):
        self.parent = parent
        self.tool_dir = tool_dir
        self.appdata_dir = appdata_dir
        self.username = username
        self.theme = theme
        self.log_action = log_action
        
        # Setup time tracking directories
        self.time_dir = tool_dir / "Time_Tracking"
        self.time_reports_dir = self.time_dir / "Reports"
        self.time_dir.mkdir(exist_ok=True)
        self.time_reports_dir.mkdir(exist_ok=True)
        
        # Data file
        self.time_data_file = self.appdata_dir / "time_tracking.json"
        
        # Encryption key for time data
        self.key_file = self.appdata_dir / "time_key.key"
        if not self.key_file.exists():
            key = Fernet.generate_key()
            with open(self.key_file, 'wb') as f:
                f.write(key)
        else:
            with open(self.key_file, 'rb') as f:
                key = f.read()
                
        self.cipher = Fernet(key)
        
        # Load time data
        self.time_data = self.load_time_data()
        
        # Current tracking state
        self.current_timer = None
        self.timer_running = False
        self.start_time = None
        self.elapsed_time = timedelta(0)  
        self.current_project = None
        self.current_task = None
        self.paused_timers = []  # Store paused timers
        self.break_start_time = None
        self.break_elapsed = timedelta(0)
        
        # Create UI
        self.create_ui()

        self.billable_type = "Billable"  # Default billable type
        self.project_estimates = {}  # Store project estimates

    def create_ui(self):
        """Create the time tracking tab UI"""
        self.time_tab = ttk.Frame(self.parent.notebook)
        self.parent.notebook.add(self.time_tab, text="Time Tracking")
        
        # Configure smaller button styles
        style = ttk.Style()
        style.configure('Small.TButton', padding=(4, 2))
        style.configure('Small.Primary.TButton', padding=(4, 2), background=self.theme["primary_color"])
        style.configure('Small.Secondary.TButton', padding=(4, 2), background=self.theme["secondary_color"])
        style.configure('Small.Accent.TButton', padding=(4, 2), background=self.theme["accent_color"])
        # In the create_ui method, replace the style configuration lines:
        success_color = self.theme.get("success_color", self.theme["accent_color"])
        style.configure('Small.Success.TButton', padding=(4, 2), background=success_color)
        # Main container with paned window for resizable sections
        main_pane = ttk.PanedWindow(self.time_tab, orient=tk.VERTICAL)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Top section - Timer and current tracking
        top_frame = ttk.LabelFrame(main_pane, text="Active Time Tracking", padding=15)
        main_pane.add(top_frame, weight=1)
        
        # Timer display - moved to the very top
        timer_frame = ttk.Frame(top_frame)
        timer_frame.pack(fill=tk.X, pady=10)
        
        self.timer_var = tk.StringVar(value="00:00:00")
        self.timer_label = ttk.Label(timer_frame, textvariable=self.timer_var, 
                                    font=('Courier', 24, 'bold'), 
                                    foreground=self.theme["primary_color"])
        self.timer_label.pack(pady=10)
        
        # Break timer display
        self.break_timer_var = tk.StringVar(value="Break: 00:00:00")
        self.break_timer_label = ttk.Label(timer_frame, textvariable=self.break_timer_var, 
                                          font=('Courier', 12), 
                                          foreground=self.theme["accent_color"])
        self.break_timer_label.pack(pady=5)
        
        
        # Timer controls
        control_frame = ttk.Frame(top_frame)
        control_frame.pack(fill=tk.X, pady=5)

        # Add this in the control_frame section, after the billable combo
        ttk.Label(control_frame, text="Paused Timers:").grid(row=4, column=0, sticky=tk.W, padx=5, pady=5)
        self.paused_btn = ttk.Button(control_frame, text="View (0)", 
                                    command=self.show_paused_timers, width=10,
                                    style='Small.Secondary.TButton')
        self.paused_btn.grid(row=4, column=1, sticky=tk.W, padx=5, pady=5)
        
        # Project selection
        ttk.Label(control_frame, text="Project:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.project_var = tk.StringVar()
        self.project_combo = ttk.Combobox(control_frame, textvariable=self.project_var, width=30)
        self.project_combo.grid(row=0, column=1, padx=5, pady=5)
        self.project_combo.bind('<KeyRelease>', self.update_project_suggestions)
        
        # Task description
        ttk.Label(control_frame, text="Task:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.task_var = tk.StringVar()
        self.task_combo = ttk.Combobox(control_frame, textvariable=self.task_var, width=30)
        self.task_combo.grid(row=1, column=1, padx=5, pady=5)
        self.task_combo.bind('<KeyRelease>', self.update_task_suggestions)
        
        # Billable type selection
        ttk.Label(control_frame, text="Billable Type:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        self.billable_var = tk.StringVar(value="Billable")
        self.billable_combo = ttk.Combobox(control_frame, textvariable=self.billable_var, 
                                        values=BILLABLE_TYPES, width=30, state="readonly")
        self.billable_combo.grid(row=2, column=1, padx=5, pady=5)

        # Add project estimate field
        ttk.Label(control_frame, text="Est. Hours:").grid(row=3, column=0, sticky=tk.W, padx=5, pady=5)
        self.estimate_var = tk.StringVar()
        self.estimate_entry = ttk.Entry(control_frame, textvariable=self.estimate_var, width=30)
        self.estimate_entry.grid(row=3, column=1, padx=5, pady=5)

        # Timer buttons - using smaller buttons
        btn_frame = ttk.Frame(top_frame)
        btn_frame.pack(fill=tk.X, pady=10)
        
        self.start_btn = ttk.Button(btn_frame, text="Start Timer", command=self.start_timer,
                                style='Small.Primary.TButton')
        self.start_btn.pack(side=tk.LEFT, padx=5)
        
        self.pause_btn = ttk.Button(btn_frame, text="Pause", command=self.pause_timer,
                                style='Small.Secondary.TButton', state=tk.DISABLED)
        self.pause_btn.pack(side=tk.LEFT, padx=5)
        
        self.stop_btn = ttk.Button(btn_frame, text="Stop & Save", command=self.stop_timer,
                                style='Small.Accent.TButton', state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        
        # Add break button - MOVED BEFORE manual entry
        self.break_btn = ttk.Button(btn_frame, text="Start Break", command=self.toggle_break,
                                   style='Small.Secondary.TButton', state=tk.DISABLED)
        self.break_btn.pack(side=tk.RIGHT, padx=5)
        
        # Add manual entry button
        self.manual_btn = ttk.Button(btn_frame, text="Add Manual Entry", 
                                    command=self.add_manual_entry, style='Small.Secondary.TButton')
        self.manual_btn.pack(side=tk.RIGHT, padx=5)
        
        # Bottom section - Time entries and reports
        bottom_frame = ttk.LabelFrame(main_pane, text="Time Entries & Reports", padding=15)
        main_pane.add(bottom_frame, weight=1)  # Reduced weight from 2 to 1
        
        # Time entries table
        table_frame = ttk.Frame(bottom_frame)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Create treeview with scrollbar - reduced height from 10 to 5
        columns = ("date", "project", "task", "billable", "start_time", "end_time", "duration", "description")
        self.time_tree = ttk.Treeview(table_frame, columns=columns, show="headings", height=5)  # Reduced height

        # Define headings  
        self.time_tree.heading("date", text="Date")
        self.time_tree.heading("project", text="Project")
        self.time_tree.heading("task", text="Task")
        self.time_tree.heading("billable", text="Billable Type")
        self.time_tree.heading("start_time", text="Start Time")
        self.time_tree.heading("end_time", text="End Time")
        self.time_tree.heading("duration", text="Duration")
        self.time_tree.heading("description", text="Description")

        # Set column widths
        self.time_tree.column("date", width=100, anchor=tk.CENTER)
        self.time_tree.column("project", width=120, anchor=tk.W)
        self.time_tree.column("task", width=120, anchor=tk.W)
        self.time_tree.column("billable", width=100, anchor=tk.CENTER)
        self.time_tree.column("start_time", width=90, anchor=tk.CENTER)
        self.time_tree.column("end_time", width=90, anchor=tk.CENTER)
        self.time_tree.column("duration", width=90, anchor=tk.CENTER)
        self.time_tree.column("description", width=150, anchor=tk.W)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.time_tree.yview)
        self.time_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.time_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Report buttons - using smaller buttons
        report_frame = ttk.Frame(bottom_frame)
        report_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(report_frame, text="Refresh", command=self.refresh_entries,
                style='Small.Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
        ttk.Button(report_frame, text="Daily Report", 
                command=lambda: self.generate_report('daily'),
                style='Small.Primary.TButton').pack(side=tk.LEFT, padx=5)
        
        ttk.Button(report_frame, text="Monthly Report", 
                command=lambda: self.generate_report('monthly'),
                style='Small.Primary.TButton').pack(side=tk.LEFT, padx=5)
        
        # Add a button for project management
        ttk.Button(report_frame, text="Manage Projects", 
                command=self.manage_projects,
                style='Small.Primary.TButton').pack(side=tk.LEFT, padx=5)

        # Add a button for billing summary
        ttk.Button(report_frame, text="Billing Summary", 
                command=self.show_billing_summary,
                style='Small.Success.TButton').pack(side=tk.LEFT, padx=5)
        
        # Add a button for burn-down chart
        ttk.Button(report_frame, text="Burn-Down Chart", 
                command=self.show_burndown_chart,
                style='Small.Success.TButton').pack(side=tk.LEFT, padx=5)
        
        # Add a button to view selected report
        ttk.Button(report_frame, text="View Selected", 
                command=self.view_selected_report,
                style='Small.Success.TButton').pack(side=tk.LEFT, padx=5)
        
        ttk.Button(report_frame, text="Delete Selected", 
                command=self.delete_entry,
                style='Small.Accent.TButton').pack(side=tk.RIGHT, padx=5)
        
        # Load initial data
        self.update_project_suggestions()
        self.update_task_suggestions()
        self.refresh_entries()
        self.update_paused_timers_display()

    def show_paused_timers(self):
        """Show paused timers in a modal dialog"""
        if not self.paused_timers:
            messagebox.showinfo("Paused Timers", "No paused timers")
            return
            
        dialog = tk.Toplevel(self.parent.root)
        dialog.title("Paused Timers")
        dialog.geometry("600x400")
        dialog.transient(self.parent.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = self.parent.root.winfo_x() + (self.parent.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.parent.root.winfo_y() + (self.parent.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Create a frame for the list
        list_frame = ttk.Frame(dialog, padding=10)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create a treeview to display paused timers
        columns = ("project", "task", "billable", "duration")
        tree = ttk.Treeview(list_frame, columns=columns, show="headings", height=10)
        
        tree.heading("project", text="Project")
        tree.heading("task", text="Task")
        tree.heading("billable", text="Billable Type")
        tree.heading("duration", text="Duration")
        
        tree.column("project", width=150)
        tree.column("task", width=150)
        tree.column("billable", width=100)
        tree.column("duration", width=100, anchor=tk.CENTER)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Populate with paused timers
        for timer in self.paused_timers:
            elapsed = timer['elapsed_time']
            hours, remainder = divmod(elapsed.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            time_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
            
            tree.insert("", tk.END, values=(
                timer['project'],
                timer['task'],
                timer['billable_type'],
                time_str
            ))
        
        # Button frame
        btn_frame = ttk.Frame(dialog, padding=10)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)
        
        def resume_selected():
            selection = tree.focus()
            if not selection:
                messagebox.showwarning("Warning", "Please select a timer to resume")
                return
                
            index = tree.index(selection)
            if index < len(self.paused_timers):
                self.resume_timer(self.paused_timers[index])
                dialog.destroy()
        
        def discard_selected():
            selection = tree.focus()
            if not selection:
                messagebox.showwarning("Warning", "Please select a timer to discard")
                return
                
            index = tree.index(selection)
            if index < len(self.paused_timers):
                self.discard_timer(self.paused_timers[index])
                # Refresh the list
                for item in tree.get_children():
                    tree.delete(item)
                for timer in self.paused_timers:
                    elapsed = timer['elapsed_time']
                    hours, remainder = divmod(elapsed.total_seconds(), 3600)
                    minutes, seconds = divmod(remainder, 60)
                    time_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
                    
                    tree.insert("", tk.END, values=(
                        timer['project'],
                        timer['task'],
                        timer['billable_type'],
                        time_str
                    ))
                
                if not self.paused_timers:
                    dialog.destroy()
        
        ttk.Button(btn_frame, text="Resume Selected", command=resume_selected, 
                style='Small.Success.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Discard Selected", command=discard_selected,
                style='Small.Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Close", command=dialog.destroy,
                style='Small.Secondary.TButton').pack(side=tk.RIGHT, padx=5)


    def toggle_break(self):
        """Toggle break on/off"""
        if self.break_start_time is None:
            # Start break
            self.break_start_time = datetime.now()
            self.break_btn.config(text="End Break", style='Small.Accent.TButton')
            self.log_action("Break started")
            # Update break timer display
            self.update_break_timer()
        else:
            # End break
            break_end = datetime.now()
            self.break_elapsed += break_end - self.break_start_time
            self.break_start_time = None
            self.break_btn.config(text="Start Break", style='Small.Secondary.TButton')
            self.log_action(f"Break ended. Duration: {self.break_elapsed}")
            # Reset break timer display
            self.break_timer_var.set("Break: 00:00:00")
            # Ensure main timer continues to update
            if self.timer_running:
                self.update_timer_display()

    def update_break_timer(self):
        """Update the break timer display"""
        if self.break_start_time is not None:
            elapsed = datetime.now() - self.break_start_time + self.break_elapsed
            hours, remainder = divmod(elapsed.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            self.break_timer_var.set(f"Break: {int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}")
            self.parent.root.after(1000, self.update_break_timer)

    def update_paused_timers_display(self):
        """Update the paused timers button text"""
        count = len(self.paused_timers)
        self.paused_btn.config(text=f"View ({count})")
        
        # Enable/disable the button based on whether there are paused timers
        if count > 0:
            self.paused_btn.state(['!disabled'])
        else:
            self.paused_btn.state(['disabled'])
            
    def resume_timer(self, timer):
        """Resume a paused timer"""
        # Remove from paused timers
        self.paused_timers = [t for t in self.paused_timers if t != timer]
        self.update_paused_timers_display()
        
        # Set up the timer
        self.project_var.set(timer['project'])
        self.task_var.set(timer['task'])
        self.billable_var.set(timer['billable_type'])
        self.elapsed_time = timer['elapsed_time']
        self.start_time = datetime.now() - self.elapsed_time
        self.current_project = timer['project']
        self.current_task = timer['task']
        self.billable_type = timer['billable_type']
        
        # Start the timer
        self.timer_running = True
        self.start_btn.config(state=tk.DISABLED)
        self.pause_btn.config(state=tk.NORMAL, text="Pause")
        self.stop_btn.config(state=tk.NORMAL)
        self.project_combo.config(state=tk.DISABLED)
        self.task_combo.config(state=tk.DISABLED)
        self.billable_combo.config(state=tk.DISABLED)
        self.estimate_entry.config(state=tk.DISABLED)
        self.break_btn.config(state=tk.NORMAL)
        
        # Start timer update
        self.update_timer_display()
        
        self.log_action(f"Resumed timer for project: {timer['project']}, task: {timer['task']}")

    def discard_timer(self, timer):
        """Discard a paused timer"""
        self.paused_timers = [t for t in self.paused_timers if t != timer]
        self.update_paused_timers_display()
        self.log_action(f"Discarded paused timer for project: {timer['project']}")

    # ... (rest of the methods remain the same, only UI positioning and button sizes changed)


    def view_selected_report(self):
        """Open the selected report file with the default application"""
        # Ask user to select a report file
        file_path = filedialog.askopenfilename(
            initialdir=self.time_reports_dir,
            title="Select Report File",
            filetypes=(("PDF files", "*.pdf"), ("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
        
        if not file_path:
            return
            
        try:
            # Open the file with the default application
            if platform.system() == "Windows":
                os.startfile(file_path)
            elif platform.system() == "Darwin":
                subprocess.Popen(["open", file_path])
            else:
                subprocess.Popen(["xdg-open", file_path])
                
            self.log_action(f"Opened report: {os.path.basename(file_path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open file: {str(e)}")
            self.log_action(f"Failed to open report: {str(e)}")

    def show_billing_summary(self):
        """Show a summary of billable vs non-billable hours"""
        # Calculate totals
        billable_hours = 0
        non_billable_hours = 0
        
        for entry in self.time_data.get('entries', []):
            hours = entry['duration_seconds'] / 3600
            if entry.get('billable', 'Billable') == 'Billable':
                billable_hours += hours
            else:
                non_billable_hours += hours
        
        total_hours = billable_hours + non_billable_hours
        
        # Create summary dialog
        dialog = tk.Toplevel(self.parent.root)
        dialog.title("Billing Summary")
        dialog.geometry("400x300")
        dialog.transient(self.parent.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = self.parent.root.winfo_x() + (self.parent.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.parent.root.winfo_y() + (self.parent.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Summary frame
        summary_frame = ttk.Frame(dialog, padding=20)
        summary_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(summary_frame, text="BILLING SUMMARY", font=('Arial', 14, 'bold')).pack(pady=10)
        
        ttk.Label(summary_frame, text=f"Total Hours: {total_hours:.2f}").pack(pady=5)
        ttk.Label(summary_frame, text=f"Billable Hours: {billable_hours:.2f} ({billable_hours/total_hours*100:.1f}%)", 
                foreground="green").pack(pady=5)
        ttk.Label(summary_frame, text=f"Non-Billable Hours: {non_billable_hours:.2f} ({non_billable_hours/total_hours*100:.1f}%)", 
                foreground="blue").pack(pady=5)
        
        # Add breakdown by project
        ttk.Separator(summary_frame, orient='horizontal').pack(fill=tk.X, pady=20)
        ttk.Label(summary_frame, text="By Project:", font=('Arial', 12, 'bold')).pack(pady=5)
        
        # Create a scrollable frame for project breakdown
        project_frame = ttk.Frame(summary_frame)
        project_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(project_frame, height=150)
        scrollbar = ttk.Scrollbar(project_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Calculate hours by project
        project_hours = {}
        project_billable = {}
        
        for entry in self.time_data.get('entries', []):
            project = entry['project']
            hours = entry['duration_seconds'] / 3600
            billable = entry.get('billable', 'Billable') == 'Billable'
            
            if project not in project_hours:
                project_hours[project] = {'billable': 0, 'non_billable': 0}
            
            if billable:
                project_hours[project]['billable'] += hours
            else:
                project_hours[project]['non_billable'] += hours
        
        # Add project breakdown
        for project, hours in project_hours.items():
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill=tk.X, pady=2)
            
            ttk.Label(frame, text=project, width=20, anchor=tk.W).pack(side=tk.LEFT)
            ttk.Label(frame, text=f"{hours['billable']:.1f}h", width=10, foreground="green").pack(side=tk.LEFT)
            ttk.Label(frame, text=f"{hours['non_billable']:.1f}h", width=10, foreground="blue").pack(side=tk.LEFT)
            ttk.Label(frame, text=f"{hours['billable'] + hours['non_billable']:.1f}h", width=10).pack(side=tk.LEFT)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Add headers
        header_frame = ttk.Frame(summary_frame)
        header_frame.pack(fill=tk.X)
        
        ttk.Label(header_frame, text="Project", width=20, anchor=tk.W, font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
        ttk.Label(header_frame, text="Billable", width=10, foreground="green", font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
        ttk.Label(header_frame, text="Non-Bill", width=10, foreground="blue", font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
        ttk.Label(header_frame, text="Total", width=10, font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
        
        ttk.Button(summary_frame, text="Close", command=dialog.destroy).pack(pady=10)

    def show_burndown_chart(self):
        """Show a burn-down chart for selected project"""
        # Get unique projects
        projects = list(set([e['project'] for e in self.time_data.get('entries', [])]))
        
        if not projects:
            messagebox.showinfo("Info", "No projects found with time entries")
            return
        
        # Dialog to select project
        dialog = tk.Toplevel(self.parent.root)
        dialog.title("Select Project for Burn-Down Chart")
        dialog.geometry("400x200")
        dialog.transient(self.parent.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = self.parent.root.winfo_x() + (self.parent.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.parent.root.winfo_y() + (self.parent.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        ttk.Label(dialog, text="Select Project:").pack(pady=20)
        
        project_var = tk.StringVar()
        project_combo = ttk.Combobox(dialog, textvariable=project_var, values=projects, state="readonly")
        project_combo.pack(pady=10)
        
        def generate_chart():
            project = project_var.get()
            if not project:
                messagebox.showwarning("Warning", "Please select a project")
                return
            
            dialog.destroy()
            self.generate_burndown_chart(project)
        
        ttk.Button(dialog, text="Generate Chart", command=generate_chart, style='Primary.TButton').pack(pady=20)

    def generate_burndown_chart(self, project):
        """Generate a burn-down chart for the selected project"""
        # Get project estimate
        estimate = self.time_data.get('project_estimates', {}).get(project, 0)
        
        if estimate <= 0:
            messagebox.showinfo("Info", f"No estimate found for project '{project}'. Please set an estimate first.")
            self.manage_projects()
            return
        
        # Get all entries for this project
        project_entries = [e for e in self.time_data.get('entries', []) if e['project'] == project]
        
        if not project_entries:
            messagebox.showinfo("Info", f"No time entries found for project '{project}'")
            return
        
        # Sort entries by date
        project_entries.sort(key=lambda x: x['date'])
        
        # Calculate cumulative hours by date
        date_hours = {}
        for entry in project_entries:
            date = entry['date']
            hours = entry['duration_seconds'] / 3600
            date_hours[date] = date_hours.get(date, 0) + hours
        
        # Create cumulative timeline
        dates = sorted(date_hours.keys())
        cumulative_hours = 0
        cumulative_data = []
        
        for date in dates:
            cumulative_hours += date_hours[date]
            cumulative_data.append((date, cumulative_hours))
        
        # Create ideal burn-down line (straight line from estimate to 0)
        start_date = min(dates)
        end_date = max(dates)
        
        # Create chart
        fig, ax = plt.subplots(figsize=(10, 6))
        
        # Plot actual burn-down
        dates_dt = [datetime.strptime(d, '%Y-%m-%d') for d, _ in cumulative_data]
        hours_remaining = [max(0, estimate - h) for _, h in cumulative_data]
        
        ax.plot(dates_dt, hours_remaining, 'o-', linewidth=2, markersize=6, label='Actual Hours Remaining')
        
        # Plot ideal burn-down
        if len(dates) > 1:
            ideal_dates = [datetime.strptime(start_date, '%Y-%m-%d'), 
                        datetime.strptime(end_date, '%Y-%m-%d')]
            ideal_hours = [estimate, 0]
            ax.plot(ideal_dates, ideal_hours, 'r--', linewidth=2, label='Ideal Burn-Down')
        
        # Format chart
        ax.set_xlabel('Date')
        ax.set_ylabel('Hours Remaining')
        ax.set_title(f'Burn-Down Chart: {project}\n(Estimate: {estimate} hours)')
        ax.legend()
        ax.grid(True, alpha=0.3)
        
        # Format x-axis dates
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        ax.xaxis.set_major_locator(mdates.WeekdayLocator())
        plt.xticks(rotation=45)
        
        plt.tight_layout()
        
        # Show chart in a new window
        chart_window = tk.Toplevel(self.parent.root)
        chart_window.title(f"Burn-Down Chart - {project}")
        chart_window.geometry("800x600")
        
        canvas = FigureCanvasTkAgg(fig, master=chart_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        ttk.Button(chart_window, text="Close", command=chart_window.destroy).pack(pady=10)

    def load_time_data(self):
        """Load encrypted time tracking data"""
        if self.time_data_file.exists():
            try:
                with open(self.time_data_file, 'rb') as f:
                    encrypted_data = f.read()
                decrypted_data = self.cipher.decrypt(encrypted_data)
                return json.loads(decrypted_data.decode())
            except:
                # If decryption fails, return empty data
                return {"entries": [], "projects": [], "tasks": [], "project_estimates": {}}
        return {"entries": [], "projects": [], "tasks": [], "project_estimates": {}}

    def save_time_data(self):
        """Save encrypted time tracking data"""
        encrypted_data = self.cipher.encrypt(json.dumps(self.time_data).encode())
        with open(self.time_data_file, 'wb') as f:
            f.write(encrypted_data)

    def update_project_suggestions(self, event=None):
        """Update project suggestions in combobox"""
        current = self.project_var.get()
        projects = self.time_data.get('projects', [])
        if current:
            self.project_combo['values'] = [p for p in projects if p.lower().startswith(current.lower())]
        else:
            self.project_combo['values'] = projects

    def update_task_suggestions(self, event=None):
        """Update task suggestions in combobox"""
        current = self.task_var.get()
        tasks = self.time_data.get('tasks', [])
        if current:
            self.task_combo['values'] = [t for t in tasks if t.lower().startswith(current.lower())]
        else:
            self.task_combo['values'] = tasks
    def manage_projects(self):
        """Dialog to manage projects and estimates"""
        dialog = tk.Toplevel(self.parent.root)
        dialog.title("Manage Projects")
        dialog.geometry("800x500")
        dialog.transient(self.parent.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = self.parent.root.winfo_x() + (self.parent.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.parent.root.winfo_y() + (self.parent.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Load project estimates
        self.project_estimates = self.time_data.get('project_estimates', {})
        
        # Create treeview for projects
        tree_frame = ttk.Frame(dialog)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        columns = ("project", "estimated_hours", "entry_count", "total_hours", "last_entry")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=10)
        
        tree.heading("project", text="Project")
        tree.heading("estimated_hours", text="Estimated Hours")
        tree.heading("entry_count", text="Entries")
        tree.heading("total_hours", text="Total Hours")
        tree.heading("last_entry", text="Last Entry")
                
        tree.column("project", width=200)
        tree.column("estimated_hours", width=120, anchor=tk.CENTER)
        tree.column("entry_count", width=80, anchor=tk.CENTER)
        tree.column("total_hours", width=100, anchor=tk.CENTER)
        tree.column("last_entry", width=100, anchor=tk.CENTER)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Populate tree with all projects
        projects = list(set([e['project'] for e in self.time_data.get('entries', [])] + 
                        self.time_data.get('projects', [])))
        
        project_stats = self.get_project_stats()
        for project in projects:
            est_hours = self.project_estimates.get(project, 0)
            stats = project_stats.get(project, {'count': 0, 'total_hours': 0, 'last_entry': None})
            
            last_entry_str = stats['last_entry'].strftime('%Y-%m-%d') if stats['last_entry'] else 'Never'
            
            tree.insert("", tk.END, values=(
                project,
                est_hours,
                stats['count'],
                f"{stats['total_hours']:.1f}",
                last_entry_str
            ))

        # Add project frame
        add_frame = ttk.Frame(dialog)
        add_frame.pack(fill=tk.X, padx=10, pady=5)
        
        ttk.Label(add_frame, text="Add New Project:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        new_project_var = tk.StringVar()
        ttk.Entry(add_frame, textvariable=new_project_var, width=30).grid(row=0, column=1, padx=5, pady=5)
        
        def add_project():
            project = new_project_var.get().strip()
            if project and project not in projects:
                # Add to tree with default values
                tree.insert("", tk.END, values=(project, 0, 0, "0.0", "Never"))
                projects.append(project)
                new_project_var.set("")
                self.log_action(f"Added new project: {project}")
            elif project in projects:
                messagebox.showwarning("Warning", f"Project '{project}' already exists")
        
        ttk.Button(add_frame, text="Add", command=add_project).grid(row=0, column=2, padx=5, pady=5)
        
        # Edit frame
        edit_frame = ttk.Frame(dialog)
        edit_frame.pack(fill=tk.X, padx=10, pady=10)
                
        ttk.Label(edit_frame, text="Selected Project:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        selected_project = tk.StringVar()
        ttk.Entry(edit_frame, textvariable=selected_project, state='readonly', width=30).grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(edit_frame, text="Estimated Hours:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        hours_var = tk.StringVar()
        ttk.Entry(edit_frame, textvariable=hours_var, width=30).grid(row=1, column=1, padx=5, pady=5)
        
        def on_select(event):
            selected = tree.focus()
            if selected:
                values = tree.item(selected, 'values')
                selected_project.set(values[0])
                hours_var.set(values[1])
        
        tree.bind('<<TreeviewSelect>>', on_select)
        
        def save_estimate():
            project = selected_project.get()
            if not project:
                messagebox.showwarning("Warning", "Please select a project first")
                return
                
            try:
                hours = float(hours_var.get())
                self.project_estimates[project] = hours
                
                # Update the tree
                for item in tree.get_children():
                    if tree.item(item, 'values')[0] == project:
                        # Get current values
                        current_values = tree.item(item, 'values')
                        # Update just the estimate, keep other values
                        tree.item(item, values=(project, hours, current_values[2], current_values[3], current_values[4]))
                        break
                        
                messagebox.showinfo("Success", f"Updated estimate for '{project}' to {hours} hours")
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid number for hours")
        
        def delete_project():
            project = selected_project.get()
            if not project:
                messagebox.showwarning("Warning", "Please select a project to delete")
                return
                
            # Get project statistics
            stats = self.get_project_stats().get(project, {'count': 0, 'total_hours': 0})
            
            # Create confirmation message
            if stats['count'] > 0:
                message = (
                    f"Project: {project}\n\n"
                    f"This project has {stats['count']} time entries totaling {stats['total_hours']:.1f} hours.\n\n"
                    "Deleting the project will also remove all associated time entries.\n\n"
                    "Are you sure you want to delete this project and all its time entries?"
                )
            else:
                message = f"Are you sure you want to delete the project '{project}'?"
            
            response = messagebox.askyesno(
                "Confirm Delete", 
                message,
                icon='warning'
            )
            if not response:
                return
            
            # Remove from tree
            for item in tree.get_children():
                if tree.item(item, 'values')[0] == project:
                    tree.delete(item)
                    break
            
            # Remove from estimates
            if project in self.project_estimates:
                del self.project_estimates[project]
            
            # Remove from projects list
            if 'projects' in self.time_data and project in self.time_data['projects']:
                self.time_data['projects'].remove(project)
            
            # Remove time entries
            self.time_data['entries'] = [e for e in self.time_data.get('entries', []) if e['project'] != project]
            
            # Clear selection
            selected_project.set("")
            hours_var.set("")
            
            self.save_time_data()
            self.refresh_entries()
            self.log_action(f"Deleted project: {project}")
            messagebox.showinfo("Success", f"Project '{project}' and all its time entries have been deleted")
    
        # Button frame
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        # Add buttons to the frame
        ttk.Button(btn_frame, text="Save Estimate", command=save_estimate).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Delete Project", command=delete_project, style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        
        def save_all():
            # Save projects and estimates
            all_projects = []
            for item in tree.get_children():
                project = tree.item(item, 'values')[0]
                all_projects.append(project)
            
            self.time_data['projects'] = all_projects
            self.time_data['project_estimates'] = self.project_estimates
            self.save_time_data()
            dialog.destroy()
            self.update_project_suggestions()
            self.log_action("Updated projects and estimates")
            messagebox.showinfo("Success", "All projects and estimates have been saved")
        

    def delete_project_entries(self, project):
        """Delete all time entries for a specific project"""
        if 'entries' in self.time_data:
            self.time_data['entries'] = [e for e in self.time_data['entries'] if e['project'] != project]
            self.save_time_data()
            self.refresh_entries()
            self.log_action(f"Deleted all time entries for project: {project}")

    def get_project_stats(self):
        """Get statistics about projects and their time entries"""
        project_stats = {}
        
        for entry in self.time_data.get('entries', []):
            project = entry['project']
            if project not in project_stats:
                project_stats[project] = {
                    'count': 0,
                    'total_hours': 0,
                    'last_entry': None
                }
            
            project_stats[project]['count'] += 1
            project_stats[project]['total_hours'] += entry['duration_seconds'] / 3600
            
            entry_date = datetime.strptime(entry['date'], '%Y-%m-%d')
            if (project_stats[project]['last_entry'] is None or 
                entry_date > project_stats[project]['last_entry']):
                project_stats[project]['last_entry'] = entry_date
        
        return project_stats

    def start_timer(self):
        """Start the timer"""
        project = self.project_var.get().strip()
        task = self.task_var.get().strip()
        billable_type = self.billable_var.get()
        
        if not project:
            messagebox.showerror("Error", "Please select a project")
            return
            
        if not task:
            messagebox.showerror("Error", "Please enter a task description")
            return
        
        # Check if project exists in saved projects
        if project not in self.time_data.get('projects', []):
            if messagebox.askyesno("New Project", f"Project '{project}' is not in your saved projects. Would you like to add it?"):
                self.time_data.setdefault('projects', []).append(project)
                self.save_time_data()
                self.update_project_suggestions()
            else:
                return
        
        # Save estimate if provided
        estimate = self.estimate_var.get().strip()
        if estimate:
            try:
                estimate_float = float(estimate)
                self.time_data.setdefault('project_estimates', {})[project] = estimate_float
                self.save_time_data()
            except ValueError:
                pass  # Silently ignore invalid estimates
        
        # Update tasks list if new
        if task not in self.time_data.get('tasks', []):
            self.time_data.setdefault('tasks', []).append(task)
            
        if billable_type not in self.time_data.get('billable_types', []):
            self.time_data.setdefault('billable_types', []).append(billable_type)
            
        self.save_time_data()
        
        # Start timer
        self.timer_running = True
        self.start_time = datetime.now()
        self.current_project = project
        self.current_task = task
        self.billable_type = billable_type
        
        # Update UI
        self.start_btn.config(state=tk.DISABLED)
        self.pause_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.NORMAL)
        self.project_combo.config(state=tk.DISABLED)
        self.task_combo.config(state=tk.DISABLED)
        self.billable_combo.config(state=tk.DISABLED)
        self.estimate_entry.config(state=tk.DISABLED)
        self.break_btn.config(state=tk.NORMAL)
        
        # Start timer update
        self.update_timer_display()
        
        self.log_action(f"Started timer for project: {project}, task: {task}, type: {billable_type}")

    def pause_timer(self):
        """Pause the timer and store it for later resumption"""
        if self.timer_running:
            # Calculate elapsed time
            elapsed = datetime.now() - self.start_time
            
            # Store paused timer
            paused_timer = {
                'project': self.current_project,
                'task': self.current_task,
                'billable_type': self.billable_type,
                'elapsed_time': elapsed,
                'start_time': self.start_time
            }
            
            self.paused_timers.append(paused_timer)
            
            # Update the main timer to show the paused time
            hours, remainder = divmod(elapsed.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            self.timer_var.set(f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}")
            
            # Reset current timer
            self.timer_running = False
            self.start_btn.config(state=tk.NORMAL)
            self.pause_btn.config(state=tk.DISABLED)
            self.stop_btn.config(state=tk.DISABLED)
            self.project_combo.config(state=tk.NORMAL)
            self.task_combo.config(state=tk.NORMAL)
            self.billable_combo.config(state="readonly")
            self.estimate_entry.config(state=tk.NORMAL)
            self.break_btn.config(state=tk.DISABLED)
            
            # Reset break timer if active
            if self.break_start_time is not None:
                break_end = datetime.now()
                self.break_elapsed += break_end - self.break_start_time
                self.break_start_time = None
                self.break_btn.config(text="Start Break", style='Secondary.TButton')
            
            # Update paused timers display
            self.update_paused_timers_display()
            
            self.log_action(f"Paused timer for project: {self.current_project}, task: {self.current_task}")

    def stop_timer(self):
        """Stop the timer and save the entry"""
        # Calculate duration (subtract break time)
        end_time = datetime.now()
        duration = end_time - self.start_time - self.break_elapsed
        
        # Create time entry
        entry = {
            "date": self.start_time.strftime("%Y-%m-%d"),
            "project": self.current_project,
            "task": self.current_task,
            "billable": self.billable_type,
            "start_time": self.start_time.strftime("%H:%M:%S"),
            "end_time": end_time.strftime("%H:%M:%S"),
            "duration": str(duration),
            "duration_seconds": duration.total_seconds(),
            "break_time": self.break_elapsed.total_seconds(),
            "description": ""  # Can be extended with a description field
        }
        
        # Add to data
        self.time_data.setdefault('entries', []).append(entry)
        self.save_time_data()
        
        # Reset UI
        self.timer_running = False
        self.timer_var.set("00:00:00")
        self.break_timer_var.set("Break: 00:00:00")
        self.start_btn.config(state=tk.NORMAL)
        self.pause_btn.config(state=tk.DISABLED, text="Pause")
        self.stop_btn.config(state=tk.DISABLED)
        self.project_combo.config(state=tk.NORMAL)
        self.task_combo.config(state=tk.NORMAL)
        self.billable_combo.config(state="readonly")
        self.estimate_entry.config(state=tk.NORMAL)
        self.break_btn.config(state=tk.DISABLED, text="Start Break", style='Secondary.TButton')
        
        # Reset elapsed time and break time
        self.elapsed_time = timedelta(0)
        self.break_elapsed = timedelta(0)
        self.break_start_time = None
        
        # Refresh entries table
        self.refresh_entries()
        
        self.log_action(f"Stopped timer and saved entry: {self.current_project} - {self.current_task}, Type: {self.billable_type}, Duration: {str(duration)}")

    def update_timer_display(self):
        """Update the timer display"""
        if self.timer_running:
            elapsed = datetime.now() - self.start_time - self.break_elapsed
            hours, remainder = divmod(elapsed.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            self.timer_var.set(f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}")
            
            # Update break timer if active
            if self.break_start_time is not None:
                self.update_break_timer()
            else:
                self.parent.root.after(1000, self.update_timer_display)
        else:
            # Display the paused time
            elapsed = self.elapsed_time
            hours, remainder = divmod(elapsed.total_seconds(), 3600)
            minutes, seconds = divmod(remainder, 60)
            self.timer_var.set(f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}")

    def add_manual_entry(self):
        """Add a manual time entry"""
        dialog = tk.Toplevel(self.parent.root)
        dialog.title("Add Manual Time Entry")
        dialog.geometry("500x550")  # Increased height to accommodate break time field
        dialog.transient(self.parent.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = self.parent.root.winfo_x() + (self.parent.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.parent.root.winfo_y() + (self.parent.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Form fields
        ttk.Label(dialog, text="Project:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=10)
        project_var = tk.StringVar()
        project_combo = ttk.Combobox(dialog, textvariable=project_var, width=30)
        project_combo.grid(row=0, column=1, padx=10, pady=10)
        project_combo['values'] = self.time_data.get('projects', [])
        
        ttk.Label(dialog, text="Task:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=10)
        task_var = tk.StringVar()
        task_combo = ttk.Combobox(dialog, textvariable=task_var, width=30)
        task_combo.grid(row=1, column=1, padx=10, pady=10)
        task_combo['values'] = self.time_data.get('tasks', [])
        
        ttk.Label(dialog, text="Billable Type:").grid(row=2, column=0, sticky=tk.W, padx=10, pady=10)
        billable_var = tk.StringVar(value="Billable")
        billable_combo = ttk.Combobox(dialog, textvariable=billable_var, 
                                    values=BILLABLE_TYPES, width=30, state="readonly")
        billable_combo.grid(row=2, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="Date:").grid(row=3, column=0, sticky=tk.W, padx=10, pady=10)
        date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        date_entry = ttk.Entry(dialog, textvariable=date_var, width=30)
        date_entry.grid(row=3, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="Start Time (HH:MM):").grid(row=4, column=0, sticky=tk.W, padx=10, pady=10)
        start_time_var = tk.StringVar(value=datetime.now().strftime("%H:%M"))
        start_time_entry = ttk.Entry(dialog, textvariable=start_time_var, width=30)
        start_time_entry.grid(row=4, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="End Time (HH:MM):").grid(row=5, column=0, sticky=tk.W, padx=10, pady=10)
        end_time_var = tk.StringVar(value=datetime.now().strftime("%H:%M"))
        end_time_entry = ttk.Entry(dialog, textvariable=end_time_var, width=30)
        end_time_entry.grid(row=5, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="Break Time (minutes):").grid(row=6, column=0, sticky=tk.W, padx=10, pady=10)
        break_time_var = tk.StringVar(value="0")
        break_time_entry = ttk.Entry(dialog, textvariable=break_time_var, width=30)
        break_time_entry.grid(row=6, column=1, padx=10, pady=10)
        
        ttk.Label(dialog, text="Description:").grid(row=7, column=0, sticky=tk.W, padx=10, pady=10)
        description_text = tk.Text(dialog, width=40, height=5)
        description_text.grid(row=7, column=1, padx=10, pady=10)
        
        # Buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.grid(row=8, column=0, columnspan=2, pady=20)
        
        def save_entry():
            # Validate inputs
            try:
                start_dt = datetime.strptime(f"{date_var.get()} {start_time_var.get()}", "%Y-%m-%d %H:%M")
                end_dt = datetime.strptime(f"{date_var.get()} {end_time_var.get()}", "%Y-%m-%d %H:%M")
                break_time = float(break_time_var.get()) * 60  # Convert to seconds
                
                if end_dt <= start_dt:
                    messagebox.showerror("Error", "End time must be after start time")
                    return
                    
                duration = end_dt - start_dt - timedelta(seconds=break_time)
                
                if duration.total_seconds() <= 0:
                    messagebox.showerror("Error", "Break time cannot exceed total time")
                    return
                    
                # Create entry
                entry = {
                    "date": date_var.get(),
                    "project": project_var.get(),
                    "task": task_var.get(),
                    "billable": billable_var.get(),
                    "start_time": start_time_var.get(),
                    "end_time": end_time_var.get(),
                    "duration": str(duration),
                    "duration_seconds": duration.total_seconds(),
                    "break_time": break_time,
                    "description": description_text.get("1.0", tk.END).strip()
                }
                
                # Update projects, tasks and billable types if new
                if project_var.get() not in self.time_data.get('projects', []):
                    self.time_data.setdefault('projects', []).append(project_var.get())
                    
                if task_var.get() not in self.time_data.get('tasks', []):
                    self.time_data.setdefault('tasks', []).append(task_var.get())
                    
                if billable_var.get() not in self.time_data.get('billable_types', []):
                    self.time_data.setdefault('billable_types', []).append(billable_var.get())
                
                # Add to data
                self.time_data.setdefault('entries', []).append(entry)
                self.save_time_data()
                
                # Refresh UI
                self.refresh_entries()
                
                dialog.destroy()
                self.log_action(f"Added manual time entry: {project_var.get()} - {task_var.get()}, Type: {billable_var.get()}")
                
            except ValueError:
                messagebox.showerror("Error", "Invalid date, time, or break time format. Use YYYY-MM-DD for date, HH:MM for time, and numbers for break time")
        
        ttk.Button(btn_frame, text="Save", command=save_entry, style='Primary.TButton').pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy, style='Secondary.TButton').pack(side=tk.LEFT, padx=10)

    def refresh_entries(self):
        """Refresh the time entries table"""
        # Clear existing entries
        for item in self.time_tree.get_children():
            self.time_tree.delete(item)
            
        # Add entries from data
        for entry in self.time_data.get('entries', []):
            # Format duration for display
            duration_seconds = entry['duration_seconds']
            hours, remainder = divmod(duration_seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            duration_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
            
            self.time_tree.insert("", tk.END, values=(
                entry['date'],
                entry['project'],
                entry['task'],
                entry.get('billable', 'Billable'),  # Default to Billable for backward compatibility
                entry['start_time'],
                entry['end_time'],
                duration_str,
                entry.get('description', '')
            ))

    def delete_entry(self):
        """Delete selected time entry"""
        selection = self.time_tree.selection()
        if not selection:
            messagebox.showwarning("Warning", "Please select an entry to delete")
            return
            
        if messagebox.askyesno("Confirm", "Are you sure you want to delete the selected entry?"):
            # Get the index of the selected item
            index = self.time_tree.index(selection[0])
            
            # Remove from data
            if index < len(self.time_data.get('entries', [])):
                self.time_data['entries'].pop(index)
                self.save_time_data()
                self.refresh_entries()
                self.log_action("Deleted time entry")

    def generate_report(self, period):
        """Generate a time tracking report"""
        # Get date range based on period
        today = datetime.now()
        if period == 'daily':
            date_str = today.strftime("%Y-%m-%d")
            entries = [e for e in self.time_data.get('entries', []) if e['date'] == date_str]
            title = f"Daily Time Report - {date_str}"
            filename = f"time_report_daily_{date_str}"
        else:  # monthly
            month_str = today.strftime("%Y-%m")
            entries = [e for e in self.time_data.get('entries', []) if e['date'].startswith(month_str)]
            title = f"Monthly Time Report - {today.strftime('%B %Y')}"
            filename = f"time_report_monthly_{month_str}"
            
        if not entries:
            messagebox.showinfo("Info", f"No time entries found for the selected {period} period")
            return
            
        # Ask for report format
        format_dialog = tk.Toplevel(self.parent.root)
        format_dialog.title("Report Format")
        format_dialog.geometry("300x150")
        format_dialog.transient(self.parent.root)
        format_dialog.grab_set()
        
        # Center the dialog
        format_dialog.update_idletasks()
        x = self.parent.root.winfo_x() + (self.parent.root.winfo_width() - format_dialog.winfo_width()) // 2
        y = self.parent.root.winfo_y() + (self.parent.root.winfo_height() - format_dialog.winfo_height()) // 2
        format_dialog.geometry(f"+{x}+{y}")
        
        ttk.Label(format_dialog, text="Select report format:").pack(pady=10)
        
        format_var = tk.StringVar(value="pdf")
        
        ttk.Radiobutton(format_dialog, text="PDF Report", variable=format_var, value="pdf").pack(anchor=tk.W, padx=20)
        ttk.Radiobutton(format_dialog, text="Excel Spreadsheet", variable=format_var, value="excel").pack(anchor=tk.W, padx=20)
        
        def generate():
            format_dialog.destroy()
            if format_var.get() == "pdf":
                self.export_to_pdf(entries, title, filename)
            else:
                self.export_to_excel(entries, title, filename)
        
        ttk.Button(format_dialog, text="Generate", command=generate, style='Primary.TButton').pack(pady=10)

    def export_to_pdf(self, entries, title, filename):
        """Export time entries to a PDF report"""
        chart_path = None
        temp_file_path = None
        qr_code_path = None
        
        try:
            # Create a temporary file first
            temp_file_path = self.time_reports_dir / f"{filename}_temp.pdf"
            final_file_path = self.time_reports_dir / f"{filename}.pdf"
            
            # Create PDF document with larger margins for header/footer
            doc = SimpleDocTemplate(str(temp_file_path), pagesize=A4,
                                rightMargin=72, leftMargin=72,
                                topMargin=120, bottomMargin=100)  # Increased bottom margin for QR code
            
            styles = getSampleStyleSheet()
            styles.add(ParagraphStyle(name='Center', alignment=1))
            
            # Create story (content)
            story = []
            
            # Summary statistics - moved before the table
            total_hours = sum(entry['duration_seconds'] for entry in entries) / 3600
            total_break_hours = sum(entry.get('break_time', 0) for entry in entries) / 3600
            billable_hours = sum(entry['duration_seconds'] for entry in entries 
                            if entry.get('billable', 'Billable') == 'Billable') / 3600
            non_billable_hours = total_hours - billable_hours
            
            project_hours = {}
            project_billable = {}
            
            for entry in entries:
                project = entry['project']
                hours = entry['duration_seconds'] / 3600
                is_billable = entry.get('billable', 'Billable') == 'Billable'
                
                project_hours[project] = project_hours.get(project, 0) + hours
                
                if project not in project_billable:
                    project_billable[project] = {'billable': 0, 'non_billable': 0}
                
                if is_billable:
                    project_billable[project]['billable'] += hours
                else:
                    project_billable[project]['non_billable'] += hours
            
            # Summary paragraph with billable information
            summary_text = f"""
            <b>Summary Statistics:</b><br/>
            Total Time: {total_hours:.2f} hours | Billable: {billable_hours:.2f} hours | Non-Billable: {non_billable_hours:.2f} hours<br/>
            Break Time: {total_break_hours:.2f} hours | Net Work Time: {total_hours - total_break_hours:.2f} hours<br/>
            Number of Entries: {len(entries)} | Number of Projects: {len(project_hours)}
            """
            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 12))
            
            # Create table data without description column
            table_data = [['Date', 'Project', 'Task', 'Billable', 'Start', 'End', 'Duration', 'Break']]
            
            for entry in entries:
                # Format duration for display
                duration_seconds = entry['duration_seconds']
                hours, remainder = divmod(duration_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                duration_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
                
                # Format break time for display
                break_seconds = entry.get('break_time', 0)
                b_hours, b_remainder = divmod(break_seconds, 3600)
                b_minutes, b_seconds = divmod(b_remainder, 60)
                break_str = f"{int(b_hours):02d}:{int(b_minutes):02d}:{int(b_seconds):02d}"
                
                table_data.append([
                    entry['date'],
                    entry['project'],
                    entry['task'],
                    entry.get('billable', 'Billable'),
                    entry['start_time'],
                    entry['end_time'],
                    duration_str,
                    break_str
                ])
            
            # Create table
            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c6fbb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8f4f8')])
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # Add project summary
            story.append(Paragraph("<b>Project Summary:</b>", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            # Create project summary table
            summary_data = [['Project', 'Billable Hours', 'Non-Billable Hours', 'Total Hours', 'Percentage']]
            
            for project, hours in project_hours.items():
                billable = project_billable[project]['billable']
                non_billable = project_billable[project]['non_billable']
                percentage = (hours / total_hours) * 100 if total_hours > 0 else 0
                
                summary_data.append([
                    project,
                    f"{billable:.2f}",
                    f"{non_billable:.2f}",
                    f"{hours:.2f}",
                    f"{percentage:.1f}%"
                ])
            
            summary_table = Table(summary_data, repeatRows=1)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c6fbb')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e8f4f8')])
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Add chart with different colors for each project
            try:
                # Create a simple bar chart of time by project with different colors
                fig, ax = plt.subplots(figsize=(8, 4))
                
                projects = list(project_hours.keys())
                hours = list(project_hours.values())
                
                # Generate distinct colors for each project
                colors_list = plt.cm.Set3(np.linspace(0, 1, len(projects)))
                
                bars = ax.bar(projects, hours, color=colors_list)
                ax.set_ylabel('Hours')
                ax.set_title('Time Distribution by Project')
                plt.xticks(rotation=45, ha='right')
                
                # Save chart to temporary file
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                chart_path = self.time_reports_dir / f"chart_{timestamp}.png"
                plt.tight_layout()
                plt.savefig(chart_path, dpi=100, bbox_inches='tight')
                plt.close()
                
                # Add chart to PDF
                story.append(Paragraph("Time Distribution by Project", styles['Heading2']))
                story.append(Spacer(1, 12))
                
                if chart_path.exists() and os.access(chart_path, os.R_OK):
                    story.append(Image(str(chart_path), width=6*inch, height=3*inch))
                else:
                    self.log_action(f"Chart file not accessible: {chart_path}")
                    story.append(Paragraph("Chart could not be generated", styles['Normal']))
                    
            except Exception as e:
                self.log_action(f"Error creating chart for PDF: {e}")
                story.append(Paragraph("Chart could not be generated", styles['Normal']))
            
            # Generate QR code for SmartReplaceTool documentation
            try:
                import qrcode
                # Create QR code with link to documentation or GitHub repository
                qr = qrcode.QRCode(
                    version=1,
                    error_correction=qrcode.constants.ERROR_CORRECT_L,
                    box_size=3,
                    border=2,
                )
                qr.add_data("https://github.com/your-username/SmartReplaceTool")  # Replace with actual URL
                qr.make(fit=True)
                
                # Create QR code image
                qr_img = qr.make_image(fill_color="black", back_color="white")
                
                # Save QR code to temporary file
                qr_code_path = self.time_reports_dir / f"qr_code_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
                qr_img.save(str(qr_code_path))
                
            except ImportError:
                self.log_action("QR code library not available. Install 'qrcode[pil]' to add QR codes to reports.")
                qr_code_path = None
            except Exception as e:
                self.log_action(f"Error creating QR code: {e}")
                qr_code_path = None
            
            # Define header function for every page
            def add_header(canvas, doc):
                canvas.saveState()
                
                # Add header background with gradient
                width = doc.pagesize[0]
                height = doc.pagesize[1]
                
                # Create gradient background
                canvas.setFillColor(colors.HexColor('#2c6fbb'))  # Dark blue
                canvas.rect(0, height - 100, width, 100, fill=1, stroke=0)
                
                # Add logo if available
                logo_path = self.tool_dir / "icon.png"
                if logo_path.exists():
                    try:
                        canvas.drawImage(str(logo_path), 72, height - 80, width=40, height=40, mask='auto')
                    except:
                        pass  # Silently fail if logo can't be loaded
                
                # Add header content with better styling
                canvas.setFillColor(colors.white)
                canvas.setFont('Helvetica-Bold', 18)
                canvas.drawCentredString(width / 2, height - 40, title)
                
                canvas.setFont('Helvetica', 10)
                canvas.drawString(120, height - 70, f"Generated by: {self.username}")
                canvas.drawString(120, height - 85, f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                canvas.drawRightString(width - 72, height - 70, "SmartReplaceTool v5.2.6")
                
                # Add decorative elements
                canvas.setStrokeColor(colors.HexColor('#ffffff'))
                canvas.setLineWidth(1)
                canvas.line(72, height - 95, width - 72, height - 95)
                
                canvas.restoreState()
            
            # Define footer function for every page with useful information and QR code
            def add_footer(canvas, doc):
                canvas.saveState()
                canvas.setFont('Helvetica', 8)
                canvas.setFillColor(colors.grey)
                
                # Footer content with useful info
                footer_text = [
                    f"Report generated by {self.username} on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                    f"Total entries: {len(entries)} | Total hours: {total_hours:.2f}",
                    f"Billable hours: {billable_hours:.2f} ({billable_hours/total_hours*100:.1f}%)" if total_hours > 0 else "No time entries",
                    "SmartReplaceTool v5.2.6 - Professional Time Tracking Module"
                ]
                
                # Draw footer text
                y_position = 60  # Increased to make room for QR code
                for text in footer_text:
                    canvas.drawCentredString(doc.pagesize[0] / 2, y_position, text)
                    y_position -= 12
                
                # Add QR code if available
                if qr_code_path and qr_code_path.exists():
                    try:
                        # Draw small QR code in the bottom right corner
                        qr_size = 40  # Size of QR code
                        canvas.drawImage(str(qr_code_path), 
                                    doc.pagesize[0] - 72 - qr_size,  # Right margin
                                    20,  # Bottom margin
                                    width=qr_size, 
                                    height=qr_size,
                                    mask='auto')
                        
                        # Add text below QR code
                        canvas.setFont('Helvetica', 6)
                        canvas.drawCentredString(doc.pagesize[0] - 72 - (qr_size/2), 15, "Scan for info")
                    except Exception as e:
                        self.log_action(f"Error adding QR code to PDF: {e}")
                
                # Add page number
                canvas.drawRightString(doc.pagesize[0] - 72, 30, f"Page {doc.page}")
                
                canvas.restoreState()
            
            # Build the PDF with header and footer on all pages
            def add_header_and_footer(canvas, doc):
                add_header(canvas, doc)
                add_footer(canvas, doc)
            
            # Build the PDF
            doc.build(story, onFirstPage=add_header_and_footer, onLaterPages=add_header_and_footer)
            
            # Make PDF uneditable
            self.lock_pdf_enhanced(temp_file_path, final_file_path)
            
            # Clean up temporary files
            if temp_file_path.exists():
                temp_file_path.unlink()
            
            if chart_path and chart_path.exists():
                try:
                    chart_path.unlink()
                except Exception as e:
                    self.log_action(f"Could not delete chart file: {chart_path} - {e}")
            
            if qr_code_path and qr_code_path.exists():
                try:
                    qr_code_path.unlink()
                except Exception as e:
                    self.log_action(f"Could not delete QR code file: {qr_code_path} - {e}")
            
            self.log_action(f"Generated PDF time report: {final_file_path.name}")
            messagebox.showinfo("Success", f"PDF report generated: {final_file_path.name}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate PDF: {str(e)}")
            self.log_action(f"PDF generation failed: {str(e)}")
            
            # Clean up temporary files if they exist
            for file_path in [temp_file_path, chart_path, qr_code_path]:
                if file_path and file_path.exists():
                    try:
                        file_path.unlink()
                    except:
                        pass

    def lock_pdf_enhanced(self, input_path, output_path):
        """Apply enhanced protection to PDF to prevent editing while allowing printing"""
        try:
            from PyPDF2 import PdfReader, PdfWriter
            
            reader = PdfReader(str(input_path))
            writer = PdfWriter()
            
            # Add all pages to the writer
            for page in reader.pages:
                writer.add_page(page)
            
            # Add metadata
            writer.add_metadata({
                '/Title': 'SmartReplaceTool Time Report',
                '/Author': f'SmartReplaceTool v5.2.6 - {self.username}',
                '/Creator': 'SmartReplaceTool Time Tracking Module',
                '/Producer': 'SmartReplaceTool v5.2.6',
                '/CreationDate': datetime.now().strftime("D:%Y%m%d%H%M%S")
            })
            
            # Encrypt with strong protection - allow printing but nothing else
            writer.encrypt(
                user_password="",  # No password needed to view
                owner_password="protected",  # Password needed to change permissions
                use_128bit=True,
                permissions_flag=0b1111001000  # Allow printing but prevent modifications
            )
            
            # Write the encrypted PDF
            with open(str(output_path), "wb") as f:
                writer.write(f)
                
        except Exception as e:
            self.log_action(f"PDF enhanced locking failed: {e}")
            # If locking fails, just copy the file
            import shutil
            shutil.copy2(str(input_path), str(output_path))

    def export_to_excel(self, entries=None, title=None, filename=None):
        """Export time entries to an Excel spreadsheet"""
        # Handle parameters for both direct call and report generation call
        if entries is None:
            entries = self.time_data.get('entries', [])
        if title is None:
            title = "Time Entries Export"
        if filename is None:
            filename = f"time_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        try:
            file_path = self.time_reports_dir / f"{filename}.xlsx"
            
            # Create a workbook and add worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Time Entries"
            
            # Add title
            ws.merge_cells('A1:I1')
            title_cell = ws['A1']  # Get the cell, not the range
            title_cell.value = title
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add user info
            ws.merge_cells('A2:I2')
            info_cell = ws['A2']
            info_cell.value = f"Generated by: {self.username} on {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            
            # Add headers
            headers = ['Date', 'Project', 'Task', 'Billable', 'Start Time', 'End Time', 'Duration', 'Break Time', 'Description']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color=self.theme["primary_color"].lstrip('#'), 
                                    end_color=self.theme["primary_color"].lstrip('#'), 
                                    fill_type="solid")
            
            # Add data
            for row, entry in enumerate(entries, 4):
                # Format duration for display
                duration_seconds = entry['duration_seconds']
                hours, remainder = divmod(duration_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                duration_str = f"{int(hours):02d}:{int(minutes):02d}:{int(seconds):02d}"
                
                # Format break time for display
                break_seconds = entry.get('break_time', 0)
                b_hours, b_remainder = divmod(break_seconds, 3600)
                b_minutes, b_seconds = divmod(b_remainder, 60)
                break_str = f"{int(b_hours):02d}:{int(b_minutes):02d}:{int(b_seconds):02d}"
                
                ws.cell(row=row, column=1, value=entry['date'])
                ws.cell(row=row, column=2, value=entry['project'])
                ws.cell(row=row, column=3, value=entry['task'])
                ws.cell(row=row, column=4, value=entry.get('billable', 'Billable'))
                ws.cell(row=row, column=5, value=entry['start_time'])
                ws.cell(row=row, column=6, value=entry['end_time'])
                ws.cell(row=row, column=7, value=duration_str)
                ws.cell(row=row, column=8, value=break_str)
                ws.cell(row=row, column=9, value=entry.get('description', ''))
            
            # Format columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary worksheet
            ws_summary = wb.create_sheet("Summary")
            
            # Calculate summary data
            project_hours = {}
            project_billable = {}
            
            for entry in entries:
                project = entry['project']
                hours = entry['duration_seconds'] / 3600
                is_billable = entry.get('billable', 'Billable') == 'Billable'
                
                project_hours[project] = project_hours.get(project, 0) + hours
                
                if project not in project_billable:
                    project_billable[project] = {'billable': 0, 'non_billable': 0}
                
                if is_billable:
                    project_billable[project]['billable'] += hours
                else:
                    project_billable[project]['non_billable'] += hours
            
            # Add summary title
            summary_title_cell = ws_summary['A1']
            summary_title_cell.value = "Time Summary by Project"
            summary_title_cell.font = Font(bold=True, size=14)
            
            # Add summary headers
            ws_summary['A2'] = "Project"
            ws_summary['B2'] = "Billable Hours"
            ws_summary['C2'] = "Non-Billable Hours"
            ws_summary['D2'] = "Total Hours"
            ws_summary['E2'] = "Percentage"
            
            # Style header cells individually
            for col in ['A', 'B', 'C', 'D', 'E']:
                cell = ws_summary[f'{col}2']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.theme["secondary_color"].lstrip('#'), 
                                    end_color=self.theme["secondary_color"].lstrip('#'), 
                                    fill_type="solid")
            
            # Add summary data
            total_hours = sum(project_hours.values())
            for row, (project, hours) in enumerate(project_hours.items(), 3):
                ws_summary.cell(row=row, column=1, value=project)
                ws_summary.cell(row=row, column=2, value=project_billable[project]['billable'])
                ws_summary.cell(row=row, column=3, value=project_billable[project]['non_billable'])
                ws_summary.cell(row=row, column=4, value=hours)
                ws_summary.cell(row=row, column=5, value=f"{(hours/total_hours)*100:.1f}%")
            
            # Format summary worksheet
            for column in ws_summary.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 30)
                ws_summary.column_dimensions[column_letter].width = adjusted_width
            
            # Protect the workbook
            for sheet in wb:
                sheet.protection.sheet = True
                sheet.protection.password = 'protected'
            
            # Save workbook
            wb.save(str(file_path))
            
            self.log_action(f"Generated Excel time report: {file_path.name}")
            messagebox.showinfo("Success", f"Excel report generated: {file_path.name}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate Excel report: {str(e)}")
            self.log_action(f"Excel generation failed: {str(e)}")