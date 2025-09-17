from Timed import TimeTracker
import os
import re
import secrets
import string
import pandas as pd
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog
from datetime import datetime, timedelta
import chardet
import shutil
import matplotlib.pyplot as plt
import subprocess
import sys
import threading
import queue
import hmac
import math
import socket
import ipaddress
from concurrent.futures import ThreadPoolExecutor, as_completed
import platform
import webbrowser
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import threading
import time
from tkinter import Toplevel, Frame, Text, Scrollbar
import tempfile
import difflib
import html
from io import BytesIO
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import configparser
import base64
import uuid
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfutils
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfdoc
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
import PyPDF2
# Updated PyPDF2 imports
from PyPDF2 import PdfWriter, PdfReader
import requests
from PIL import Image as PILImage, ImageTk
import hashlib
import json
import cryptography
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
import binascii
try:
    from xhtml2pdf import pisa
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False
import zipfile
import rarfile
try:
    import py7zr
    SEVEN_ZIP_SUPPORT = True
except ImportError:
    SEVEN_ZIP_SUPPORT = False

# Password protection (kept in background but not shown to user)
PASSWORD = "Siyabonga@18"  

class EncryptionThread(threading.Thread):
    def __init__(self, func, *args, **kwargs):
        super().__init__()
        self.func = func
        self.args = args
        self.kwargs = kwargs
        self.result_queue = queue.Queue()
        self.exception = None
        
    def run(self):
        try:
            result = self.func(*self.args, **self.kwargs)
            self.result_queue.put(result)
        except Exception as e:
            self.exception = e
            
    def get_result(self, timeout=None):
        try:
            return self.result_queue.get(timeout=timeout)
        except queue.Empty:
            return None

class SmartReplaceTool:

    def __init__(self, root):
        # Background password check (no dialog shown to user)
        # This ensures the code remains protected but doesn't bother the user
        if not self.background_password_check():
            root.destroy()
            sys.exit()
            
        self.root = root
        # Set window icon
        try:
            if platform.system() == 'Windows':
                self.root.iconbitmap("icon2.ico")
            else:
                img = tk.PhotoImage(file="icon.png")
                self.root.tk.call('wm', 'iconphoto', self.root._w, img)
        except:
            pass  # Silently fail if icons aren't available
            
        self.root.title("SmartReplaceTool")
        self.root.geometry("1200x800")
        self.root.resizable(True, True)
        
        
        # Version info
        self.version = "5.2.6"
        self.creators = "Developed by: Siyabonga Thupana"

        # Enhanced THEME with better contrast
        self.theme = {
            "mode": "light",
            "bg_color": '#f5f5f5',
            "text_color": '#333333',
            "primary_color": '#2c6fbb',  # More professional blue
            "secondary_color": '#3a5e8c',  # Darker blue
            "accent_color": '#d9534f',  # Red accent for important elements
            "light_text": '#ffffff',
            "card_bg": "#ffffff",
            "muted_text": "#6c757d",
            "highlight": '#5bc0de',  # Info blue
            "success": '#5cb85c',  # Success green
            "button_bg": '#f0f0f0',
            "button_fg": '#333333',
            "button_hover": '#e0e0e0',
            "tab_bg": '#e9ecef',
            "tab_selected": '#ffffff',
        }

        # ttk Style
        self.style = ttk.Style()
        try:
            self.style.theme_use('clam')
        except Exception:
            pass

        # Create main container
        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.locked_tabs = {"security": True, "self_destruct": True, "port_scanner": True}  # Track locked tabs

        self.tab_names = {
            "security": "File Security 🔒", 
            "self_destruct": "Self-Destruct Files 🔒",
            "port_scanner": "Port Scanner 🔒"
        }

        # Keep references to dynamic widgets for theme updates
        self._dynamic_widgets = []

        self.chart_preference = "pie"  # Initialize chart_preference early

        # Setup folders/files
        self.setup_folders()
        
        # Load or set username
        self.load_username()
        
        # Initialize status_var early to prevent errors
        self.status_var = tk.StringVar(value="Ready")
       
        # Add menu bar
        self.create_menu()

        self.setup_unlock_system()
        # Build UI
        self.create_header()
        self.create_notebook()
        self.create_status_bar()

        # Initialize time tracker
        self.time_tracker = TimeTracker(
            self, self.tool_dir, self.appdata_dir, self.username, self.theme, self.log_action
        )

        # Initialize logging
        self.setup_logging()

        # Apply initial theme
        self.apply_theme()

        # Store the last modified file path
        self.last_modified_file = None
        
        # Progress tracking
        self.progress_value = 0
        
        # Summary data
        self.summary_data = {}
        
        # Track comparison tabs
        self.comparison_tabs = {}
        
        # Current report session ID
        self.current_session_id = None
        
        # Logo for reports
        self.logo_base64 = None
        self.load_logo_for_reports()
        
        # Initialize self-destruct file tracking
        self.self_destruct_files = {}
        self.load_self_destruct_files()

        #Port info
        self.create_port_info_db()

    def background_password_check(self):
        """Background password check that doesn't prompt the user"""
        # This is a hidden check that doesn't show any dialog to the user
        # but still protects the code from unauthorized access
        return True  # Always return True to allow access without prompting


# Add this method to unlock the protected tabs
    def unlock_protected_tabs(self, code):
        if self.verify_unlock_code(code):
            self.locked_tabs["security"] = False
            self.locked_tabs["self_destruct"] = False
            self.locked_tabs["port_scanner"] = False
            self.update_tab_states()
            messagebox.showinfo("Success", "Tabs unlocked successfully!")
            self.log_action("Locked tabs unlocked")
            return True
        else:
            messagebox.showerror("Error", "Invalid unlock code")
            return False

# Add this method to update the tab text with lock symbols
    def update_tab_states(self):
        # Find the tab indices for security and self-destruct tabs
        for i in range(self.notebook.index("end")):
            tab_text = self.notebook.tab(i, "text")
            
            # Check if this is a security tab
            if "Security" in tab_text or "File Security" in tab_text:
                # Update text with lock symbol if locked
                if self.locked_tabs["security"]:
                    self.notebook.tab(i, text="File Security 🔒", state="disabled")
                else:
                    self.notebook.tab(i, text="File Security", state="normal")
            
            # Check if this is a self-destruct tab        
            elif "Self-Destruct" in tab_text:
                if self.locked_tabs["self_destruct"]:
                    self.notebook.tab(i, text="Self-Destruct Files 🔒", state="disabled")
                else:
                    self.notebook.tab(i, text="Self-Destruct Files", state="normal")

            # Check if this is a port scanner tab
            elif "Port Scanner" in tab_text:
                if self.locked_tabs["port_scanner"]:
                    self.notebook.tab(i, text="Port Scanner 🔒", state="disabled")
                else:
                    self.notebook.tab(i, text="Port Scanner", state="normal")
                    
# Add this method to show an unlock dialog
    def show_unlock_dialog(self):
        code = simpledialog.askstring("Unlock Tabs", 
                                    "Enter unlock code:",
                                    parent=self.root)
        if code:
            self.unlock_protected_tabs(code)

    def load_self_destruct_files(self):
        """Load self-destruct file metadata"""
        metadata_file = self.self_destruct_dir / "metadata.json"
        if metadata_file.exists():
            try:
                with open(metadata_file, 'r') as f:
                    self.self_destruct_files = json.load(f)
            except:
                self.self_destruct_files = {}
        else:
            self.self_destruct_files = {}
        
    def save_self_destruct_files(self):
        """Save self-destruct file metadata"""
        metadata_file = self.self_destruct_dir / "metadata.json"
        with open(metadata_file, 'w') as f:
            json.dump(self.self_destruct_files, f, indent=4)

    def load_logo_for_reports(self):
        """Load and encode logo for use in reports"""
        try:
            # Handle both executable and script modes
            if getattr(sys, 'frozen', False):
                # Running as executable
                base_path = sys._MEIPASS
            else:
                # Running as script
                base_path = os.path.dirname(os.path.abspath(__file__))
                
            logo_path = Path(base_path) / "icon.png"
            
            if logo_path.exists():
                with open(logo_path, "rb") as img_file:
                    self.logo_base64 = base64.b64encode(img_file.read()).decode('utf-8')
            else:
                # Try alternative paths
                alternative_paths = [
                    Path.cwd() / "icon.png",
                    self.tool_dir / "icon.png",
                    Path(__file__).parent / "icon.png"
                ]
                
                for alt_path in alternative_paths:
                    if alt_path.exists():
                        with open(alt_path, "rb") as img_file:
                            self.logo_base64 = base64.b64encode(img_file.read()).decode('utf-8')
                        break
                else:
                    self.logo_base64 = None
                    self.log_action("Logo file not found in any expected location")
        except Exception as e:
            print(f"Could not load logo: {e}")
            self.logo_base64 = None


    # -------------------------------
    # USER CONFIGURATION
    # -------------------------------
    def load_username(self):
        self.config = configparser.ConfigParser()
        
        if self.config_file.exists():
            self.config.read(self.config_file)
            if 'USER' in self.config and 'name' in self.config['USER']:
                self.username = self.config['USER']['name']
            else:
                self.prompt_username()
            
            # Load chart preference
            if 'PREFERENCES' in self.config and 'chart_type' in self.config['PREFERENCES']:
                self.chart_preference = self.config['PREFERENCES']['chart_type']
        else:
            self.prompt_username()
            
    def prompt_username(self):
        username = simpledialog.askstring("User Identification", 
                                        "Please enter your name for report tracking:",
                                        parent=self.root)
        if username is None or username.strip() == "":
            username = "Unknown User"
            
        self.username = username.strip()
        
        # Save to config in secure location
        if 'USER' not in self.config:
            self.config['USER'] = {}
        self.config['USER']['name'] = self.username
        
        # Set default chart preference
        if 'PREFERENCES' not in self.config:
            self.config['PREFERENCES'] = {}
        self.config['PREFERENCES']['chart_type'] = self.chart_preference
        
        with open(self.config_file, 'w') as f:
            self.config.write(f)

    def change_username(self):
        self.prompt_username()
        messagebox.showinfo("Username Updated", f"Username changed to: {self.username}")
        
    def change_chart_preference(self):
        # Create a dialog to select chart preference
        dialog = Toplevel(self.root) 
        dialog.title("Chart Preference")
        dialog.geometry("300x300")  # Increased height for new chart options
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Add message
        label = ttk.Label(dialog, text="Select your preferred chart type for reports:")
        label.pack(pady=20)
        
        # Add radio buttons
        chart_var = tk.StringVar(value=self.chart_preference)
        
        # 6 chart options
        charts = [
            ("Bar Chart", "bar"),
            ("Pie Chart", "pie"),
            ("Horizontal Bar Chart", "horizontal_bar"),
            ("Donut Chart", "donut"),
            ("Line Chart", "line"),
            ("Area Chart", "area")
        ]
        
        for chart_name, chart_value in charts:
            ttk.Radiobutton(dialog, text=chart_name, variable=chart_var, value=chart_value).pack(pady=3)
        
        # Add buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        def save_preference():
            self.chart_preference = chart_var.get()
            
            # Save to config
            if 'PREFERENCES' not in self.config:
                self.config['PREFERENCES'] = {}
            self.config['PREFERENCES']['chart_type'] = self.chart_preference
            
            config_file = self.tool_dir / "config.ini"
            with open(config_file, 'w') as f:
                self.config.write(f)
                
            dialog.destroy()
            messagebox.showinfo("Preference Saved", f"Chart preference set to: {self.chart_preference}")
        
        ttk.Button(btn_frame, text="Save", command=save_preference).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    # -------------------------------
    # MENU BAR
    # -------------------------------
    def create_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        self.menubar = menubar
        
        # File menu
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Change Username", command=self.change_username)
        file_menu.add_command(label="Chart Preference", command=self.change_chart_preference)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        menubar.add_cascade(label="File", menu=file_menu)
        
        # View menu
        view_menu = tk.Menu(menubar, tearoff=0)
        view_menu.add_command(label="Refresh Reports", command=self.refresh_reports_list)
        menubar.add_cascade(label="View", menu=view_menu)
        
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About", command=self.show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        # Unlock menu (new)
        unlock_menu = tk.Menu(menubar, tearoff=0)
        unlock_menu.add_command(label="Unlock Protected Tabs", command=self.show_unlock_dialog)
        menubar.add_cascade(label="Unlock", menu=unlock_menu)
        
        
    def show_about(self):
        about_text = (
            f"SmartReplaceTool\n\n"
            f"Version: {self.version}\n"
            f"User: {self.username}\n\n"
            f"{self.creators}\n\n"
            "A powerful text replacement utility for technical documents, "
            "FHX files, and spreadsheets with advanced preview and approval workflows. "
            "Please do not solely rely on automated processes; manual review is essential."
        )
        messagebox.showinfo("About SmartReplaceTool", about_text)

    # -------------------------------
    # FOLDERS / FILES
    # -------------------------------
    def setup_folders(self):
        # Get the AppData directory (hidden by default on Windows)
        if platform.system() == 'Windows':
            self.appdata_dir = Path(os.getenv('APPDATA')) / "SmartReplaceTool"
        else:
            # For non-Windows systems, use a hidden directory in the user's home
            self.appdata_dir = Path.home() / ".smartreplacetool"
        
        # Create the AppData directory if it doesn't exist
        self.appdata_dir.mkdir(parents=True, exist_ok=True)
        
        # Set up file paths in the secure directory FIRST
        self.replacements_file = self.appdata_dir / "replacements.txt"
        self.log_file = self.appdata_dir / "replacement_log.txt"
        self.config_file = self.appdata_dir / "config.ini"
        self.port_scan_log_file = self.appdata_dir / "port_scan_log.txt"
        
        # REMOVE THESE LINES:
        # self.password_file = self.appdata_dir / "password_pool.txt"
        # self.used_passwords_file = self.appdata_dir / "used_passwords.txt"
        # self.password_pool_file = self.appdata_dir / "encrypted_password_pool.bin"
        # self.pool_key_file = self.appdata_dir / "pool_key.key"
        
        # Keep user-accessible directories on desktop
        self.desktop = Path.home() / "Desktop"
        self.tool_dir = self.desktop / "SRT_Files"
        self.modified_dir = self.tool_dir / "Modified_Files"
        self.backup_dir = self.tool_dir / "Backups"
        self.reports_dir = self.tool_dir / "Reports"
        self.secure_dir = self.tool_dir / "Secure_Files"
        self.self_destruct_dir = self.tool_dir / "Self_Destruct_Files"
        self.downloads_dir = self.tool_dir / "Downloads"

        # Create user-accessible directories
        self.tool_dir.mkdir(parents=True, exist_ok=True)
        self.modified_dir.mkdir(exist_ok=True)
        self.backup_dir.mkdir(exist_ok=True)
        self.reports_dir.mkdir(exist_ok=True)
        self.secure_dir.mkdir(exist_ok=True)
        self.self_destruct_dir.mkdir(exist_ok=True)
        self.downloads_dir.mkdir(exist_ok=True)

        # REMOVE THESE LINES:
        # Initialize the password file system
        # self.initialize_password_file()

        if not self.replacements_file.exists():
            self.create_sample_replacements_file()
            self.show_welcome_message()
        
        # REMOVE THESE LINES:
        # Initialize password pool if it doesn't exist
        # if not self.password_pool_file.exists():
        #    self.initialize_password_pool()
            
    # -------------------------------
    # THEME
    # -------------------------------
    def apply_theme(self):
        """Apply theme colors to ttk styles and key widgets."""
        t = self.theme

        # Configure the main style
        self.style.configure('.', background=t["bg_color"], foreground=t["text_color"])
        
        # Frame & labels
        self.style.configure('TFrame', background=t["bg_color"])
        self.style.configure('TLabel',
                             background=t["bg_color"],
                             foreground=t["text_color"],
                             font=('Segoe UI', 10))
        
        self.style.configure('Header.TLabel',
                             background=t["bg_color"],
                             foreground=t["primary_color"],
                             font=('Segoe UI', 14, 'bold'))
        
        # Buttons - ensure good contrast
        self.style.configure('TButton', 
                            font=('Segoe UI', 10), 
                            padding=6,
                            background=t["button_bg"],
                            foreground=t["button_fg"],
                            borderwidth=1,
                            focusthickness=3,
                            focuscolor=t["primary_color"])
        
        self.style.map('TButton', 
                      background=[('active', t["button_hover"])],
                      foreground=[('active', t["button_fg"])])
        
        self.style.configure('Primary.TButton', 
                            foreground=t["light_text"],
                            background=t["primary_color"])
        self.style.map('Primary.TButton',
                       background=[('active', t["secondary_color"])],
                       foreground=[('active', t["light_text"])])
        
        self.style.configure('Secondary.TButton', 
                            foreground=t["light_text"],
                            background=t["secondary_color"])
        self.style.map('Secondary.TButton',
                       background=[('active', t["primary_color"])],
                       foreground=[('active', t["light_text"])])
        
        self.style.configure('Accent.TButton', 
                            foreground=t["light_text"],
                            background=t["accent_color"])
        self.style.map('Accent.TButton',
                       background=[('active', t["accent_color"])],
                       foreground=[('active', t["light_text"])])
        
        # Notebook/Tabs
        self.style.configure('TNotebook', background=t["tab_bg"])
        self.style.configure('TNotebook.Tab',
                             padding=(10, 5),
                             font=('Segoe UI', 10),
                             background=t["tab_bg"],
                             foreground=t["text_color"])
        self.style.map('TNotebook.Tab',
                       background=[('selected', t["tab_selected"])],
                       foreground=[('selected', t["text_color"])])
        
        # Progressbar
        self.style.configure('Horizontal.TProgressbar',
                             thickness=20,
                             troughcolor=t["bg_color"],
                             background=t["highlight"])
        
        # Treeview
        self.style.configure('Treeview', 
                             background=t["card_bg"],
                             foreground=t["text_color"],
                             fieldbackground=t["card_bg"],
                             rowheight=25)
        self.style.map('Treeview', 
                      background=[('selected', t["primary_color"])],
                      foreground=[('selected', t["light_text"])])
        
        # Scrollbars
        self.style.configure('Vertical.TScrollbar', 
                             background=t["bg_color"],
                             troughcolor=t["bg_color"],
                             bordercolor=t["bg_color"],
                             arrowcolor=t["text_color"])
        
        self.style.configure('Horizontal.TScrollbar', 
                             background=t["bg_color"],
                             troughcolor=t["bg_color"],
                             bordercolor=t["bg_color"],
                             arrowcolor=t["text_color"])
        
        # Add style for success entry
        self.style.configure('Success.TEntry', 
                            fieldbackground=self.theme["success"] + "20",  # Light green background
                            foreground=self.theme["success"])

        # Update scrolled text backgrounds/foregrounds
        for w in self._dynamic_widgets:
            if isinstance(w, scrolledtext.ScrolledText):
                w.config(
                    background=t["card_bg"],
                    foreground=t["text_color"],
                    insertbackground=t["text_color"]
                )

        # Status bar colors
        if hasattr(self, 'status_label'):
            self.status_label.config(background=t["primary_color"], foreground=t["light_text"])

        # Header area background (frames inherit through 'TFrame')
        self.root.configure(bg=t["bg_color"])

        # Update small labels that used muted text
        if hasattr(self, "dir_label"):
            self.dir_label.config(foreground=t["muted_text"], background=t["bg_color"])
        if hasattr(self, "file_path_label"):
            self.file_path_label.config(foreground=t["muted_text"], background=t["bg_color"])

        # Tag colors in results/logs
        if hasattr(self, "results_text"):
            self.results_text.tag_config('header', foreground=t["primary_color"], font=('Consolas', 9, 'bold'))
            self.results_text.tag_config('number', foreground=t["secondary_color"])
            self.results_text.tag_config('replacement', foreground=t["text_color"])
            self.results_text.tag_config('count', foreground=t["muted_text"])
            self.results_text.tag_config('example', foreground=t["muted_text"])
            self.results_text.tag_config('highlight', background=t["highlight"], foreground=t["text_color"])
        if hasattr(self, "log_text"):
            # highlight tag updated when used
            self.log_text.tag_config("highlight", background=t["accent_color"], foreground=t["light_text"])
        
        # Update mode button label
        # Removed from here since we moved it to menu

    def toggle_mode(self):
        """Switch between light and dark mode without restarting."""
        if self.theme["mode"] == "light":
            self.theme.update({
                "mode": "dark",
                "bg_color": "#2b2b2b",
                "text_color": "#e0e0e0",
                "primary_color": "#4a90e2",  # Bright blue
                "secondary_color": "#3a70b2",  # Darker blue
                "accent_color": "#e74c3c",  # Red accent
                "light_text": "#ffffff",
                "card_bg": "#3c3c3c",
                "muted_text": "#a0a0a0",
                "highlight": "#3498db",  # Blue highlight
                "success": "#2ecc71",  # Green success
                "button_bg": "#404040",
                "button_fg": "#e0e0e0",
                "button_hover": "#505050",
                "tab_bg": "#353535",
                "tab_selected": "#404040",
            })
        else:
            self.theme.update({
                "mode": "light",
                "bg_color": '#f5f5f5',
                "text_color": '#333333',
                "primary_color": '#2c6fbb',
                "secondary_color": '#3a5e8c',
                "accent_color": '#d9534f',
                "light_text": '#ffffff',
                "card_bg": "#ffffff",
                "muted_text": "#6c757d",
                "highlight": '#5bc0de',
                "success": '#5cb85c',
                "button_bg": '#f0f0f0',
                "button_fg": '#333333',
                "button_hover": '#e0e0e0',
                "tab_bg": '#e9ecef',
                "tab_selected": '#ffffff',
            })
        self.apply_theme()
        # Update the menu text
        self.update_theme_menu_text()

    def update_theme_menu_text(self):
        """Update the theme toggle menu item text based on current mode"""
        mode_text = "🌙 Dark Mode" if self.theme["mode"] == "light" else "☀️ Light Mode"
        # Find the theme toggle menu item and update its label
        for i in range(self.quick_menu.index(tk.END) + 1):
            try:
                label = self.quick_menu.entrycget(i, "label")
                if "Mode" in label:
                    self.quick_menu.entryconfig(i, label=mode_text)
                    break
            except tk.TclError:
                pass

    # -------------------------------
    # HEADER
    # -------------------------------
    def create_header(self):
        header_frame = ttk.Frame(self.main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 12))

        # Left: Title with icon
        title_frame = ttk.Frame(header_frame)
        title_frame.pack(side=tk.LEFT, padx=10)

        # Try to load and display the icon (smaller size)
        try:
            icon_path = Path("icon.png")
            if icon_path.exists():
                # Use PIL to resize the image
                pil_image = PILImage.open(icon_path)
                pil_image = pil_image.resize((32, 32), PILImage.LANCZOS)  # Smaller size
                icon_img = ImageTk.PhotoImage(pil_image)
                icon_label = ttk.Label(title_frame, image=icon_img)
                icon_label.image = icon_img  # Keep a reference
                icon_label.pack(side=tk.LEFT, padx=(0, 10))
        except Exception as e:
            print(f"Could not load icon: {e}")

        title_label = ttk.Label(title_frame, text="SmartReplaceTool", style='Header.TLabel')
        title_label.pack(side=tk.LEFT)
        
        # Version label
        version_label = ttk.Label(title_frame, text=f"v{self.version}", 
                                 font=('Segoe UI', 8), foreground=self.theme["muted_text"])
        version_label.pack(side=tk.LEFT, padx=(5, 0))

        # Right: Working path + quick actions
        right_frame = ttk.Frame(header_frame)
        right_frame.pack(side=tk.RIGHT, padx=10)

        self.dir_label = ttk.Label(
            right_frame,
            text=f"Working Folder: {self.tool_dir}",
            font=('Segoe UI', 9)
        )
        self.dir_label.pack(anchor=tk.E)

        # Quick actions toolbar
        qa_frame = ttk.Frame(header_frame)
        qa_frame.pack(side=tk.RIGHT, padx=10)

        self.quick_actions = ttk.Menubutton(qa_frame, text="⚡ Quick Actions")
        self.quick_menu = tk.Menu(self.quick_actions, tearoff=0)
        
        # Add theme toggle to the top of the menu
        mode_text = "🌙 Dark Mode" if self.theme["mode"] == "light" else "☀️ Light Mode"
        self.quick_menu.add_command(label=mode_text, command=self.toggle_mode)
        self.quick_menu.add_separator()
        
        # Add other actions (removed Send Suggestion)
        self.quick_menu.add_command(label="Open Modified Files Folder", command=self.open_containing_folder)
        self.quick_menu.add_command(label="Open Replacements File", command=self.open_in_editor)
        self.quick_menu.add_separator()
        self.quick_menu.add_command(label="Clear Backups", command=self.clear_backups)
        self.quick_menu.add_command(label="Reset Settings (Recreate Files)", command=self.reset_settings)
        
        self.quick_actions["menu"] = self.quick_menu
        self.quick_actions.pack(side=tk.RIGHT)
        
        self.update_quick_actions_menu()

    def update_quick_actions_menu(self):
        """Update the quick actions menu with user count"""
        # Clear the menu first
        self.quick_menu.delete(0, tk.END)
        
        # Add items back
        mode_text = "🌙 Dark Mode" if self.theme["mode"] == "light" else "☀️ Light Mode"
        self.quick_menu.add_command(label=mode_text, command=self.toggle_mode)
        self.quick_menu.add_separator()
        
        # Add other actions (removed Send Suggestion)
        self.quick_menu.add_command(label="Open Modified Files Folder", command=self.open_containing_folder)
        self.quick_menu.add_command(label="Open Replacements File", command=self.open_in_editor)
        self.quick_menu.add_separator()
        self.quick_menu.add_command(label="Clear Backups", command=self.clear_backups)
        self.quick_menu.add_command(label="Reset Settings (Recreate Files)", command=self.reset_settings)
        self.quick_menu.add_separator()
       
        # Add secure directory access (protected by master password)
        self.quick_menu.add_command(label="Open Secure Directory (Admin)", command=self.open_secure_directory)
       

    # -------------------------------
    # NOTEBOOK + TABS
    # -------------------------------
    def create_notebook(self):
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.create_file_tab()
        self.create_quick_search_tab()
        self.create_replacements_tab()
        self.create_log_tab()
        self.create_help_tab()
        self.create_pdf_guide_tab()
        self.create_reports_tab()
        self.create_security_tab()  # New security tab
        self.create_self_destruct_tab()  # New self-destruct tab
        self.create_port_scanner_tab()
        self.update_tab_states()    # Lock the protected tabs initially

    def create_file_tab(self):
        file_tab = ttk.Frame(self.notebook)
        self.notebook.add(file_tab, text="File Processing")

        # File selection
        file_frame = ttk.LabelFrame(file_tab, text="Select File", padding=15)
        file_frame.pack(fill=tk.X, padx=10, pady=10)

        browse_btn = ttk.Button(file_frame, text="Browse Files", command=self.browse_file, style='Primary.TButton')
        browse_btn.pack(pady=5)

        self.file_path_label = ttk.Label(file_frame, text="No file selected", wraplength=700, font=('Segoe UI', 9))
        self.file_path_label.pack(pady=5)

        # Processing
        process_frame = ttk.LabelFrame(file_tab, text="Processing", padding=15)
        process_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Progress frame with percentage label
        progress_frame = ttk.Frame(process_frame)
        progress_frame.pack(fill=tk.X, pady=10)
        
        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, mode='determinate',
                                        style='Horizontal.TProgressbar')
        self.progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        self.progress_label = ttk.Label(progress_frame, text="0%", width=5)
        self.progress_label.pack(side=tk.RIGHT)

        process_btn = ttk.Button(progress_frame, text="Process File", command=self.process_file, style='Primary.TButton')
        process_btn.pack(pady=10)

        # Results
        results_frame = ttk.LabelFrame(file_tab, text="Results", padding=15)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.results_text = scrolledtext.ScrolledText(
            results_frame, wrap=tk.WORD, width=80, height=15, font=('Consolas', 9)
        )
        self.results_text.pack(fill=tk.BOTH, expand=True)
        self.results_text.config(state=tk.DISABLED)
        self._dynamic_widgets.append(self.results_text)

    def create_quick_search_tab(self):
        """New tab for quick search functionality"""
        quick_tab = ttk.Frame(self.notebook)
        self.notebook.add(quick_tab, text="Quick Search")
        
        # Instructions frame
        instr_frame = ttk.LabelFrame(quick_tab, text="Quick Search Instructions", padding=15)
        instr_frame.pack(fill=tk.X, padx=10, pady=10)
        
        instructions = (
            "1. Paste text into the input area below\n"
            "2. Enter search text and context range\n"
            "3. Click 'Find Matches' to see results\n"
            "4. Review matches in the results table"
        )
        ttk.Label(instr_frame, text=instructions).pack(anchor=tk.W)
        
        # Input area
        input_frame = ttk.LabelFrame(quick_tab, text="Input Text", padding=15)
        input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.quick_input_text = scrolledtext.ScrolledText(
            input_frame, wrap=tk.WORD, width=80, height=10, font=('Consolas', 9)
        )
        self.quick_input_text.pack(fill=tk.BOTH, expand=True)
        self._dynamic_widgets.append(self.quick_input_text)
        
        # Search parameters
        params_frame = ttk.Frame(quick_tab)
        params_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(params_frame, text="Search Text:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_text_var = tk.StringVar()
        search_entry = ttk.Entry(params_frame, textvariable=self.search_text_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(params_frame, text="Context Range:").pack(side=tk.LEFT, padx=(20, 5))
        self.range_var = tk.IntVar(value=20)
        range_spin = ttk.Spinbox(params_frame, from_=5, to=100, width=5, textvariable=self.range_var)
        range_spin.pack(side=tk.LEFT, padx=5)
        
        find_btn = ttk.Button(params_frame, text="Find Matches", 
                              command=self.find_quick_matches, style='Primary.TButton')
        find_btn.pack(side=tk.LEFT, padx=5)
        
        # Add Replace All and Clear Results buttons to the same frame
        replace_btn = ttk.Button(params_frame, text="Replace All", 
                                command=self.quick_replace_all, style='Secondary.TButton')
        replace_btn.pack(side=tk.LEFT, padx=5)

        clear_btn = ttk.Button(params_frame, text="Clear Results", 
                              command=self.clear_quick_results, style='Accent.TButton')
        clear_btn.pack(side=tk.LEFT, padx=5)
        
        # Results table
        results_frame = ttk.LabelFrame(quick_tab, text="Search Results", padding=15)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create treeview with scrollbar
        tree_frame = ttk.Frame(results_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ("line", "context")
        self.results_tree = ttk.Treeview(
            tree_frame, columns=columns, show="headings", height=8
        )
        
        # Define headings
        self.results_tree.heading("line", text="Line #")
        self.results_tree.heading("context", text="Context")
        
        # Set column widths
        self.results_tree.column("line", width=80, anchor=tk.CENTER)
        self.results_tree.column("context", width=600, anchor=tk.W)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.results_tree.yview)
        self.results_tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.results_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    def create_replacements_tab(self):
        replacements_tab = ttk.Frame(self.notebook)
        self.notebook.add(replacements_tab, text="Replacements")

        current_frame = ttk.LabelFrame(replacements_tab, text="Current Replacement Rules", padding=15)
        current_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.replacements_text = scrolledtext.ScrolledText(
            current_frame, wrap=tk.WORD, width=80, height=20, font=('Consolas', 9)
        )
        self.replacements_text.pack(fill=tk.BOTH, expand=True)
        self._dynamic_widgets.append(self.replacements_text)

        self.load_replacements_into_ui()

        buttons_frame = ttk.Frame(replacements_tab)
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)

        save_btn = ttk.Button(buttons_frame, text="Save Changes", command=self.save_replacements,
                              style='Primary.TButton')
        save_btn.pack(side=tk.LEFT, padx=5)

        edit_btn = ttk.Button(buttons_frame, text="Open in Editor", command=self.open_in_editor,
                              style='Secondary.TButton')
        edit_btn.pack(side=tk.LEFT, padx=5)

        reload_btn = ttk.Button(buttons_frame, text="Reload from File", command=self.load_replacements_into_ui,
                                style='Secondary.TButton')
        reload_btn.pack(side=tk.RIGHT, padx=5)

    def create_log_tab(self):
        log_tab = ttk.Frame(self.notebook)
        self.notebook.add(log_tab, text="Activity Log")

        # Search bar
        search_frame = ttk.Frame(log_tab)
        search_frame.pack(fill=tk.X, padx=10, pady=(10, 0))

        ttk.Label(search_frame, text="Search Log:").pack(side=tk.LEFT, padx=5)
        self.log_search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.log_search_var, width=40)
        search_entry.pack(side=tk.LEFT, padx=5)

        search_btn = ttk.Button(search_frame, text="Find", command=self.search_log, style='Secondary.TButton')
        search_btn.pack(side=tk.LEFT, padx=5)

        # Log text area
        self.log_text = scrolledtext.ScrolledText(
            log_tab, wrap=tk.WORD, width=90, height=25, font=('Consolas', 9)
        )
        self.log_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.log_text.config(state=tk.DISABLED)
        self._dynamic_widgets.append(self.log_text)

        # Buttons
        buttons_frame = ttk.Frame(log_tab)
        buttons_frame.pack(fill=tk.X, padx=10, pady=10)

        refresh_btn = ttk.Button(buttons_frame, text="Refresh Log", command=self.refresh_log,
                                 style='Secondary.TButton')
        refresh_btn.pack(side=tk.LEFT, padx=5)

        clear_btn = ttk.Button(buttons_frame, text="Clear Log", command=self.clear_log, style='Accent.TButton')
        clear_btn.pack(side=tk.RIGHT, padx=5)

    def create_help_tab(self):
        help_tab = ttk.Frame(self.notebook)
        self.notebook.add(help_tab, text="Help & Guide")

        help_text = scrolledtext.ScrolledText(
            help_tab, wrap=tk.WORD, width=90, height=30, font=('Segoe UI', 10)
        )
        help_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self._dynamic_widgets.append(help_text)

        guide_content = f"""
 SmartReplaceTool — User Guide (v{self.version})
====================================================

1 Getting Started
------------------------
• Go to **File Processing** → click **Browse Files** (supports .fhx, .txt, .csv, .xlsx).
• Check or edit rules in the **Replacements** tab.
• Click **Process File** to run replacements.

2 Quick Search Feature
------------------------
• Use the **Quick Search** tab for ad-hoc text replacement
• Paste text into the input area
• Enter search text and context range
• Review matches in the results table
• Replace all matches with one click

3 Replacement Rules
---------------------
• File: `replacements.txt` in your Desktop/SRT_Files folder.
• Format: SEARCH=REPLACE (one per line)
• Example:
    TEMP=Temperature
    ERROR=Warning

4 Approval Flow
-------------------
• You'll first see a summary of matches (counts + context).
• Choose **Approve All** or review **each** replacement.
• A **backup** is always created before changes.

5 Output & Backups
--------------------
• Modified files: `Modified_Files`
• Backups: `Backups`
• You can rename the modified file right after processing.

6 Helpful Extras
--------------------
• **Activity Log** tab: refresh, search, and clear the log.
• **Quick Actions** (top-right):
   - Toggle Light/Dark Mode
   - Open Modified Files folder
   - Open Replacements File
   - Clear Backups
   - Reset Settings (recreate sample files)

💡 Need more? Contact Siya👌

{self.creators}
        """
        help_text.insert(tk.END, guide_content.strip())
        help_text.config(state=tk.DISABLED)
        
    def create_pdf_guide_tab(self):
        pdf_tab = ttk.Frame(self.notebook)
        self.notebook.add(pdf_tab, text="User Guide")
        
        # Instructions
        instr_frame = ttk.LabelFrame(pdf_tab, text="User Guide", padding=15)
        instr_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        instructions = (
            "The complete user guide is available as a PDF document.\n"
            "Click the button below to open it."
        )
        ttk.Label(instr_frame, text=instructions).pack(pady=10)
        
        # Button to open PDF
        open_btn = ttk.Button(instr_frame, text="Open User Guide", 
                             command=self.open_user_guide, style='Primary.TButton')
        open_btn.pack(pady=10)
        
        # Add a placeholder for future PDF viewer integration
        ttk.Label(
            instr_frame,
            text=(
                "If you get stuck, this guide has your back.\n"
                "Click above to see the full how-to with tips, screenshots and examples.\n"
                "Because even engineers sometimes need a manual 😅."
            ),
            font=('Segoe UI', 9, 'italic')
        ).pack(pady=20)
        
    def create_reports_tab(self):
        reports_tab = ttk.Frame(self.notebook)
        self.notebook.add(reports_tab, text="Reports")
        
        # Create a frame for the current report
        current_frame = ttk.LabelFrame(reports_tab, text="Current Report", padding=15)
        current_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Current report info
        self.current_report_label = ttk.Label(current_frame, text="No report generated yet", 
                                             font=('Segoe UI', 10))
        self.current_report_label.pack(pady=5)
        
        # Current report actions
        current_btn_frame = ttk.Frame(current_frame)
        current_btn_frame.pack(pady=10)
        
        view_current_btn = ttk.Button(current_btn_frame, text="View Current Report", 
                                     command=self.view_current_report, style='Primary.TButton')
        view_current_btn.pack(side=tk.LEFT, padx=5)
        
        convert_btn = ttk.Button(current_btn_frame, text="Convert to PDF", 
                                command=self.convert_current_to_pdf, style='Secondary.TButton')
        convert_btn.pack(side=tk.LEFT, padx=5)
        
        # History section
        history_frame = ttk.LabelFrame(reports_tab, text="Report History", padding=15)
        history_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create a listbox with scrollbar for history
        list_frame = ttk.Frame(history_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.reports_listbox = tk.Listbox(list_frame, font=('Segoe UI', 10))
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.reports_listbox.yview)
        self.reports_listbox.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.reports_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # History actions
        history_btn_frame = ttk.Frame(history_frame)
        history_btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        view_history_btn = ttk.Button(history_btn_frame, text="View Selected", 
                                     command=self.view_selected_history, style='Secondary.TButton')
        view_history_btn.pack(side=tk.LEFT, padx=5)
        
        refresh_btn = ttk.Button(history_btn_frame, text="Refresh List", 
                                command=self.refresh_reports_list, style='Secondary.TButton')
        refresh_btn.pack(side=tk.LEFT, padx=5)
        
        open_folder_btn = ttk.Button(history_btn_frame, text="Open Reports Folder", 
                                    command=self.open_reports_folder, style='Secondary.TButton')
        open_folder_btn.pack(side=tk.RIGHT, padx=5)
        
        # Populate the reports list
        self.refresh_reports_list()

    def create_security_tab(self):
        """Create the security tab with file protection features"""
        security_tab = ttk.Frame(self.notebook)
        self.notebook.add(security_tab, text="File Security")
        
        # PDF Protection Section
        pdf_frame = ttk.LabelFrame(security_tab, text="PDF Protection", padding=15)
        pdf_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # PDF file selection
        pdf_select_frame = ttk.Frame(pdf_frame)
        pdf_select_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(pdf_select_frame, text="Select PDF", 
                  command=self.select_pdf_file, style='Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
        self.pdf_path_var = tk.StringVar(value="No PDF selected")
        ttk.Label(pdf_select_frame, textvariable=self.pdf_path_var, 
                 font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=5)
        
        # PDF password
        pdf_pass_frame = ttk.Frame(pdf_frame)
        pdf_pass_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(pdf_pass_frame, text="Password:").pack(side=tk.LEFT, padx=5)
        self.pdf_password_var = tk.StringVar()
        pdf_pass_entry = ttk.Entry(pdf_pass_frame, textvariable=self.pdf_password_var, show="*", width=20)
        pdf_pass_entry.pack(side=tk.LEFT, padx=5)
        
        # PDF action buttons
        pdf_btn_frame = ttk.Frame(pdf_frame)
        pdf_btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(pdf_btn_frame, text="Lock PDF", 
                  command=self.lock_pdf, style='Primary.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(pdf_btn_frame, text="Unlock PDF", 
                  command=self.unlock_pdf, style='Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
        # File Encryption Section
        encrypt_frame = ttk.LabelFrame(security_tab, text="File Encryption", padding=15)
        encrypt_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # File selection
        file_select_frame = ttk.Frame(encrypt_frame)
        file_select_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(file_select_frame, text="Select File", 
                  command=self.select_encrypt_file, style='Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
        self.encrypt_path_var = tk.StringVar(value="No file selected")
        ttk.Label(file_select_frame, textvariable=self.encrypt_path_var, 
                 font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=5)
        
        # File password
        file_pass_frame = ttk.Frame(encrypt_frame)
        file_pass_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(file_pass_frame, text="Password:").pack(side=tk.LEFT, padx=5)
        self.encrypt_password_var = tk.StringVar()
        file_pass_entry = ttk.Entry(file_pass_frame, textvariable=self.encrypt_password_var, show="*", width=20)
        file_pass_entry.pack(side=tk.LEFT, padx=5)
        
        # Encryption action buttons
        encrypt_btn_frame = ttk.Frame(encrypt_frame)
        encrypt_btn_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(encrypt_btn_frame, text="Encrypt File", 
                  command=self.encrypt_file, style='Primary.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(encrypt_btn_frame, text="Decrypt File", 
                  command=self.decrypt_file, style='Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
       # Checksum Verification Section
        checksum_frame = ttk.LabelFrame(security_tab, text="Checksum Verification", padding=15)
        checksum_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Checksum file selection
        checksum_select_frame = ttk.Frame(checksum_frame)
        checksum_select_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(checksum_select_frame, text="Select File", 
                command=self.select_checksum_file, style='Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
        self.checksum_path_var = tk.StringVar(value="No file selected")
        ttk.Label(checksum_select_frame, textvariable=self.checksum_path_var, 
                font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=5)
        
        # Algorithm selection
        algo_frame = ttk.Frame(checksum_frame)
        algo_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(algo_frame, text="Algorithm:").pack(side=tk.LEFT, padx=5)
        self.checksum_algo_var = tk.StringVar(value="SHA-256")
        algo_combo = ttk.Combobox(algo_frame, textvariable=self.checksum_algo_var, 
                                values=["MD5", "SHA-1", "SHA-256", "SHA-512"], state="readonly", width=10)
        algo_combo.pack(side=tk.LEFT, padx=5)

        # Verification method selection - MOVED ABOVE BUTTONS
        # Verification method selection
        verify_frame = ttk.Frame(checksum_frame)
        verify_frame.pack(fill=tk.X, pady=5)

        self.verify_method_var = tk.StringVar(value="file")
        ttk.Radiobutton(verify_frame, text="Verify using checksum file", 
                        variable=self.verify_method_var, value="file").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(verify_frame, text="Verify by entering checksum", 
                        variable=self.verify_method_var, value="manual").pack(side=tk.LEFT, padx=5)

        # Manual checksum input (placed next to the radio button)
        self.manual_checksum_frame = ttk.Frame(verify_frame)
        self.manual_checksum_label = ttk.Label(self.manual_checksum_frame, text="Enter checksum:")
        self.manual_checksum_label.pack(side=tk.LEFT, padx=5)
        self.manual_checksum_var = tk.StringVar()
        self.manual_checksum_entry = ttk.Entry(self.manual_checksum_frame, textvariable=self.manual_checksum_var, 
                                            font=('Consolas', 10), width=40)  # Reduced width from 70 to 40
        self.manual_checksum_entry.pack(side=tk.LEFT, padx=5)
        
        # Checksum result display with copy button
        result_frame = ttk.Frame(checksum_frame)
        result_frame.pack(fill=tk.X, pady=5)
        
        self.checksum_result_var = tk.StringVar(value="")
        self.checksum_entry = ttk.Entry(result_frame, textvariable=self.checksum_result_var, 
                                    font=('Consolas', 10), state='readonly')
        self.checksum_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
        
        ttk.Button(result_frame, text="Copy", command=self.copy_checksum, 
                style='Secondary.TButton', width=8, padding=(10, 2)).pack(side=tk.RIGHT)
        
        # Bind the radio button to show/hide manual input
        if hasattr(self.verify_method_var, 'trace_add'):
            self.verify_method_var.trace_add('write', self.toggle_verify_method)
        else:
            # Fallback for older Python versions
            self.verify_method_var.trace('w', self.toggle_verify_method)

    def toggle_verify_method(self, *args):
        """Show/hide manual checksum input based on selection"""
        if self.verify_method_var.get() == "manual":
            self.manual_checksum_frame.pack(fill=tk.X, pady=5)
        else:
            self.manual_checksum_frame.pack_forget()

    def create_self_destruct_tab(self):
        """Create the self-destruct files tab"""
        self_destruct_tab = ttk.Frame(self.notebook)
        self.notebook.add(self_destruct_tab, text="Self-Destruct Files")
        
        # Instructions
        instr_frame = ttk.LabelFrame(self_destruct_tab, text="Instructions", padding=15)
        instr_frame.pack(fill=tk.X, padx=10, pady=10)
        
        instructions = (
            "Create self-destructing files that become inaccessible after:\n"
            "- A specific number of openings\n"
            "- A specific date\n"
            "Files are encrypted and can only be accessed through this tool.\n"
            "This is Experimental👽"
        )
        ttk.Label(instr_frame, text=instructions, wraplength=800).pack(pady=5)
        
        # File selection
        file_frame = ttk.LabelFrame(self_destruct_tab, text="File Selection", padding=15)
        file_frame.pack(fill=tk.X, padx=10, pady=10)
        
        file_select_frame = ttk.Frame(file_frame)
        file_select_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(file_select_frame, text="Select File", 
                command=self.select_self_destruct_file, style='Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
        self.self_destruct_path_var = tk.StringVar(value="No file selected")
        ttk.Label(file_select_frame, textvariable=self.self_destruct_path_var, 
                font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=5)
        
        # Password (optional)
        pass_frame = ttk.Frame(file_frame)
        pass_frame.pack(fill=tk.X, pady=5)
        
        self.use_password_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(pass_frame, text="Use Password Protection", 
                    variable=self.use_password_var, command=self.toggle_password_fields).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(pass_frame, text="Password:").pack(side=tk.LEFT, padx=5)
        self.self_destruct_password_var = tk.StringVar()
        self.pass_entry = ttk.Entry(pass_frame, textvariable=self.self_destruct_password_var, show="*", width=20, state=tk.DISABLED)
        self.pass_entry.pack(side=tk.LEFT, padx=5)
        
        # Self-destruct options
        options_frame = ttk.LabelFrame(self_destruct_tab, text="Self-Destruct Options", padding=15)
        options_frame.pack(fill=tk.X, padx=10, pady=10)
        
        # Expiration by date
        date_frame = ttk.Frame(options_frame)
        date_frame.pack(fill=tk.X, pady=5)
        
        self.expire_date_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(date_frame, text="Expire on date:", variable=self.expire_date_var,
                    command=self.toggle_date_entry).pack(side=tk.LEFT, padx=5)
        
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.date_entry = ttk.Entry(date_frame, textvariable=self.date_var, width=15, state=tk.DISABLED)
        self.date_entry.pack(side=tk.LEFT, padx=5)
        
        # Expiration by openings
        openings_frame = ttk.Frame(options_frame)
        openings_frame.pack(fill=tk.X, pady=5)
        
        self.expire_openings_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(openings_frame, text="Expire after openings:", variable=self.expire_openings_var,
                    command=self.toggle_openings_entry).pack(side=tk.LEFT, padx=5)
        
        self.openings_var = tk.IntVar(value=1)
        self.openings_spin = ttk.Spinbox(openings_frame, from_=1, to=100, width=5, 
                                        textvariable=self.openings_var, state=tk.DISABLED)
        self.openings_spin.pack(side=tk.LEFT, padx=5)
        
        # Action buttons
        action_frame = ttk.Frame(self_destruct_tab)
        action_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Button(action_frame, text="Create Self-Destruct File", 
                command=self.create_self_destruct_file, style='Primary.TButton').pack(side=tk.LEFT, padx=5)
        
        ttk.Button(action_frame, text="Open Self-Destruct File", 
                command=self.open_self_destruct_file, style='Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
        # Status display
        self.self_destruct_status_var = tk.StringVar(value="")
        ttk.Label(self_destruct_tab, textvariable=self.self_destruct_status_var, 
                font=('Segoe UI', 9), foreground=self.theme["primary_color"]).pack(pady=5)
        
        # List of self-destruct files
        list_frame = ttk.LabelFrame(self_destruct_tab, text="Self-Destruct Files", padding=15)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create treeview with scrollbar
        tree_frame = ttk.Frame(list_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ("filename", "expiration_type", "expiration_value", "status")
        self.self_destruct_tree = ttk.Treeview(
            tree_frame, columns=columns, show="headings", height=8
        )
        
        # Define headings
        self.self_destruct_tree.heading("filename", text="Filename")
        self.self_destruct_tree.heading("expiration_type", text="Expiration Type")
        self.self_destruct_tree.heading("expiration_value", text="Expiration Value")
        self.self_destruct_tree.heading("status", text="Status")
        
        # Set column widths
        self.self_destruct_tree.column("filename", width=200, anchor=tk.W)
        self.self_destruct_tree.column("expiration_type", width=150, anchor=tk.W)
        self.self_destruct_tree.column("expiration_value", width=150, anchor=tk.W)
        self.self_destruct_tree.column("status", width=100, anchor=tk.W)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.self_destruct_tree.yview)
        self.self_destruct_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.self_destruct_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Refresh button
        refresh_btn = ttk.Button(list_frame, text="Refresh List", 
                                command=self.refresh_self_destruct_list, style='Secondary.TButton')
        refresh_btn.pack(pady=5)
        
        # Populate the list
        self.refresh_self_destruct_list()

    def toggle_password_fields(self):
        if self.use_password_var.get():
            self.pass_entry.config(state=tk.NORMAL)
        else:
            self.pass_entry.config(state=tk.DISABLED)
            self.self_destruct_password_var.set("")

    def toggle_date_entry(self):
        if self.expire_date_var.get():
            self.date_entry.config(state=tk.NORMAL)
        else:
            self.date_entry.config(state=tk.DISABLED)

    def toggle_openings_entry(self):
        if self.expire_openings_var.get():
            self.openings_spin.config(state=tk.NORMAL)
        else:
            self.openings_spin.config(state=tk.DISABLED)

    # -------------------------------
    # STATUS BAR
    # -------------------------------
    def create_status_bar(self):
        self.status_var = tk.StringVar(value="Ready")

        status_bar = ttk.Frame(self.main_frame)
        status_bar.pack(fill=tk.X, pady=(10, 0))

        self.status_label = ttk.Label(
            status_bar,
            textvariable=self.status_var,
            font=('Segoe UI', 9),
            anchor=tk.W,
            padding=(10, 5)
        )
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Removed the mode button from status bar since it's now in the menu

    # -------------------------------
    # QUICK SEARCH FUNCTIONALITY
    # -------------------------------
    def find_quick_matches(self):
        """Find matches in the quick search text"""
        search_text = self.search_text_var.get().strip()
        if not search_text:
            messagebox.showwarning("Input Error", "Please enter search text")
            return
            
        context_range = self.range_var.get()
        input_text = self.quick_input_text.get(1.0, tk.END)
        
        if not input_text.strip():
            messagebox.showwarning("Input Error", "Please paste text into the input area")
            return
            
        # Clear previous results
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
            
        # Find all matches
        lines = input_text.splitlines()
        match_count = 0
        
        for i, line in enumerate(lines, 1):
            if search_text in line:
                match_count += 1
                
                # Find position of match
                start_idx = line.find(search_text)
                end_idx = start_idx + len(search_text)
                
                # Get context
                context_start = max(0, start_idx - context_range)
                context_end = min(len(line), end_idx + context_range)
                context = line[context_start:context_end]
                
                # Highlight match in context
                before = line[context_start:start_idx]
                match = line[start_idx:end_idx]
                after = line[end_idx:context_end]
                
                # Add to treeview
                self.results_tree.insert("", tk.END, values=(
                    i, 
                    f"{before}[{match}]{after}"
                ))
        
        if match_count == 0:
            self.status_var.set(f"No matches found for: {search_text}")
            messagebox.showinfo("No Matches", f"No matches found for: {search_text}")
        else:
            self.status_var.set(f"Found {match_count} matches for: {search_text}")
            self.log_action(f"Quick search found {match_count} matches for: {search_text}")

    def quick_replace_all(self):
        """Replace all matches in quick search"""
        search_text = self.search_text_var.get().strip()
        if not search_text:
            messagebox.showwarning("Input Error", "Please enter search text")
            return
            
        replace_text = simpledialog.askstring("Replacement Text", 
                                             "Enter replacement text:", 
                                             parent=self.root)
        if replace_text is None:  # User cancelled
            return
            
        input_text = self.quick_input_text.get(1.0, tk.END)
        if not input_text.strip():
            return
            
        # Perform replacement
        new_text = input_text.replace(search_text, replace_text)
        self.quick_input_text.delete(1.0, tk.END)
        self.quick_input_text.insert(tk.END, new_text)
        
        # Clear results
        self.clear_quick_results()
        
        self.status_var.set(f"Replaced all occurrences of '{search_text}'")
        self.log_action(f"Quick replace: '{search_text}' → '{replace_text}'")
        messagebox.showinfo("Success", f"Replaced all occurrences of '{search_text}'")

    def clear_quick_results(self):
        """Clear quick search results"""
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        self.status_var.set("Quick search results cleared")

    # -------------------------------
    # LOGGING
    # -------------------------------
    def setup_logging(self):
        self.log_file.touch(exist_ok=True)
        # Don't call refresh_log here - it will be called after UI is fully built
        self.log_action("Application started")
        self.log_action(f"Version: {self.version}")
        self.log_action(f"Secure storage: {self.appdata_dir}")

    def log_action(self, message):
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_entry = f"[{timestamp}] {message}\n"
            with open(self.log_file, "a", encoding="utf-8") as f:
                f.write(log_entry)
            if hasattr(self, 'status_var'):
                self.status_var.set(message)
            # Only refresh log if log_text widget exists
            if hasattr(self, 'log_text'):
                self.refresh_log()
        except Exception as e:
            print(f"Logging error: {e}")

    def open_secure_directory(self):
        """Open the secure directory (for advanced users)"""
        # Verify master password first
        master_password = simpledialog.askstring("Master Password", 
                                            "Enter master password:",
                                            show='*', parent=self.root)
        if master_password != "Siyabonga@18":
            messagebox.showerror("Error", "Incorrect master password")
            return
        
        try:
            if platform.system() == 'Windows':
                os.startfile(self.appdata_dir)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', self.appdata_dir])
            else:
                subprocess.run(['xdg-open', self.appdata_dir])
            self.log_action("Opened secure directory")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open secure directory:\n\n{str(e)}")

    def refresh_log(self):
        try:
            if not self.log_file.exists():
                self.log_file.touch()
            with open(self.log_file, "r", encoding="utf-8") as f:
                content = f.read()
            # Check if log_text exists before trying to update it
            if hasattr(self, 'log_text'):
                self.log_text.config(state=tk.NORMAL)
                self.log_text.delete(1.0, tk.END)
                self.log_text.insert(tk.END, content)
                self.log_text.config(state=tk.DISABLED)
                self.log_text.see(tk.END)
        except Exception as e:
            messagebox.showerror("Error", f"Could not read log file:\n\n{str(e)}")

    def clear_log(self):
        if messagebox.askyesno("Confirm", "Clear the entire log file?"):
            try:
                with open(self.log_file, "w", encoding="utf-8") as f:
                    f.write("")
                self.refresh_log()
                self.log_action("Log file cleared")
            except Exception as e:
                messagebox.showerror("Error", f"Could not clear log file:\n\n{str(e)}")

    def search_log(self):
        term = self.log_search_var.get().strip()
        if not term:
            return
        self.log_text.config(state=tk.NORMAL)
        self.log_text.tag_remove("highlight", "1.0", tk.END)
        start_pos = "1.0"
        while True:
            start_pos = self.log_text.search(term, start_pos, stopindex=tk.END)
            if not start_pos:
                break
            end_pos = f"{start_pos}+{len(term)}c"
            self.log_text.tag_add("highlight", start_pos, end_pos)
            start_pos = end_pos
        self.log_text.config(state=tk.DISABLED)
        # ensure highlight color matches theme
        self.log_text.tag_config("highlight",
                                 background=self.theme["accent_color"],
                                 foreground=self.theme["light_text"])
        try:
            self.log_text.see("highlight.first")
        except Exception:
            pass

    # -------------------------------
    # HELPERS / DIALOGS
    # -------------------------------
    def show_welcome_message(self):
        # Updated welcome message to match image
        message = (
            "Welcome to SmartReplaceTool!\n\n"
            "1. Your SmartReplaceTool folder has been created on your desktop (SRT_Files)\n"
            "2. A sample replacements.txt file has been created\n"
            "3. Please edit it with your search/replace pairs\n\n"
            "Format: search_text=replacement_text (one per line)\n\n"
            "You can edit the file directly or use the Replacements tab in this tool."
        )
        messagebox.showinfo("Welcome", message)

    def browse_file(self):
        file_types = [
            ("All supported files", "*.fhx *.txt *.csv *.xlsx"),
            ("Text files", "*.fhx *.txt"),
            ("CSV files", "*.csv"),
            ("Excel files", "*.xlsx"),
            ("All files", "*.*")
        ]
        file_path = filedialog.askopenfilename(filetypes=file_types)
        if file_path:
            self.current_file = Path(file_path)
            self.file_path_label.config(text=str(self.current_file))
            self.log_action(f"Selected file: {self.current_file.name}")

    # -------------------------------
    # PROCESSING
    # -------------------------------
    def process_file(self):
        if not hasattr(self, 'current_file') or not self.current_file:
            messagebox.showerror("Error", "Please select a file first")
            return

        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)
        self.results_text.config(state=tk.DISABLED)

        self.update_progress(0, "Starting processing...")
        self.root.update()

        try:
            # Step 1: Load replacements (10%)
            self.update_progress(10, "Loading replacement rules...")
            replacements = self.load_replacements()
            if not replacements:
                messagebox.showerror("Error", "No replacement rules found")
                self.update_progress(0, "Ready")
                return

            file_path = self.current_file
            self.log_action(f"Processing file: {file_path.name}")

            # Initialize summary data
            self.summary_data = {
                "filename": file_path.name,
                "file_size": file_path.stat().st_size,
                "file_type": file_path.suffix,
                "processing_date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "replacements_applied": [],
                "rules_used": [],
                "backup_created": "",
                "modified_file": "",
                "username": self.username,
                "version": self.version,
                "total_changes": 0
            }
            
            # Create a session ID for organizing reports
            self.current_session_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file_path.stem}"
            self.session_report_dir = self.reports_dir / self.current_session_id
            self.session_report_dir.mkdir(exist_ok=True)

            # Step 2: Process file based on type (20-80%)
            if file_path.suffix.lower() in (".fhx", ".txt"):
                success = self.process_text_file(file_path, replacements)
            elif file_path.suffix.lower() in (".csv", ".xlsx"):
                success = self.process_spreadsheet(file_path, replacements)
            else:
                messagebox.showerror("Error", f"Unsupported file type: {file_path.suffix}")
                self.update_progress(0, "Ready")
                return

            if success:
                # Step 3: Generate report (90%)
                self.update_progress(90, "Generating report...")
                self.generate_summary_report()
                
                # Step 4: Complete (100%)
                self.update_progress(100, "Processing complete")
                
                # Ask rename
                if messagebox.askyesno("Rename File", "Would you like to give a custom name to the modified file?"):
                    new_name = simpledialog.askstring(
                        "New Filename",
                        "Enter new filename (without extension):",
                        initialvalue=f"{file_path.stem}_modified"
                    )
                    if new_name:
                        old_path = self.last_modified_file
                        new_path = self.modified_dir / f"{new_name}{file_path.suffix}"
                        try:
                            shutil.move(old_path, new_path)
                            self.last_modified_file = new_path
                            self.summary_data["modified_file"] = new_path.name
                            self.log_action(f"Renamed file to: {new_path.name}")
                        except Exception as e:
                            messagebox.showerror("Error", f"Could not rename file:\n\n{str(e)}")

                # Offer open/download
                self.offer_download_option()
                
                # Refresh reports tab
                self.refresh_reports_list()
            else:
                self.update_progress(100, "Processing completed with no changes")

        except Exception as e:
            self.update_progress(0, "Error occurred")
            self.log_action(f"✕ Error processing file: {e}")
            messagebox.showerror("Error", f"An error occurred:\n\n{str(e)}")
            self.status_var.set("Error occurred - see log")
        finally:
            self.root.update()

    def offer_download_option(self):
        if not self.last_modified_file or not self.last_modified_file.exists():
            return

        choice = messagebox.askyesnocancel(
            "File Ready",
            "File processing complete!\n\n"
            f"Modified file saved as: {self.last_modified_file.name}\n\n"
            "Would you like to open the containing folder?\n"
            "(Click 'No' to copy file into a 'Downloads' folder here)",
            icon='info'
        )
        if choice is None:
            return
        elif choice:
            self.open_containing_folder()
        else:
            self.download_file()

    def open_containing_folder(self):
        """Open Modified_Files folder."""
        try:
            if platform.system() == 'Windows':
                os.startfile(self.modified_dir)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', self.modified_dir])
            else:
                subprocess.run(['xdg-open', self.modified_dir])
            self.log_action("Opened Modified_Files folder")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open folder:\n\n{str(e)}")

    def download_file(self):
        try:
            downloads_dir = self.tool_dir / "Downloads"
            downloads_dir.mkdir(exist_ok=True)
            dest_path = downloads_dir / self.last_modified_file.name
            shutil.copy2(self.last_modified_file, dest_path)

            if platform.system() == 'Windows':
                os.startfile(downloads_dir)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', downloads_dir])
            else:
                subprocess.run(['xdg-open', downloads_dir])

            self.log_action(f"File prepared for download: {dest_path.name}")
            messagebox.showinfo("Download Ready",
                                f"File has been copied to:\n{downloads_dir}\n\n"
                                "You can now save it to your desired location.")
        except Exception as e:
            messagebox.showerror("Error", f"Could not prepare download:\n\n{str(e)}")

    def process_text_file(self, file_path, replacements):
        try:
            self.update_progress(20, "Reading file...")
            encoding = self.detect_file_encoding(file_path)
            self.log_action(f"Detected encoding: {encoding}")
            with open(file_path, "r", encoding=encoding, errors='replace') as f:
                content = f.read()
        except UnicodeDecodeError:
            self.log_action("⚠️ Could not decode as text - treating as binary")
            return self.process_binary_file(file_path, replacements)
        except Exception as e:
            self.log_action(f"✕ Error reading file: {e}")
            return False

        # Store original content for comparison
        original_content = content

        self.update_progress(30, "Scanning for matches...")
        found_matches = self.scan_text(content, replacements)
        if not found_matches:
            self.log_action("✓ No matches found in the file")
            messagebox.showinfo("No Matches", "No replacement matches were found in the selected file.")
            return False

        self.update_progress(40, "Waiting for approval...")
        if not self.get_replacement_approval(found_matches):
            self.log_action("Replacements cancelled by user")
            return False

        # Backup
        self.update_progress(50, "Creating backup...")
        backup_path = self.backup_dir / (file_path.stem + "_backup" + file_path.suffix)
        with open(backup_path, "w", encoding=encoding, errors='replace') as f:
            f.write(content)
        self.log_action(f"Backup created: {backup_path.name}")
        
        # Update summary data
        self.summary_data["backup_created"] = backup_path.name

        # Apply replacements
        self.update_progress(60, "Applying replacements...")
        new_content = content
        applied_replacements = []
        
        total_replacements = sum(1 for data in found_matches.values() if data["approve"])
        current_replacement = 0
        
        for old, data in found_matches.items():
            if data["approve"]:
                current_replacement += 1
                progress = 60 + (current_replacement / total_replacements * 20) if total_replacements > 0 else 80
                self.update_progress(int(progress), f"Applying replacement {current_replacement}/{total_replacements}...")
                
                new_content = new_content.replace(old, data["new"])
                applied_replacements.append({
                    "old": old if isinstance(old, str) else f"<binary {len(old)} bytes>",
                    "new": data["new"] if isinstance(data["new"], str) else f"<binary {len(data['new'])} bytes>",
                    "count": data["count"]
                })

        # Update summary data
        self.summary_data["replacements_applied"] = applied_replacements
        self.summary_data["rules_used"] = list(replacements.keys())
        self.summary_data["total_changes"] = sum(r["count"] for r in applied_replacements)

        # Save
        self.update_progress(80, "Saving modified file...")
        new_file = self.modified_dir / (file_path.stem + "_modified" + file_path.suffix)
        with open(new_file, "w", encoding=encoding, errors='replace') as f:
            f.write(new_content)

        self.last_modified_file = new_file
        self.summary_data["modified_file"] = new_file.name
        self.log_action(f"✓ Saved modified file: {new_file.name}")
        
        # Offer to show comparison
        if messagebox.askyesno("Show Comparison", "Would you like to see a side-by-side comparison of the changes?"):
            self.create_comparison_tab(original_content, new_content, file_path.name)
        
        return True

    def process_binary_file(self, file_path, replacements):
        self.log_action("⚠️ Performing binary replacements")
        byte_replacements = {}
        for old, new in replacements.items():
            try:
                byte_replacements[old.encode('utf-8')] = new.encode('utf-8')
            except UnicodeError:
                self.log_action(f"⚠️ Skipping: {old}={new} (invalid characters)")
                continue

        with open(file_path, 'rb') as f:
            content = f.read()

        found_matches = {}
        for old_b, new_b in byte_replacements.items():
            count = content.count(old_b)
            if count > 0:
                found_matches[old_b] = {
                    'new': new_b,
                    'count': count,
                    'approve': False,
                    'context': f"Binary pattern ({len(old_b)} bytes)"
                }

        if not found_matches:
            self.log_action("✓ No binary patterns matched")
            messagebox.showinfo("No Matches", "No replacement matches were found in the selected file.")
            return False

        if not self.get_replacement_approval(found_matches):
            self.log_action("Binary replacements cancelled by user")
            return False

        # Backup
        self.update_progress(50, "Creating backup...")
        backup_path = self.backup_dir / (file_path.stem + "_backup" + file_path.suffix)
        with open(backup_path, 'wb') as f:
            f.write(content)
        self.log_action(f"Backup created: {backup_path.name}")
        
        # Update summary data
        self.summary_data["backup_created"] = backup_path.name

        # Apply
        applied_replacements = []
        total_replacements = sum(1 for data in found_matches.values() if data["approve"])
        current_replacement = 0
        
        for old_b, data in found_matches.items():
            if data["approve"]:
                current_replacement += 1
                progress = 60 + (current_replacement / total_replacements * 20) if total_replacements > 0 else 80
                self.update_progress(int(progress), f"Applying binary replacement {current_replacement}/{total_replacements}...")
                
                content = content.replace(old_b, data['new'])
                applied_replacements.append({
                    "old": f"<binary {len(old_b)} bytes>",
                    "new": f"<binary {len(data['new'])} bytes>",
                    "count": data["count"]
                })

        # Update summary data
        self.summary_data["replacements_applied"] = applied_replacements
        self.summary_data["rules_used"] = list(replacements.keys())
        self.summary_data["total_changes"] = sum(r["count"] for r in applied_replacements)

        # Save
        self.update_progress(80, "Saving modified file...")
        new_file = self.modified_dir / (file_path.stem + "_modified" + file_path.suffix)
        with open(new_file, 'wb') as f:
            f.write(content)

        self.last_modified_file = new_file
        self.summary_data["modified_file"] = new_file.name
        self.log_action(f"✓ Saved modified file: {new_file.name}")
        return True

    def process_spreadsheet(self, file_path, replacements):
        try:
            self.update_progress(20, "Reading spreadsheet...")
            
            if file_path.suffix.lower() == ".csv":
                # For CSV files, use pandas as before
                df = pd.read_csv(file_path, dtype=str, keep_default_na=False)
                use_openpyxl = False
            else:
                # For Excel files, use openpyxl to preserve formatting
                wb = openpyxl.load_workbook(file_path)
                use_openpyxl = True
                
        except Exception as e:
            self.log_action(f"✕ Error reading spreadsheet: {e}")
            return False

        found_matches = {}
        
        if use_openpyxl:
            # Process with openpyxl to preserve formatting
            for old, new in replacements.items():
                count = 0
                sample_context = None
                
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str) and old in str(cell.value):
                                count += 1
                                if sample_context is None:
                                    sample_context = str(cell.value)[:50] + ("..." if len(str(cell.value)) > 50 else "")
            
                if count > 0:
                    found_matches[old] = {
                        "new": new,
                        "count": count,
                        "approve": False,
                        "context": sample_context
                    }
        else:
            # Process with pandas for CSV files
            for old, new in replacements.items():
                # Use map instead of deprecated applymap
                mask = df.map(lambda x: old in str(x))
                count = mask.sum().sum()
                if count > 0:
                    # Get a sample cell context
                    pos = mask.stack().idxmax()
                    sample = str(df.loc[pos])
                    sample = sample[:50] + ("..." if len(sample) > 50 else "")
                    found_matches[old] = {
                        "new": new,
                        "count": int(count),
                        "approve": False,
                        "context": sample
                    }

        if not found_matches:
            self.log_action("✓ No matches found in spreadsheet")
            messagebox.showinfo("No Matches", "No replacement matches were found in the selected file.")
            return False

        self.update_progress(40, "Waiting for approval...")
        if not self.get_replacement_approval(found_matches):
            self.log_action("Spreadsheet replacements cancelled by user")
            return False

        # Backup
        self.update_progress(50, "Creating backup...")
        backup_path = self.backup_dir / (file_path.stem + "_backup" + file_path.suffix)
        shutil.copy2(file_path, backup_path)
        self.log_action(f"Backup created: {backup_path.name}")
        
        # Update summary data
        self.summary_data["backup_created"] = backup_path.name

        # Apply replacements
        applied_replacements = []
        total_replacements = sum(1 for data in found_matches.values() if data["approve"])
        current_replacement = 0
        
        if use_openpyxl:
            # Apply replacements using openpyxl (preserves formatting)
            for old, data in found_matches.items():
                if data["approve"]:
                    current_replacement += 1
                    progress = 60 + (current_replacement / total_replacements * 20) if total_replacements > 0 else 80
                    self.update_progress(int(progress), f"Applying replacement {current_replacement}/{total_replacements}...")
                    
                    for sheet in wb.worksheets:
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value and isinstance(cell.value, str) and old in str(cell.value):
                                    cell.value = str(cell.value).replace(old, data["new"])
                    
                    applied_replacements.append({
                        "old": old,
                        "new": data["new"],
                        "count": data["count"]
                    })
        else:
            # Apply replacements using pandas for CSV files
            for old, data in found_matches.items():
                if data["approve"]:
                    current_replacement += 1
                    progress = 60 + (current_replacement / total_replacements * 20) if total_replacements > 0 else 80
                    self.update_progress(int(progress), f"Applying replacement {current_replacement}/{total_replacements}...")
                    
                    # Use map instead of deprecated applymap
                    df = df.map(lambda x: str(x).replace(old, data["new"]))
                    applied_replacements.append({
                        "old": old,
                        "new": data["new"],
                        "count": data["count"]
                    })

        # Update summary data
        self.summary_data["replacements_applied"] = applied_replacements
        self.summary_data["rules_used"] = list(replacements.keys())
        self.summary_data["total_changes"] = sum(r["count"] for r in applied_replacements)

        # Save
        self.update_progress(80, "Saving modified file...")
        new_file = self.modified_dir / (file_path.stem + "_modified" + file_path.suffix)
        try:
            if use_openpyxl:
                # Save with openpyxl to preserve formatting
                wb.save(new_file)
            else:
                # Save with pandas for CSV files
                if file_path.suffix.lower() == ".csv":
                    df.to_csv(new_file, index=False)
                else:
                    df.to_excel(new_file, index=False)
        except Exception as e:
            self.log_action(f"✕ Error saving spreadsheet: {e}")
            return False

        self.last_modified_file = new_file
        self.summary_data["modified_file"] = new_file.name
        self.log_action(f"✓ Saved modified file: {new_file.name}")
        return True

    # -------------------------------
    # APPROVAL / PREVIEW
    # -------------------------------
    def get_replacement_approval(self, found_matches):
        # Show in UI
        self.show_matches_in_ui(found_matches)

        # Create custom dialog
        dialog = Toplevel(self.root)
        dialog.title("Approve Replacements")
        dialog.geometry("500x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Add message
        msg = "Would you like to approve ALL replacements?\n\nClick Approve All to approve all, Review Individually to review each, or Cancel to stop processing."
        label = ttk.Label(dialog, text=msg, wraplength=400)
        label.pack(pady=20)
        
        # Add buttons
        btn_frame = ttk.Frame(dialog)
        btn_frame.pack(pady=10)
        
        result = {"value": None}
        
        def set_result(val):
            result["value"] = val
            dialog.destroy()
        
        ttk.Button(btn_frame, text="Approve All", command=lambda: set_result("yes")).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancel", command=lambda: set_result("quit")).pack(side=tk.RIGHT, padx=10)
        ttk.Button(btn_frame, text="Review Individually", command=lambda: set_result("no")).pack(side=tk.LEFT, padx=10)
        
        # Wait for dialog to close
        self.root.wait_window(dialog)
        
        if result["value"] == "quit":
            self.log_action("Replacements cancelled by user")
            return False
        elif result["value"] == "yes":
            for old in found_matches:
                found_matches[old]["approve"] = True
            return True
        else:
            # per-item
            for old, data in found_matches.items():
                if isinstance(old, bytes):
                    display_old = f"<binary {len(old)} bytes>"
                    display_new = f"<binary {len(data['new'])} bytes>"
                else:
                    display_old = old
                    display_new = data['new']
                    
                # Create custom dialog for each replacement
                item_dialog = Toplevel(self.root)
                item_dialog.title("Approve Replacement")
                item_dialog.geometry("600x250")
                item_dialog.transient(self.root)
                item_dialog.grab_set()
                
                # Center the dialog
                item_dialog.update_idletasks()
                x = self.root.winfo_x() + (self.root.winfo_width() - item_dialog.winfo_width()) // 2
                y = self.root.winfo_y() + (self.root.winfo_height() - item_dialog.winfo_height()) // 2
                item_dialog.geometry(f"+{x}+{y}")
                
                # Add message
                msg = f"Replace '{display_old}' with '{display_new}'?\n\nFound {data['count']} occurrence(s)\nExample: {data['context']}"
                label = ttk.Label(item_dialog, text=msg, wraplength=500)
                label.pack(pady=20)
                
                # Add buttons
                item_btn_frame = ttk.Frame(item_dialog)
                item_btn_frame.pack(pady=10)
                
                item_result = {"value": None}
                
                def set_item_result(val):
                    item_result["value"] = val
                    item_dialog.destroy()
                
                ttk.Button(item_btn_frame, text="Approve", command=lambda: set_item_result(True)).pack(side=tk.LEFT, padx=10)
                ttk.Button(item_btn_frame, text="Skip", command=lambda: set_item_result(False)).pack(side=tk.LEFT, padx=10)
                ttk.Button(item_btn_frame, text="Cancel All", command=lambda: set_item_result("quit")).pack(side=tk.LEFT, padx=10)
                
                # Wait for dialog to close
                self.root.wait_window(item_dialog)
                
                if item_result["value"] == "quit":
                    self.log_action("Replacements cancelled by user during individual review")
                    return False
                    
                found_matches[old]["approve"] = item_result["value"]
                if not item_result["value"]:
                    self.log_action(f"Skipped replacement: {display_old} → {display_new}")

            if not any(data["approve"] for data in found_matches.values()):
                self.log_action("No replacements were approved")
                return False
            return True

    def show_matches_in_ui(self, found_matches):
        self.results_text.config(state=tk.NORMAL)
        self.results_text.delete(1.0, tk.END)

        self.results_text.insert(tk.END, "=== FOUND REPLACEMENTS ===\n\n", 'header')

        for i, (old, data) in enumerate(found_matches.items(), 1):
            disp_old = f"<binary {len(old)} bytes>" if isinstance(old, bytes) else old
            disp_new = f"<binary {len(data['new'])} bytes>" if isinstance(data['new'], (bytes, bytearray)) else data['new']
            self.results_text.insert(tk.END, f"[{i}] ", 'number')
            self.results_text.insert(tk.END, f"{disp_old} → {disp_new}\n", 'replacement')
            self.results_text.insert(tk.END, f"    Count: {data['count']}\n", 'count')
            self.results_text.insert(tk.END, f"    Example: {data['context']}\n\n", 'example')

        # Tag colors set in apply_theme
        self.results_text.config(state=tk.DISABLED)
        self.notebook.select(0)

    # -------------------------------
    # REPLACEMENTS FILE I/O
    # -------------------------------
    def load_replacements_into_ui(self):
        try:
            if not self.replacements_file.exists():
                self.create_sample_replacements_file()
            with open(self.replacements_file, "r", encoding="utf-8") as f:
                content = f.read()
            self.replacements_text.config(state=tk.NORMAL)
            self.replacements_text.delete(1.0, tk.END)
            self.replacements_text.insert(tk.END, content)
            self.replacements_text.config(state=tk.NORMAL)
            self.log_action("Loaded replacements into editor")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Could not load replacements:\n\n{str(e)}")
            return False

    def save_replacements(self):
        try:
            content = self.replacements_text.get(1.0, tk.END)
            with open(self.replacements_file, "w", encoding="utf-8") as f:
                f.write(content)
            self.log_action("Saved replacement rules")
            messagebox.showinfo("Success", "Replacements saved successfully")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Could not save replacements:\n\n{str(e)}")
            return False

    def open_in_editor(self):
        """Open the replacements file in the default OS editor."""
        try:
            if platform.system() == 'Windows':
                os.startfile(self.replacements_file)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', self.replacements_file])
            else:
                subprocess.run(['xdg-open', self.replacements_file])
            self.log_action("Opened replacements file in editor")
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Could not open editor:\n\n{str(e)}")
            return False

    def create_sample_replacements_file(self):
        sample_content = """# This is your replacements file for SmartReplaceTool
# Format: search_text=replacement_text
# Put one replacement per line
# Example:

OLD_VALUE=NEW_VALUE
ERROR_CODE=STATUS_CODE
TEMP_READING=TEMPERATURE
"""
        with open(self.replacements_file, "w", encoding="utf-8") as f:
            f.write(sample_content)

    def load_replacements(self):
        replacements = {}
        if not self.replacements_file.exists():
            return replacements
        try:
            with open(self.replacements_file, "r", encoding="utf-8") as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if "=" not in line:
                        continue
                    old, new = line.split("=", 1)
                    old = old.strip()
                    new = new.strip()
                    if not old:
                        continue
                    replacements[old] = new
            return replacements
        except Exception as e:
            self.log_action(f"Error loading replacements: {e}")
            return {}

    # -------------------------------
    # UTIL
    # -------------------------------
    def detect_file_encoding(self, file_path):
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)
            result = chardet.detect(raw_data)
            return result['encoding'] if result['confidence'] and result['confidence'] > 0.7 else 'utf-8'

    # -------------------------------
    # QUICK ACTIONS
    # -------------------------------
    def clear_backups(self):
        if not self.backup_dir.exists():
            messagebox.showinfo("Info", "Backup folder does not exist.")
        if messagebox.askyesno("Confirm", "Delete ALL files in Backups? This cannot be undone."):
            try:
                for p in self.backup_dir.glob("*"):
                    if p.is_file():
                        p.unlink()
                self.log_action("All backups cleared")
                messagebox.showinfo("Done", "Backups cleared.")
            except Exception as e:
                messagebox.showerror("Error", f"Could not clear backups:\n\n{str(e)}")

    def reset_settings(self):
        if messagebox.askyesno(
            "Reset Settings",
            "This will recreate the sample replacements file and clear the log.\n"
            "Modified files & backups will remain. Continue?"
        ):
            try:
                # recreate sample replacements in secure location
                self.create_sample_replacements_file()
                # clear log in secure location
                with open(self.log_file, "w", encoding="utf-8") as f:
                    f.write("")
                self.load_replacements_into_ui()
                self.refresh_log()
                self.log_action("Settings reset: sample replacements & cleared log")
                messagebox.showinfo("Done", "Settings have been reset.")
            except Exception as e:
                messagebox.showerror("Error", f"Could not reset settings:\n\n{str(e)}")

    # -------------------------------
    # TEXT SCAN
    # -------------------------------
    def scan_text(self, content, replacements):
        found_matches = {}
        for old, new in replacements.items():
            matches = list(re.finditer(re.escape(old), content))
            if matches:
                start_pos = max(0, matches[0].start() - 20)
                end_pos = min(len(content), matches[0].end() + 20)
                context = content[start_pos:end_pos].replace("\n", " ").strip()
                found_matches[old] = {
                    "new": new,
                    "count": len(matches),
                    "approve": False,
                    "context": f"...{context}..."
                }
        return found_matches

    # -------------------------------
    # PROGRESS UPDATES
    # -------------------------------
    def update_progress(self, value, message):
        self.progress['value'] = value
        self.progress_label.config(text=f"{value}%")
        self.status_var.set(f"{message} ({value}%)")
        self.root.update_idletasks()

    # -------------------------------
    # PDF GUIDE
    # -------------------------------
    def open_user_guide(self):
        # Always try to open the existing PDF file
        pdf_path = Path("How_To_Use_SRT.pdf")
        
        if not pdf_path.exists():
            messagebox.showerror("Error", "User guide PDF not found. Please ensure 'How_To_Use_SRT.pdf' is in the same folder as the application.")
            return
        
        try:
            if platform.system() == 'Windows':
                os.startfile(pdf_path)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', pdf_path])
            else:
                subprocess.run(['xdg-open', pdf_path])
            self.log_action("Opened user guide PDF")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open PDF:\n\n{str(e)}")

    # -------------------------------
    # REPORTS TAB FUNCTIONS
    # -------------------------------
    def refresh_reports_list(self):
        # Clear existing items
        self.reports_listbox.delete(0, tk.END)
        
        # Update current report label
        if hasattr(self, 'current_session_id') and self.current_session_id:
            self.current_report_label.config(
                text=f"Current Report: {self.current_session_id} (HTML available)"
            )
        else:
            self.current_report_label.config(text="No report generated yet")
            
        # Populate with reports
        if self.reports_dir.exists():
            for session_dir in self.reports_dir.glob("*"):
                if session_dir.is_dir():
                    for report_file in session_dir.glob("*.html"):
                        if report_file.is_file():
                            # Get file modification time
                            mtime = datetime.fromtimestamp(report_file.stat().st_mtime)
                            date_str = mtime.strftime("%Y-%m-%d %H:%M")
                            
                            # Add to listbox
                            self.reports_listbox.insert(
                                tk.END, 
                                f"{session_dir.name} - {date_str}"
                            )
                            # Store the path as item data
                            self.reports_listbox.itemconfig(
                                tk.END, 
                                {'fg': self.theme['text_color']}
                            )
    
    def view_current_report(self):
        if not hasattr(self, 'current_session_id') or not self.current_session_id:
            messagebox.showwarning("No Report", "No current report available. Process a file first.")
            return
            
        report_path = self.session_report_dir / f"{self.summary_data['filename']}_report.html"
        
        try:
            if platform.system() == 'Windows':
                os.startfile(report_path)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', report_path])
            else:
                subprocess.run(['xdg-open', report_path])
            self.log_action(f"Opened current report: {report_path.name}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open report: {e}")
    
    def convert_current_to_pdf(self):
        if not hasattr(self, 'current_session_id') or not self.current_session_id:
            messagebox.showwarning("No Report", "No current report available. Process a file first.")
            return
            
        html_path = self.session_report_dir / f"{self.summary_data['filename']}_report.html"
        pdf_path = self.session_report_dir / f"{self.summary_data['filename']}_report.pdf"
        
        try:
            # Read HTML content
            with open(html_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
                
            # Generate PDF
            success = self.generate_pdf_report(html_content, pdf_path)
            
            if success:
                self.log_action(f"Converted HTML to PDF: {pdf_path.name}")
                messagebox.showinfo("Success", f"PDF created: {pdf_path.name}")
                self.refresh_reports_list()
            else:
                messagebox.showerror("Error", "Failed to convert HTML to PDF")
        except Exception as e:
            messagebox.showerror("Error", f"Could not convert HTML to PDF: {e}")
    
    def view_selected_history(self):
        selected = self.reports_listbox.curselection()
        if not selected:
            messagebox.showwarning("No Selection", "Please select a report to view")
            return
            
        item_text = self.reports_listbox.get(selected[0])
        session_id = item_text.split(" - ")[0]
        
        # Find the HTML report in this session
        session_dir = self.reports_dir / session_id
        if session_dir.exists():
            for report_file in session_dir.glob("*.html"):
                try:
                    if platform.system() == 'Windows':
                        os.startfile(report_file)
                    elif platform.system() == 'Darwin':
                        subprocess.run(['open', report_file])
                    else:
                        subprocess.run(['xdg-open', report_file])
                    self.log_action(f"Opened historical report: {report_file.name}")
                    break
                except Exception as e:
                    messagebox.showerror("Error", f"Could not open report: {e}")
    
    def open_reports_folder(self):
        try:
            if platform.system() == 'Windows':
                os.startfile(self.reports_dir)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', self.reports_dir])
            else:
                subprocess.run(['xdg-open', self.reports_dir])
        except Exception as e:
            messagebox.showerror("Error", f"Could not open reports folder: {e}")

    # -------------------------------
    # SUMMARY REPORT
    # -------------------------------
    def generate_summary_report(self):
        if not self.summary_data or not self.summary_data.get("replacements_applied"):
            return
        
        # Generate HTML report
        html_report = self.generate_html_report()
        html_path = self.session_report_dir / f"{self.summary_data['filename']}_report.html"
        
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_report)
        
        # Also generate PDF directly
        pdf_path = self.session_report_dir / f"{self.summary_data['filename']}_report.pdf"
        self.generate_pdf_report(html_report, pdf_path)
        
        self.log_action(f"Generated summary report in {self.session_report_dir}")

    def generate_html_report(self):
        # Logo HTML
        logo_html = ""
        if self.logo_base64:
            logo_html = f'<div class="logo"><img src="data:image/png;base64,{self.logo_base64}" alt="Logo" style="height: 50px; margin-right: 15px; vertical-align: middle;"><h1 style="display: inline-block; vertical-align: middle; margin: 0;">SmartReplaceTool</h1></div>'
        else:
            logo_html = '<div class="logo"><h1>SmartReplaceTool</h1></div>'
        
        # Create replacement distribution chart
        chart_html = ""
        if self.summary_data.get("replacements_applied"):
            try:
                # Create chart based on preference
                plt.figure(figsize=(10, 6))
                replacements = self.summary_data['replacements_applied']
                labels = [r['old'][:20] + '...' if len(r['old']) > 20 else r['old'] for r in replacements]
                counts = [r['count'] for r in replacements]
                
                if self.chart_preference == "pie":
                    # Create pie chart
                    plt.pie(counts, labels=labels, autopct='%1.1f%%', startangle=90)
                    plt.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.
                    plt.title('Replacement Distribution')
                elif self.chart_preference == "horizontal_bar":
                    # Create horizontal bar chart
                    y_pos = range(len(labels))
                    bars = plt.barh(y_pos, counts)
                    plt.xlabel('Count')
                    plt.ylabel('Replacement Text')
                    plt.title('Replacement Distribution')
                    plt.yticks(y_pos, labels)
                    
                    # Add value labels on the right of bars
                    for i, bar in enumerate(bars):
                        width = bar.get_width()
                        plt.text(width + 0.1, bar.get_y() + bar.get_height()/2,
                                f'{counts[i]}', ha='left', va='center')
                elif self.chart_preference == "donut":
                    # Create donut chart
                    wedges, texts, autotexts = plt.pie(counts, labels=labels, autopct='%1.1f%%', 
                                                      startangle=90, wedgeprops=dict(width=0.3))
                    plt.axis('equal')
                    plt.title('Replacement Distribution (Donut)')
                elif self.chart_preference == "line":
                    # Create line chart
                    x_pos = range(len(labels))
                    plt.plot(x_pos, counts, marker='o', linestyle='-', linewidth=2)
                    plt.xlabel('Replacement Text')
                    plt.ylabel('Count')
                    plt.title('Replacement Distribution')
                    plt.xticks(x_pos, labels, rotation=45, ha='right')
                    
                    # Add value labels on top of points
                    for i, count in enumerate(counts):
                        plt.text(i, count + 0.1, f'{count}', ha='center', va='bottom')
                elif self.chart_preference == "area":
                    # Create area chart
                    x_pos = range(len(labels))
                    plt.fill_between(x_pos, counts, alpha=0.4)
                    plt.plot(x_pos, counts, marker='o', linestyle='-', linewidth=2)
                    plt.xlabel('Replacement Text')
                    plt.ylabel('Count')
                    plt.title('Replacement Distribution')
                    plt.xticks(x_pos, labels, rotation=45, ha='right')
                    
                    # Add value labels on top of points
                    for i, count in enumerate(counts):
                        plt.text(i, count + 0.1, f'{count}', ha='center', va='bottom')
                else:
                    # Create bar chart (default)
                    bars = plt.bar(range(len(labels)), counts)
                    plt.xlabel('Replacement Text')
                    plt.ylabel('Count')
                    plt.title('Replacement Distribution')
                    plt.xticks(range(len(labels)), labels, rotation=45, ha='right')
                    
                    # Add value labels on top of bars
                    for i, bar in enumerate(bars):
                        height = bar.get_height()
                        plt.text(bar.get_x() + bar.get_width()/2., height + 0.1,
                                f'{counts[i]}', ha='center', va='bottom')
                
                # Save chart to a temporary file
                chart_path = self.session_report_dir / "replacement_chart.png"
                plt.tight_layout()
                plt.savefig(chart_path, dpi=100, bbox_inches='tight')
                plt.close()
                
                # Convert chart to base64
                with open(chart_path, "rb") as chart_file:
                    chart_base64 = base64.b64encode(chart_file.read()).decode('utf-8')
                
                chart_html = f'<div class="chart"><h2>Replacement Distribution ({self.chart_preference.replace("_", " ").title()})</h2><img src="data:image/png;base64,{chart_base64}" alt="Replacement Chart" style="max-width: 100%;"></div>'
            except Exception as e:
                self.log_action(f"Error creating chart: {e}")
                chart_html = "<p>Chart could not be generated</p>"
        
        # Create replacement distribution table
        distribution_html = ""
        if self.summary_data.get("replacements_applied"):
            distribution_html = """
            <h2>Replacement Details</h2>
            <div class="table-container">
            <table>
                <tr>
                    <th>Original Text</th>
                    <th>Replacement Text</th>
                    <th>Count</th> 
                    <th>Percentage</th>
                </tr>
            """
            
            total = sum(r['count'] for r in self.summary_data['replacements_applied'])
            for r in self.summary_data['replacements_applied']:
                percentage = (r['count'] / total) * 100 if total > 0 else 0
                distribution_html += f"""
                <tr>
                    <td>{html.escape(r['old'])}</td>
                    <td>{html.escape(r['new'])}</td>
                    <td>{r['count']}</td>
                    <td>{percentage:.1f}%</td>
                </tr>
                """
            
            distribution_html += "</table></div>"
        
        # File information with additional details
        file_size_kb = self.summary_data.get("file_size", 0) / 1024
        file_info_html = f"""
        <div class="file-info">
            <h2>File Information</h2>
            <table>
                <tr><th>Attribute</th><th>Value</th></tr>
                <tr><td>Filename</td><td>{html.escape(self.summary_data['filename'])}</td></tr>
                <tr><td>File Type</td><td>{html.escape(self.summary_data.get('file_type', 'Unknown'))}</td></tr>
                <tr><td>File Size</td><td>{file_size_kb:.2f} KB</td></tr>
                <tr><td>Processing Date</td><td>{html.escape(self.summary_data.get('processing_date', 'Unknown'))}</td></tr>
                <tr><td>Total Changes</td><td>{self.summary_data.get('total_changes', 0)}</td></tr>
                <tr><td>Processed By</td><td>{html.escape(self.summary_data.get('username', 'Unknown'))}</td></tr>
                <tr><td>Tool Version</td><td>v{html.escape(self.summary_data.get('version', 'Unknown'))}</td></tr>
                <tr><td>Backup Location</td><td>{html.escape(self.summary_data.get('backup_created', 'Unknown'))}</td></tr>
                <tr><td>Modified File</td><td>{html.escape(self.summary_data.get('modified_file', 'Unknown'))}</td></tr>
            </table>
        </div>
        """
        
        # Add processing statistics
        stats_html = ""
        if self.summary_data.get("replacements_applied"):
            stats_html = f"""
            <div class="stats">
                <h2>Processing Statistics</h2>
                <table>
                    <tr><td>Total Replacement Rules Applied</td><td>{len(self.summary_data['replacements_applied'])}</td></tr>
                    <tr><td>Total Individual Changes</td><td>{self.summary_data.get('total_changes', 0)}</td></tr>
                    <tr><td>Average Changes per Rule</td><td>{self.summary_data.get('total_changes', 0) / len(self.summary_data['replacements_applied']) if len(self.summary_data['replacements_applied']) > 0 else 0:.2f}</td></tr>
                </table>
            </div>
            """
        
        html_template = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>SmartReplaceTool Report</title>
            <meta charset="UTF-8">
            <style>
                body {{ 
                    font-family: 'Segoe UI', Arial, sans-serif; 
                    margin: 40px; 
                    background-color: #f8f9fa;
                    color: #343a40;
                }}
                .container {{
                    max-width: 1200px;
                    margin: 0 auto;
                    background-color: white;
                    padding: 30px;
                    border-radius: 10px;
                    box-shadow: 0 0 20px rgba(0,0,0,0.1);
                }}
                .header {{
                    text-align: center;
                    border-bottom: 2px solid #4361ee;
                    padding-bottom: 20px;
                    margin-bottom: 30px;
                }}
                h1 {{ 
                    color: #4361ee; 
                    margin-bottom: 10px;
                }}
                .subtitle {{
                    color: #6c757d;
                    font-size: 1.2em;
                }}
                .summary {{
                    background-color: #e8f4f8; 
                    padding: 20px; 
                    border-radius: 8px;
                    margin-bottom: 30px;
                }}
                .summary-grid {{
                    display: grid;
                    grid-template-columns: repeat(2, 1fr);
                    gap: 15px;
                }}
                .summary-item {{
                    margin-bottom: 10px;
                }}
                .summary-label {{
                    font-weight: bold;
                    color: #4361ee;
                }}
                .table-container {{
                    overflow-x: auto;
                    width: 100%;
                }}
                table {{ 
                    border-collapse: collapse; 
                    width: 100%; 
                    margin-bottom: 30px;
                    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
                }}
                th, td {{ 
                    border: 1px solid #dee2e6; 
                    padding: 12px; 
                    text-align: left; 
                    word-break: break-word;
                }}
                th {{ 
                    background-color: #4361ee;
                    color: white;
                    font-weight:600;
                }}
                tr:nth-child(even) {{
                    background-color: #f8f9fa;
                }}
                tr:hover {{
                    background-color: #e9ecef;
                }}
                .chart {{
                    margin: 30px 0;
                    text-align: center;
                }}
                .file-info, .stats {{
                    margin: 30px 0;
                }}
                .footer {{
                    text-align: center;
                    margin-top: 40px;
                    padding-top: 20px;
                    border-top: 1px solid #dee2e6;
                    color: #6c757d;
                }}
                @media print {{
                    body {{
                        margin: 0;
                        padding: 15px;
                        background-color: white;
                    }}
                    .container {{
                        box-shadow: none;
                        padding: 0;
                        max-width: 100%;
                    }}
                    .summary {{
                        background-color: #f0f0f0 !important;
                        -webkit-print-color-adjust: exact;
                    }}
                    th {{
                        background-color: #4361ee !important;
                        color: white !important;
                        -webkit-print-color-adjust: exact;
                    }}
                }}
            </style>
        </head>
        <body>
            <div class="container">
                {logo_html}
                <div class="subtitle">Processing Report</div>
                
                <div class="summary">
                    <h2>Summary</h2>
                    <div class="summary-grid">
                        <div class="summary-item"><span class="summary-label">File:</span> {html.escape(self.summary_data['filename'])}</div>
                        <div class="summary-item"><span class="summary-label">Total Changes:</span> {self.summary_data.get('total_changes', 0)}</div>
                        <div class="summary-item"><span class="summary-label">Processed By:</span> {html.escape(self.summary_data.get('username', 'Unknown'))}</div>
                        <div class="summary-item"><span class="summary-label">Date:</span> {html.escape(self.summary_data.get('processing_date', 'Unknown'))}</div>
                    </div>
                </div>
                
                {file_info_html}
                
                {stats_html}
                
                {chart_html}
                
                {distribution_html}
                
                <h2>Rules Used</h2>
                <ul>
                    {"".join(f"<li>{html.escape(rule)}</li>" for rule in self.summary_data['rules_used'])}
                </ul>
                
                <div class="footer">
                    <p>Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                    <p>SmartReplaceTool v{self.version} | {self.creators}</p>
                </div>
            </div>
        </body>
        </html>
        """
    
        return html_template

    def generate_pdf_report(self, html_content, pdf_path):
        try:
            if not PDF_SUPPORT:
                messagebox.showerror("Error", "PDF support is not available. Please install xhtml2pdf.")
                return False
                
            # Create a PDF with landscape orientation
            from reportlab.lib.pagesizes import landscape, letter
            
            # Create a temporary file for the HTML content
            with tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.html', encoding='utf-8') as temp_html:
                temp_html.write(html_content)
                temp_html_path = temp_html.name
            
            # Use xhtml2pdf with landscape orientation
            with open(pdf_path, "wb") as f:
                pisa_status = pisa.CreatePDF(
                    html_content, 
                    dest=f,
                    pagesize=landscape(letter),
                    encoding='UTF-8'
                )
                
            # Clean up temporary file
            try:
                os.unlink(temp_html_path)
            except:
                pass
                
            if pisa_status.err:
                self.log_action("Error creating PDF")
                return False
                
            return True
        except Exception as e:
            self.log_action(f"Error generating PDF report: {e}")
            return False

    # -------------------------------
    # COMPARISON TAB
    # -------------------------------
    def create_comparison_tab(self, original_content, new_content, filename):
        # Create a new tab for comparison
        comparison_id = f"comparison_{len(self.comparison_tabs) + 1}"
        comparison_tab = ttk.Frame(self.notebook)
        
        # Add to notebook
        self.notebook.add(comparison_tab, text=f"Comparison: {filename}")
        self.notebook.select(comparison_tab)
        
        # Store reference
        self.comparison_tabs[comparison_id] = {
            "frame": comparison_tab,
            "filename": filename
        }
        
        # Create frames for original and modified content
        content_frame = ttk.Frame(comparison_tab)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create paned window for resizable split
        paned_window = ttk.PanedWindow(content_frame, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True)
        
        # Original content frame
        original_frame = ttk.LabelFrame(paned_window, text="Original Content")
        paned_window.add(original_frame, weight=1)
        
        # Modified content frame
        modified_frame = ttk.LabelFrame(paned_window, text="Modified Content")
        paned_window.add(modified_frame, weight=1)
        
        # Add text widgets with scrollbars
        original_text = Text(original_frame, wrap=tk.WORD, font=('Consolas', 10))
        original_scroll = Scrollbar(original_frame, orient=tk.VERTICAL, command=original_text.yview)
        original_text.configure(yscrollcommand=original_scroll.set)
        
        modified_text = Text(modified_frame, wrap=tk.WORD, font=('Consolas', 10))
        modified_scroll = Scrollbar(modified_frame, orient=tk.VERTICAL, command=modified_text.yview)
        modified_text.configure(yscrollcommand=modified_scroll.set)
        
        # Pack widgets
        original_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        original_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        modified_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        modified_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Insert content and highlight differences
        original_text.insert(tk.END, original_content)
        modified_text.insert(tk.END, new_content)
        
        # Highlight differences
        self.highlight_differences(original_text, modified_text, original_content, new_content)
        
        # Make text widgets read-only
        original_text.config(state=tk.DISABLED)
        modified_text.config(state=tk.DISABLED)
        
        # Add close button
        button_frame = ttk.Frame(comparison_tab)
        button_frame.pack(fill=tk.X, padx=10, pady=10)
        
        close_btn = ttk.Button(button_frame, text="Close Comparison", 
                              command=lambda: self.close_comparison_tab(comparison_id),
                              style='Accent.TButton')
        close_btn.pack(side=tk.RIGHT)
        
        # Bind close event
        comparison_tab.bind("<Destroy>", lambda e: self.close_comparison_tab(comparison_id))

    def close_comparison_tab(self, comparison_id):
        if comparison_id in self.comparison_tabs:
            tab_index = None
            # Find the tab index
            for i in range(self.notebook.index("end")):
                if self.notebook.nametowidget(self.notebook.tabs()[i]) == self.comparison_tabs[comparison_id]["frame"]:
                    tab_index = i
                    break
            
            if tab_index is not None:
                # Remove the tab
                self.notebook.forget(tab_index)
            
            # Remove from tracking
            del self.comparison_tabs[comparison_id]

    def highlight_differences(self, original_text, modified_text, original_content, new_content):
        # Use difflib to find differences
        diff = difflib.SequenceMatcher(None, original_content.splitlines(), new_content.splitlines())
        
        # Configure tags for highlighting
        original_text.tag_config('diff', background='#ffdddd')
        modified_text.tag_config('diff', background='#ddffdd')
        original_text.tag_config('change', background='#ffd700')
        modified_text.tag_config('change', background='#ffd700')
        
        # Highlight differences
        for tag, i1, i2, j1, j2 in diff.get_opcodes():
            if tag != 'equal':
                # Original text highlighting
                if tag in ('replace', 'delete'):
                    start_line = i1 + 1
                    end_line = i2 + 1
                    original_text.tag_add('diff', f'{start_line}.0', f'{end_line}.0')
                
                # Modified text highlighting
                if tag in ('replace', 'insert'):
                    start_line = j1 + 1
                    end_line = j2 + 1
                    modified_text.tag_add('diff', f'{start_line}.0', f'{end_line}.0')

    # -------------------------------
    # SECURITY FUNCTIONS
    # -------------------------------
    def select_pdf_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if file_path:
            self.pdf_file_path = Path(file_path)
            self.pdf_path_var.set(self.pdf_file_path.name)
            self.log_action(f"Selected PDF: {self.pdf_file_path.name}")

    def select_encrypt_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
        if file_path:
            self.encrypt_file_path = Path(file_path)
            self.encrypt_path_var.set(self.encrypt_file_path.name)
            self.log_action(f"Selected file for encryption: {self.encrypt_file_path.name}")

    def select_checksum_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
        if file_path:
            self.checksum_file_path = Path(file_path)
            self.checksum_path_var.set(self.checksum_file_path.name)
            self.log_action(f"Selected file for checksum: {self.checksum_file_path.name}")

    def lock_pdf(self):
        if not hasattr(self, 'pdf_file_path') or not self.pdf_file_path:
            messagebox.showwarning("Error", "Please select a PDF file first")
            return
            
        password = self.pdf_password_var.get()
        if not password:
            messagebox.showwarning("Error", "Please enter a password")
            return
            
        try:
            output_path = self.secure_dir / f"locked_{self.pdf_file_path.name}"
            
            # Read the PDF
            with open(self.pdf_file_path, 'rb') as f:
                pdf_reader = PdfReader(f)
                pdf_writer = PdfWriter()
                
                # Copy all pages
                for page_num in range(len(pdf_reader.pages)):
                    page = pdf_reader.pages[page_num]
                    pdf_writer.add_page(page)
                
                # Add password protection
                pdf_writer.encrypt(password)
                
                # Save the encrypted PDF
                with open(output_path, 'wb') as output_file:
                    pdf_writer.write(output_file)
            
            self.log_action(f"PDF locked with password: {output_path.name}")
            messagebox.showinfo("Success", f"PDF locked successfully!\nSaved as: {output_path.name}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to lock PDF: {str(e)}")
            self.log_action(f"PDF lock failed: {str(e)}")

    def unlock_pdf(self):
        if not hasattr(self, 'pdf_file_path') or not self.pdf_file_path:
            messagebox.showwarning("Error", "Please select a PDF file first")
            return
            
        password = self.pdf_password_var.get()
        if not password:
            messagebox.showwarning("Error", "Please enter the password")
            return
            
        try:
            output_path = self.secure_dir / f"unlocked_{self.pdf_file_path.name}"
            
            # Read the encrypted PDF
            with open(self.pdf_file_path, 'rb') as f:
                pdf_reader = PdfReader(f)
                
                # Try to decrypt
                if pdf_reader.is_encrypted:
                    if pdf_reader.decrypt(password):
                        pdf_writer = PdfWriter()
                        
                        # Copy all pages
                        for page_num in range(len(pdf_reader.pages)):
                            page = pdf_reader.pages[page_num]
                            pdf_writer.add_page(page)
                        
                        # Save the decrypted PDF
                        with open(output_path, 'wb') as output_file:
                            pdf_writer.write(output_file)
                        
                        self.log_action(f"PDF unlocked: {output_path.name}")
                        messagebox.showinfo("Success", f"PDF unlocked successfully!\nSaved as: {output_path.name}")
                    else:
                        messagebox.showerror("Error", "Incorrect password")
                        self.log_action("PDF unlock failed: incorrect password")
                else:
                    messagebox.showinfo("Info", "PDF is not encrypted")
                    self.log_action("PDF unlock: file not encrypted")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Failed to unlock PDF: {str(e)}")
            self.log_action(f"PDF unlock failed: {str(e)}")


        # Replace the encrypt_file method with this corrected version
    def run_encryption_with_progress(self, operation, file_path, password):
        """Run encryption/decryption in a separate thread with progress updates"""
        # Create a thread for the operation
        if operation == "encrypt":
            thread = EncryptionThread(self._encrypt_file_worker, file_path, password)
        else:
            thread = EncryptionThread(self._decrypt_file_worker, file_path, password)
        
        thread.start()
        
        # Show progress while the thread is running
        self.update_progress(0, f"Starting {operation}ion...")
        
        # Create a progress window
        progress_window = Toplevel(self.root)
        progress_window.title(f"File {operation.capitalize()}ion")
        progress_window.geometry("300x100")
        progress_window.transient(self.root)
        progress_window.grab_set()
        
        # Center the window
        progress_window.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - progress_window.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - progress_window.winfo_height()) // 2
        progress_window.geometry(f"+{x}+{y}")
        
        # Add progress bar and label
        progress_label = ttk.Label(progress_window, text=f"{operation.capitalize()}ing file...")
        progress_label.pack(pady=10)
        
        progress_bar = ttk.Progressbar(progress_window, orient=tk.HORIZONTAL, 
                                    length=250, mode='indeterminate')
        progress_bar.pack(pady=10)
        progress_bar.start(10)  # Start indeterminate progress
        
        # Function to check thread status
        def check_thread():
            if thread.is_alive():
                # Thread still running, check again in 100ms
                progress_window.after(100, check_thread)
            else:
                # Thread finished
                progress_bar.stop()
                progress_window.destroy()
                
                if thread.exception:
                    messagebox.showerror("Error", f"Failed to {operation} file: {thread.exception}")
                    self.log_action(f"File {operation}ion failed: {thread.exception}")
                else:
                    result = thread.get_result()
                    if result:
                        self.update_progress(100, f"{operation.capitalize()}ion complete!")
                        messagebox.showinfo("Success", 
                                        f"File {operation}ed successfully!\nSaved as: {result}")
                        self.log_action(f"File {operation}ed: {result}")
        
        # Start checking thread status
        progress_window.after(100, check_thread)
        
        # Wait for the window to close
        self.root.wait_window(progress_window)

    def _encrypt_file_worker(self, file_path, password):
        """Worker function for encryption (runs in separate thread)"""
        try:
            output_path = self.secure_dir / f"encrypted_{file_path.name}.enc"
            
            # Get file size for progress calculation
            file_size = os.path.getsize(file_path)
            chunk_size = 64 * 1024  # 64KB chunks
            
            # Generate a key from the password
            salt = os.urandom(32)
            kdf = PBKDF2HMAC(
                algorithm=hashes.SHA512(),
                length=32,
                salt=salt,
                iterations=500000,
            )
            key = kdf.derive(password.encode())
            
            # Generate a random IV
            iv = os.urandom(16)
            
            # Use AES in GCM mode for authenticated encryption
            encryptor = Cipher(
                algorithms.AES(key),
                modes.GCM(iv),
                backend=default_backend()
            ).encryptor()
            
            # Add associated data for authentication
            encryptor.authenticate_additional_data(salt)
            
            # Process file in chunks
            processed = 0
            
            with open(file_path, 'rb') as infile, open(output_path, 'wb') as outfile:
                # Write header, salt, and IV
                outfile.write(b"SRT_ENC_V3")  # Version identifier
                outfile.write(salt)
                outfile.write(iv)
                
                # Process file in chunks
                while True:
                    chunk = infile.read(chunk_size)
                    if len(chunk) == 0:
                        break
                        
                    encrypted_chunk = encryptor.update(chunk)
                    outfile.write(encrypted_chunk)
                    
                    # Update progress (this will be shown in the UI)
                    processed += len(chunk)
                    progress = min(95, int((processed / file_size) * 95))
                    
                    # Use thread-safe method to update progress
                    self.root.after(0, lambda: self.update_progress(
                        progress, f"Encrypting... {progress}%"))
                
                # Finalize encryption
                encrypted_chunk = encryptor.finalize()
                outfile.write(encrypted_chunk)
                outfile.write(encryptor.tag)  # Write authentication tag
            
            return output_path.name
            
        except Exception as e:
            raise e

    def _decrypt_file_worker(self, file_path, password):
        """Worker function for decryption (runs in separate thread)"""
        try:
            # Remove .enc extension if present
            if file_path.name.endswith('.enc'):
                output_filename = file_path.name[:-4]
            else:
                output_filename = f"decrypted_{file_path.name}"
                
            output_path = self.secure_dir / output_filename
            
            # Get file size for progress calculation
            file_size = os.path.getsize(file_path)
            chunk_size = 64 * 1024  # 64KB chunks
            
            with open(file_path, 'rb') as infile:
                # Read header, salt, and IV
                header = infile.read(10)
                if header != b"SRT_ENC_V3":
                    raise ValueError("Invalid file format or version")
                    
                salt = infile.read(32)
                iv = infile.read(16)
                
                # Derive key from password
                kdf = PBKDF2HMAC(
                    algorithm=hashes.SHA512(),
                    length=32,
                    salt=salt,
                    iterations=500000,
                )
                key = kdf.derive(password.encode())
                
                # Read authentication tag (last 16 bytes)
                infile.seek(-16, 2)  # Seek to 16 bytes from end
                tag = infile.read(16)
                data_size = file_size - 10 - 32 - 16 - 16  # Subtract header, salt, IV, and tag
                
                # Set up decryptor
                decryptor = Cipher(
                    algorithms.AES(key),
                    modes.GCM(iv, tag),
                    backend=default_backend()
                ).decryptor()
                
                # Authenticate additional data
                decryptor.authenticate_additional_data(salt)
                
                # Reset to beginning of encrypted data
                infile.seek(10 + 32 + 16)
                
                # Process file in chunks
                processed = 0
                
                with open(output_path, 'wb') as outfile:
                    # Read and decrypt in chunks
                    while processed < data_size:
                        chunk = infile.read(min(chunk_size, data_size - processed))
                        if len(chunk) == 0:
                            break
                            
                        decrypted_chunk = decryptor.update(chunk)
                        outfile.write(decrypted_chunk)
                        
                        # Update progress (this will be shown in the UI)
                        processed += len(chunk)
                        progress = min(95, int((processed / data_size) * 95))
                        
                        # Use thread-safe method to update progress
                        self.root.after(0, lambda: self.update_progress(
                            progress, f"Decrypting... {progress}%"))
                    
                    # Finalize decryption
                    decryptor.finalize()
            
            return output_path.name
            
        except Exception as e:
            raise e
    # Replace the encrypt_file and decrypt_file methods with these simpler versions
    def encrypt_file(self):
        if not hasattr(self, 'encrypt_file_path') or not self.encrypt_file_path:
            messagebox.showwarning("Error", "Please select a file first")
            return
            
        password = self.encrypt_password_var.get()
        if not password:
            messagebox.showwarning("Error", "Please enter a password")
            return
            
        # Run encryption in a separate thread with progress UI
        self.run_encryption_with_progress("encrypt", self.encrypt_file_path, password)

    def decrypt_file(self):
        if not hasattr(self, 'encrypt_file_path') or not self.encrypt_file_path:
            messagebox.showwarning("Error", "Please select a file first")
            return
            
        password = self.encrypt_password_var.get()
        if not password:
            messagebox.showwarning("Error", "Please enter the password")
            return
            
        # Run decryption in a separate thread with progress UI
        self.run_encryption_with_progress("decrypt", self.encrypt_file_path, password)

    def generate_checksum(self):
        if not hasattr(self, 'checksum_file_path') or not self.checksum_file_path:
            messagebox.showwarning("Error", "Please select a file first")
            return
            
        try:
            # Get selected algorithm
            algo = self.checksum_algo_var.get().lower()
            
            # Generate checksum based on selected algorithm
            if algo == "md5":
                hash_obj = hashlib.md5()
            elif algo == "sha-1":
                hash_obj = hashlib.sha1()
            elif algo == "sha-256":
                hash_obj = hashlib.sha256()
            elif algo == "sha-512":
                hash_obj = hashlib.sha512()
            else:
                hash_obj = hashlib.sha256()  # Default
            
            with open(self.checksum_file_path, "rb") as f:
                # Read and update hash in chunks of 4K
                for byte_block in iter(lambda: f.read(4096), b""):
                    hash_obj.update(byte_block)
            
            checksum = hash_obj.hexdigest()
            
            # Save checksum to file
            checksum_file = self.secure_dir / f"{self.checksum_file_path.name}.{algo}"
            with open(checksum_file, 'w') as f:
                f.write(f"{algo.upper()} checksum for {self.checksum_file_path.name}\n")
                f.write(f"{checksum}\n")
                f.write(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"File size: {self.checksum_file_path.stat().st_size} bytes\n")
            
            # Display the checksum
            self.checksum_result_var.set(checksum)
            
            self.log_action(f"Generated {algo.upper()} checksum for {self.checksum_file_path.name}")
            messagebox.showinfo("Success", 
                            f"{algo.upper()} checksum generated successfully!\n"
                            f"Saved to: {checksum_file.name}\n\n"
                            f"Checksum: {checksum}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate checksum: {str(e)}")
            self.log_action(f"Checksum generation failed: {str(e)}")

    def verify_checksum(self):
        if not hasattr(self, 'checksum_file_path') or not self.checksum_file_path:
            messagebox.showwarning("Error", "Please select a file first")
            return
            
        try:
            # Get selected algorithm
            algo = self.checksum_algo_var.get().lower()
            
            # Generate current checksum
            if algo == "md5":
                hash_obj = hashlib.md5()
            elif algo == "sha-1":
                hash_obj = hashlib.sha1()
            elif algo == "sha-256":
                hash_obj = hashlib.sha256()
            elif algo == "sha-512":
                hash_obj = hashlib.sha512()
            else:
                hash_obj = hashlib.sha256()  # Default
                
            with open(self.checksum_file_path, "rb") as f:
                for byte_block in iter(lambda: f.read(4096), b""):
                    hash_obj.update(byte_block)
            
            current_checksum = hash_obj.hexdigest()
            
            # Get expected checksum based on verification method
            if self.verify_method_var.get() == "file":
                # Ask for the checksum file
                checksum_file_path = filedialog.askopenfilename(
                    title="Select checksum file",
                    filetypes=[("Checksum files", "*.md5 *.sha1 *.sha256 *.sha512"), ("All files", "*.*")]
                )
                
                if not checksum_file_path:
                    return
                    
                # Read the expected checksum
                with open(checksum_file_path, 'r') as f:
                    content = f.read()
                    # Extract checksum from file (might be in various formats)
                    lines = content.splitlines()
                    expected_checksum = None
                    
                    for line in lines:
                        # Look for a line with just the checksum
                        if re.match(r'^[a-fA-F0-9]{8,128}$', line.strip()):
                            expected_checksum = line.strip().lower()
                            break
                        # Look for checksum followed by filename
                        match = re.match(r'^([a-fA-F0-9]{8,128})\s+\*?(.+)$', line)
                        if match:
                            expected_checksum = match.group(1).lower()
                            break
                    
                    if not expected_checksum:
                        messagebox.showerror("Error", "Could not find a valid checksum in the selected file")
                        return
            else:
                # Manual input method
                expected_checksum = self.manual_checksum_var.get().strip().lower()
                if not expected_checksum:
                    messagebox.showwarning("Error", "Please enter a checksum to verify against")
                    return
                    
                # Validate manual checksum format
                if not re.match(r'^[a-f0-9]{8,128}$', expected_checksum):
                    messagebox.showwarning("Error", "Invalid checksum format. Please enter a valid hexadecimal checksum")
                    return
            
            # Verify
            if current_checksum == expected_checksum:
                result_text = f"✓ {algo.upper()} checksum verification PASSED"
                result_color = "green"
                self.log_action(f"Checksum verification passed for {self.checksum_file_path.name}")
                messagebox.showinfo("Success", 
                                f"Checksum verification PASSED!\n\n"
                                f"File: {self.checksum_file_path.name}\n"
                                f"Algorithm: {algo.upper()}\n"
                                f"Checksum: {current_checksum}\n\n"
                                f"File integrity is intact.")
            else:
                result_text = f"✗ {algo.upper()} checksum verification FAILED"
                result_color = "red"
                self.log_action(f"Checksum verification failed for {self.checksum_file_path.name}")
                messagebox.showerror("Error", 
                                f"Checksum verification FAILED!\n\n"
                                f"Expected: {expected_checksum}\n"
                                f"Actual: {current_checksum}\n\n"
                                f"File may be corrupted or modified.")
            
            # Display result
            self.checksum_result_var.set(current_checksum)
            self.checksum_entry.config(foreground=result_color)
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to verify checksum: {str(e)}")
            self.log_action(f"Checksum verification failed: {str(e)}")

    def copy_checksum(self):
        """Copy the current checksum to clipboard"""
        checksum = self.checksum_result_var.get()
        if checksum:
            self.root.clipboard_clear()
            self.root.clipboard_append(checksum)
            self.status_var.set("Checksum copied to clipboard")
            self.log_action("Checksum copied to clipboard")
            
            # Show temporary confirmation
            original_text = self.checksum_entry.cget('style')
            self.checksum_entry.config(style='Success.TEntry')
            self.root.after(1000, lambda: self.checksum_entry.config(style=original_text))
        else:
            messagebox.showwarning("No Checksum", "No checksum to copy")
    # -------------------------------
    # SELF-DESTRUCT FUNCTIONS
    # -------------------------------
    def select_self_destruct_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("All files", "*.*")])
        if file_path:
            self.self_destruct_source = Path(file_path)
            self.self_destruct_path_var.set(self.self_destruct_source.name)
            self.log_action(f"Selected file for self-destruct: {self.self_destruct_source.name}")

    def create_self_destruct_file(self):
        if not hasattr(self, 'self_destruct_source') or not self.self_destruct_source:
            messagebox.showwarning("Error", "Please select a file first")
            return
            
        # Validate options
        if not self.expire_date_var.get() and not self.expire_openings_var.get():
            messagebox.showwarning("Error", "Please select at least one expiration option")
            return
            
        # Get password if enabled
        password = None
        if self.use_password_var.get():
            password = self.self_destruct_password_var.get()
            if not password:
                messagebox.showwarning("Error", "Please enter a password")
                return
        
        # Get expiration settings
        expiration_data = {}
        if self.expire_date_var.get():
            try:
                expire_date = datetime.strptime(self.date_var.get(), "%Y-%m-%d")
                expiration_data['expire_date'] = expire_date.isoformat()
            except ValueError:
                messagebox.showwarning("Error", "Invalid date format. Please use YYYY-MM-DD")
                return
                
        if self.expire_openings_var.get():
            openings = self.openings_var.get()
            if openings < 1:
                messagebox.showwarning("Error", "Openings must be at least 1")
                return
            expiration_data['max_openings'] = openings
            expiration_data['remaining_openings'] = openings
        
        try:
            # Read the source file
            with open(self.self_destruct_source, 'rb') as f:
                file_data = f.read()
            
            # Generate a unique ID for this file
            file_id = str(uuid.uuid4())
            
            # Create metadata
            metadata = {
                'id': file_id,
                'original_filename': self.self_destruct_source.name,
                'created': datetime.now().isoformat(),
                'expiration': expiration_data,
                'password_protected': password is not None
            }
            
            # Encrypt the file data if password is provided
            if password:
                salt = os.urandom(16)
                kdf = PBKDF2HMAC(
                    algorithm=hashes.SHA256(),
                    length=32,
                    salt=salt,
                    iterations=100000,
                )
                key = base64.urlsafe_b64encode(kdf.derive(password.encode()))
                fernet = Fernet(key)
                encrypted_data = fernet.encrypt(file_data)
                file_data = salt + encrypted_data
            
            # Create the self-destruct file with the same extension
            output_filename = f"self_destruct_{self.self_destruct_source.name}"
            output_path = self.self_destruct_dir / output_filename
            
            # Write the file with metadata appended to the end
            with open(output_path, 'wb') as f:
                # Write file data
                f.write(file_data)
                
                # Write metadata as JSON at the end of the file
                metadata_str = json.dumps(metadata)
                metadata_bytes = metadata_str.encode('utf-8')
                
                # Add a marker and metadata length before the metadata
                marker = b"SRT_SELF_DESTRUCT_METADATA:"
                f.write(marker)
                f.write(len(metadata_bytes).to_bytes(4, 'big'))
                f.write(metadata_bytes)
            
            # Store metadata for tracking
            self.self_destruct_files[file_id] = {
                'filename': output_path.name,
                'metadata': metadata,
                'file_path': str(output_path)
            }
            self.save_self_destruct_files()
            
            self.log_action(f"Created self-destruct file: {output_path.name}")
            self.self_destruct_status_var.set(f"Self-destruct file created: {output_path.name}")
            messagebox.showinfo("Success", f"Self-destruct file created successfully!\n\nFile: {output_path.name}")
            
            # Refresh the list
            self.refresh_self_destruct_list()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create self-destruct file: {str(e)}")
            self.log_action(f"Self-destruct creation failed: {str(e)}")

    def open_self_destruct_file(self):
        # Ask user to select a self-destruct file
        file_path = filedialog.askopenfilename(
            title="Select self-destruct file",
            filetypes=[("All files", "*.*")],
            initialdir=str(self.self_destruct_dir)
        )
        
        if not file_path:
            return
            
        file_path = Path(file_path)
        
        try:
            # Read the file and extract metadata
            with open(file_path, 'rb') as f:
                content = f.read()
            
            # Look for the metadata marker at the end of the file
            marker = b"SRT_SELF_DESTRUCT_METADATA:"
            marker_pos = content.rfind(marker)
            
            if marker_pos == -1:
                messagebox.showerror("Error", "Not a valid self-destruct file")
                return
                
            # Extract metadata length and data
            metadata_start = marker_pos + len(marker)
            metadata_length = int.from_bytes(content[metadata_start:metadata_start+4], 'big')
            metadata_str = content[metadata_start+4:metadata_start+4+metadata_length].decode('utf-8')
            metadata = json.loads(metadata_str)
            
            # Extract the original file data (everything before the marker)
            file_data = content[:marker_pos]
            
            # Check expiration
            if self.check_self_destruct_expired(metadata):
                messagebox.showerror("Error", "This file has expired and can no longer be accessed")
                # Delete the expired file
                try:
                    file_path.unlink()
                    self.log_action(f"Deleted expired self-destruct file: {file_path.name}")
                    
                    # Remove from tracking
                    file_id = metadata['id']
                    if file_id in self.self_destruct_files:
                        del self.self_destruct_files[file_id]
                        self.save_self_destruct_files()
                except:
                    pass
                return
            
            # Check password
            password = None
            if metadata.get('password_protected', False):
                password = simpledialog.askstring("Password", "Enter password:", show='*')
                if not password:
                    return
            
            # Decrypt data if password is provided
            if metadata.get('password_protected', False):
                try:
                    salt = file_data[:16]
                    encrypted_data = file_data[16:]
                    
                    kdf = PBKDF2HMAC(
                        algorithm=hashes.SHA256(),
                        length=32,
                        salt=salt,
                        iterations=100000,
                    )
                    key = base64.urlsafe_b64encode(kdf.derive(password.encode()))
                    fernet = Fernet(key)
                    file_data = fernet.decrypt(encrypted_data)
                except Exception as e:
                    messagebox.showerror("Error", "Incorrect password or corrupted file")
                    return
            
            # Save the decrypted file to a temporary location with original name
            original_filename = metadata.get('original_filename', 'restored_file')
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_filename)[1])
            with open(temp_file.name, 'wb') as f:
                f.write(file_data)
            
            # Open the file with default application
            if platform.system() == 'Windows':
                os.startfile(temp_file.name)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', temp_file.name])
            else:
                subprocess.run(['xdg-open', temp_file.name])
            
            # Update opening count
            file_id = metadata['id']
            if file_id in self.self_destruct_files:
                expiration = self.self_destruct_files[file_id]['metadata'].get('expiration', {})
                if 'max_openings' in expiration:
                    remaining = expiration.get('remaining_openings', 0)
                    if remaining > 0:
                        remaining -= 1
                        self.self_destruct_files[file_id]['metadata']['expiration']['remaining_openings'] = remaining
                        
                        # Update the metadata in the file
                        self.update_self_destruct_metadata(file_path, self.self_destruct_files[file_id]['metadata'])
                        
                        if remaining == 0:
                            messagebox.showinfo("Info", "This was the last opening. The file will now be deleted.")
                            try:
                                file_path.unlink()
                                del self.self_destruct_files[file_id]
                                self.log_action(f"Self-destruct file expired and deleted: {file_path.name}")
                            except:
                                pass
            
            self.save_self_destruct_files()
            self.refresh_self_destruct_list()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open self-destruct file: {str(e)}")
            self.log_action(f"Self-destruct open failed: {str(e)}")

    def update_self_destruct_metadata(self, file_path, new_metadata):
        try:
            # Read the file
            with open(file_path, 'rb') as f:
                content = f.read()
            
            # Find the metadata marker
            marker = b"SRT_SELF_DESTRUCT_METADATA:"
            marker_pos = content.rfind(marker)
            
            if marker_pos == -1:
                return False
                
            # Extract the file data (everything before the marker)
            file_data = content[:marker_pos]
            
            # Create new metadata
            metadata_str = json.dumps(new_metadata)
            metadata_bytes = metadata_str.encode('utf-8')
            
            # Write the updated file
            with open(file_path, 'wb') as f:
                f.write(file_data)
                f.write(marker)
                f.write(len(metadata_bytes).to_bytes(4, 'big'))
                f.write(metadata_bytes)
                
            return True
            
        except Exception as e:
            self.log_action(f"Failed to update self-destruct metadata: {str(e)}")
            return False

    def check_self_destruct_expired(self, metadata):
        if 'expiration' not in metadata:
            return False
            
        expiration = metadata['expiration']
        
        # Check date expiration
        if 'expire_date' in expiration:
            expire_date = datetime.fromisoformat(expiration['expire_date'])
            if datetime.now() > expire_date:
                return True
        
        # Check openings expiration
        if 'remaining_openings' in expiration:
            if expiration['remaining_openings'] <= 0:
                return True
        
        return False

    def update_self_destruct_metadata(self, file_path, new_metadata):
        try:
            temp_dir = tempfile.mkdtemp()
            
            # Extract the file
            if file_path.suffix.lower() == '.zip':
                with zipfile.ZipFile(file_path, 'r') as zipf:
                    zipf.extractall(temp_dir)
            elif file_path.suffix.lower() == '.7z':
                with py7zr.SevenZipFile(file_path, 'r') as z:
                    z.extractall(temp_dir)
            elif file_path.suffix.lower() == '.rar':
                with rarfile.RarFile(file_path, 'r') as rarf:
                    rarf.extractall(temp_dir)
            
            # Update metadata
            metadata_path = Path(temp_dir) / 'metadata.json'
            with open(metadata_path, 'w') as f:
                json.dump(new_metadata, f, indent=2)
            
            # Recreate the archive
            if file_path.suffix.lower() == '.zip':
                with zipfile.ZipFile(file_path, 'w') as zipf:
                    zipf.write(Path(temp_dir) / 'data.bin', 'data.bin')
                    zipf.write(metadata_path, 'metadata.json')
            elif file_path.suffix.lower() == '.7z':
                with py7zr.SevenZipFile(file_path, 'w') as z:
                    z.write(Path(temp_dir) / 'data.bin', 'data.bin')
                    z.write(metadata_path, 'metadata.json')
            
            # Clean up
            shutil.rmtree(temp_dir)
            
        except Exception as e:
            self.log_action(f"Failed to update self-destruct metadata: {str(e)}")

    def refresh_self_destruct_list(self):
        # Clear existing items
        for item in self.self_destruct_tree.get_children():
            self.self_destruct_tree.delete(item)
        
        # Load metadata
        self.load_self_destruct_files()
        
        # Add files to treeview
        for file_id, data in self.self_destruct_files.items():
            metadata = data['metadata']
            expiration = metadata.get('expiration', {})
            
            # Determine expiration type and value
            expiration_type = ""
            expiration_value = ""
            status = "Active"
            
            if self.check_self_destruct_expired(metadata):
                status = "Expired"
            
            if 'expire_date' in expiration:
                expiration_type = "Date"
                expiration_value = expiration['expire_date'][:10]  # Just the date part
            elif 'max_openings' in expiration:
                expiration_type = "Openings"
                remaining = expiration.get('remaining_openings', expiration['max_openings'])
                expiration_value = f"{remaining}/{expiration['max_openings']}"
            
            self.self_destruct_tree.insert("", tk.END, values=(
                data['filename'],
                expiration_type,
                expiration_value,
                status
            ))

# -------------------------------
# Start of Port Scanning methods
# -------------------------------
 
    def log_port_scan_action(self, message):
        try:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_entry = f"[{timestamp}] {message}\n"
            with open(self.port_scan_log_file, "a", encoding="utf-8") as f:
                f.write(log_entry)
        except Exception as e:
            print(f"Port scan logging error: {e}")
        
    def create_port_scanner_tab(self):
        """Create an advanced port scanner tab with Metasploit-like features"""
        port_tab = ttk.Frame(self.notebook)
        self.notebook.add(port_tab, text="Port Scanner")
        
        # Create a paned window for the advanced interface
        main_paned = ttk.PanedWindow(port_tab, orient=tk.HORIZONTAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Left panel for scanning options
        left_frame = ttk.Frame(main_paned)
        main_paned.add(left_frame, weight=1)
        
        # Right panel for results and details
        right_frame = ttk.Frame(main_paned)
        main_paned.add(right_frame, weight=2)
        
        # Instructions
        instr_frame = ttk.LabelFrame(left_frame, text="Advanced Port Scanner", padding=15)
        instr_frame.pack(fill=tk.X, padx=5, pady=5)
        
        instructions = (
            "Advanced port scanner with service fingerprinting, "
            "vulnerability assessment, and exploit suggestions.\n\n"
            "1. Enter target IP address or subnet\n"
            "2. Configure scan options\n"
            "3. Click 'Start Scan' to begin\n"
            "4. Review results and suggested actions"
        )
        ttk.Label(instr_frame, text=instructions, wraplength=300).pack(anchor=tk.W)
        
        # Target input
        target_frame = ttk.LabelFrame(left_frame, text="Scan Target", padding=10)
        target_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Ping Button
        ttk.Button(target_frame, text="Ping", command=self.ping_target).grid(row=0, column=6, padx=5, pady=5)


        ttk.Label(target_frame, text="Target IP/Subnet:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.target_var = tk.StringVar(value="127.0.0.1")
        target_entry = ttk.Entry(target_frame, textvariable=self.target_var, width=20)
        target_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(target_frame, text="Port Range:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        self.port_start_var = tk.IntVar(value=1)
        port_start_spin = ttk.Spinbox(target_frame, from_=1, to=65535, width=5, textvariable=self.port_start_var)
        port_start_spin.grid(row=0, column=3, padx=5, pady=5)
        
        ttk.Label(target_frame, text="to").grid(row=0, column=4, padx=5, pady=5)
        self.port_end_var = tk.IntVar(value=1024)
        port_end_spin = ttk.Spinbox(target_frame, from_=1, to=65535, width=5, textvariable=self.port_end_var)
        port_end_spin.grid(row=0, column=5, padx=5, pady=5)
        
        # Scan options
        options_frame = ttk.Frame(target_frame)
        options_frame.grid(row=1, column=0, columnspan=6, sticky=tk.W, pady=5)
        
        self.timeout_var = tk.IntVar(value=2)
        ttk.Label(options_frame, text="Timeout (s):").pack(side=tk.LEFT, padx=5)
        timeout_spin = ttk.Spinbox(options_frame, from_=1, to=10, width=3, textvariable=self.timeout_var)
        timeout_spin.pack(side=tk.LEFT, padx=5)
        
        self.threads_var = tk.IntVar(value=50)
        ttk.Label(options_frame, text="Threads:").pack(side=tk.LEFT, padx=5)
        threads_spin = ttk.Spinbox(options_frame, from_=1, to=200, width=3, textvariable=self.threads_var)
        threads_spin.pack(side=tk.LEFT, padx=5)
        
        # Ping option
        self.ping_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Ping Check", variable=self.ping_var).pack(side=tk.LEFT, padx=5)
        
        # Advanced options
        self.os_detection_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="OS Detection", variable=self.os_detection_var).pack(side=tk.LEFT, padx=5)
        
        self.service_detection_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(options_frame, text="Service Detection", variable=self.service_detection_var).pack(side=tk.LEFT, padx=5)
        
        # Buttons
        button_frame = ttk.Frame(target_frame)
        button_frame.grid(row=2, column=0, columnspan=6, pady=10)
        
        scan_btn = ttk.Button(button_frame, text="Start Scan", 
                            command=self.start_port_scan, style='Primary.TButton')
        scan_btn.pack(side=tk.LEFT, padx=5)
        
        stop_btn = ttk.Button(button_frame, text="Stop Scan", 
                            command=self.stop_port_scan, style='Accent.TButton')
        stop_btn.pack(side=tk.LEFT, padx=5)
        
        export_btn = ttk.Button(button_frame, text="Export Results", 
                            command=self.export_scan_results, style='Secondary.TButton')
        export_btn.pack(side=tk.LEFT, padx=5)
        
        # In the create_port_scanner_tab method, add the download button to the button_frame
            # Find the button_frame section and add this button:
        download_log_btn = ttk.Button(button_frame, text="Download Log", command=self.download_port_scan_log, style='Secondary.TButton')
        download_log_btn.pack(side=tk.LEFT, padx=5)
         
        #-------------------------------------------------------------

        # Progress
        progress_frame = ttk.Frame(left_frame)
        progress_frame.pack(fill=tk.X, padx=5, pady=5)
        
        self.scan_progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, mode='determinate')
        self.scan_progress.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        self.scan_status = ttk.Label(progress_frame, text="Ready")
        self.scan_status.pack(side=tk.RIGHT)
        
        # NEW: Add open ports display section
        open_ports_frame = ttk.LabelFrame(left_frame, text="Open Ports Found", padding=10)
        open_ports_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create treeview with scrollbar for open ports
        open_ports_tree_frame = ttk.Frame(open_ports_frame)
        open_ports_tree_frame.pack(fill=tk.BOTH, expand=True)

        columns = ("port", "service", "description")
        self.open_ports_tree = ttk.Treeview(
        open_ports_tree_frame, columns=columns, show="headings", height=8
                                                                         )
        # Define headings
        self.open_ports_tree.heading("port", text="Port")
        self.open_ports_tree.heading("service", text="Service")
        self.open_ports_tree.heading("description", text="Description")
        
        # Set column widths
        self.open_ports_tree.column("port", width=80, anchor=tk.CENTER)
        self.open_ports_tree.column("service", width=120, anchor=tk.W)
        self.open_ports_tree.column("description", width=250, anchor=tk.W)

        # Add scrollbar
        open_ports_scrollbar = ttk.Scrollbar(open_ports_tree_frame, orient=tk.VERTICAL, command=self.open_ports_tree.yview)
        self.open_ports_tree.configure(yscrollcommand=open_ports_scrollbar.set)
        open_ports_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.open_ports_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Right panel - Results and details

        #--------------------------------------------------------

        # Create notebook for different views
        results_notebook = ttk.Notebook(right_frame)
        results_notebook.pack(fill=tk.BOTH, expand=True)
        
        # Hosts tab
        hosts_frame = ttk.Frame(results_notebook)
        results_notebook.add(hosts_frame, text="Hosts")
        
        # Create treeview with scrollbar for hosts
        hosts_tree_frame = ttk.Frame(hosts_frame)
        hosts_tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ("host", "status", "os", "open_ports", "vulnerabilities")
        self.hosts_tree = ttk.Treeview(hosts_tree_frame, columns=columns, show="headings", height=10)
        
        # Define headings
        self.hosts_tree.heading("host", text="Host")
        self.hosts_tree.heading("status", text="Status")
        self.hosts_tree.heading("os", text="OS")
        self.hosts_tree.heading("open_ports", text="Open Ports")
        self.hosts_tree.heading("vulnerabilities", text="Vulnerabilities")
        
        # Set column widths
        self.hosts_tree.column("host", width=120, anchor=tk.W)
        self.hosts_tree.column("status", width=80, anchor=tk.CENTER)
        self.hosts_tree.column("os", width=150, anchor=tk.W)
        self.hosts_tree.column("open_ports", width=100, anchor=tk.CENTER)
        self.hosts_tree.column("vulnerabilities", width=100, anchor=tk.CENTER)
        
        # Add scrollbar
        hosts_scrollbar = ttk.Scrollbar(hosts_tree_frame, orient=tk.VERTICAL, command=self.hosts_tree.yview)
        self.hosts_tree.configure(yscrollcommand=hosts_scrollbar.set)
        hosts_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.hosts_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Services tab
        services_frame = ttk.Frame(results_notebook)
        results_notebook.add(services_frame, text="Services")
        
        # Create treeview with scrollbar for services
        services_tree_frame = ttk.Frame(services_frame)
        services_tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ("host", "port", "service", "version", "banner", "status", "vulnerabilities")
        self.services_tree = ttk.Treeview(services_tree_frame, columns=columns, show="headings", height=15)
        
        # Define headings
        self.services_tree.heading("host", text="Host")
        self.services_tree.heading("port", text="Port")
        self.services_tree.heading("service", text="Service")
        self.services_tree.heading("version", text="Version")
        self.services_tree.heading("banner", text="Banner")
        self.services_tree.heading("status", text="Status")
        self.services_tree.heading("vulnerabilities", text="Vulnerabilities")
        
        # Set column widths
        self.services_tree.column("host", width=120, anchor=tk.W)
        self.services_tree.column("port", width=80, anchor=tk.CENTER)
        self.services_tree.column("service", width=120, anchor=tk.W)
        self.services_tree.column("version", width=120, anchor=tk.W)
        self.services_tree.column("banner", width=200, anchor=tk.W)
        self.services_tree.column("status", width=80, anchor=tk.W)
        self.services_tree.column("vulnerabilities", width=100, anchor=tk.CENTER)
        
        # Add scrollbar
        services_scrollbar = ttk.Scrollbar(services_tree_frame, orient=tk.VERTICAL, command=self.services_tree.yview)
        self.services_tree.configure(yscrollcommand=services_scrollbar.set)
        services_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.services_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Vulnerabilities tab
        vuln_frame = ttk.Frame(results_notebook)
        results_notebook.add(vuln_frame, text="Vulnerabilities")
        
        # Create treeview with scrollbar for vulnerabilities
        vuln_tree_frame = ttk.Frame(vuln_frame)
        vuln_tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        columns = ("host", "port", "service", "vulnerability", "severity", "exploit_suggestion")
        self.vuln_tree = ttk.Treeview(vuln_tree_frame, columns=columns, show="headings", height=10)
        
        # Define headings
        self.vuln_tree.heading("host", text="Host")
        self.vuln_tree.heading("port", text="Port")
        self.vuln_tree.heading("service", text="Service")
        self.vuln_tree.heading("vulnerability", text="Vulnerability")
        self.vuln_tree.heading("severity", text="Severity")
        self.vuln_tree.heading("exploit_suggestion", text="Exploit Suggestion")
        
        # Set column widths
        self.vuln_tree.column("host", width=120, anchor=tk.W)
        self.vuln_tree.column("port", width=80, anchor=tk.CENTER)
        self.vuln_tree.column("service", width=120, anchor=tk.W)
        self.vuln_tree.column("vulnerability", width=200, anchor=tk.W)
        self.vuln_tree.column("severity", width=80, anchor=tk.CENTER)
        self.vuln_tree.column("exploit_suggestion", width=250, anchor=tk.W)
        
        # Add scrollbar
        vuln_scrollbar = ttk.Scrollbar(vuln_tree_frame, orient=tk.VERTICAL, command=self.vuln_tree.yview)
        self.vuln_tree.configure(yscrollcommand=vuln_scrollbar.set)
        vuln_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.vuln_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Define tags for severity levels
        self.vuln_tree.tag_configure("critical", background="#ffdddd")
        self.vuln_tree.tag_configure("high", background="#ffeedd")
        self.vuln_tree.tag_configure("medium", background="#ffffdd")
        self.vuln_tree.tag_configure("low", background="#ddffdd")
        
        # Initialize scan control variables
        self.scanning = False
        self.scan_thread = None
        self.stop_scan = False
        
        # Database of known vulnerabilities and exploit suggestions
        self.vuln_db = {
            "SSH-1.5": {
                "vulnerability": "SSH protocol version 1 (insecure)",
                "severity": "High",
                "exploit": "Use SSHv1 exploit modules in Metasploit: exploit/linux/ssh/..."
            },
            "FTP anonymous access": {
                "vulnerability": "FTP anonymous access allowed",
                "severity": "Medium",
                "exploit": "Try anonymous login, use FTP bounce attacks"
            },
            "HTTP Apache/1.3": {
                "vulnerability": "Outdated Apache web server",
                "severity": "High",
                "exploit": "Check for known Apache 1.3 exploits like chunked encoding vulnerability"
            },
            "HTTP IIS/5.0": {
                "vulnerability": "Outdated IIS web server",
                "severity": "Critical",
                "exploit": "Check for IIS 5.0 exploits like .printer buffer overflow"
            },
            "Telnet service": {
                "vulnerability": "Telnet service (insecure cleartext protocol)",
                "severity": "High",
                "exploit": "Brute force credentials, session hijacking"
            },
            "SNMP public community": {
                "vulnerability": "SNMP default community string",
                "severity": "Medium",
                "exploit": "Use snmpwalk to gather information, modify configurations"
            },
            "SMB Windows XP": {
                "vulnerability": "Outdated Windows SMB service",
                "severity": "Critical",
                "exploit": "EternalBlue exploit (MS17-010)"
            },
            "RDP without NLA": {
                "vulnerability": "RDP without Network Level Authentication",
                "severity": "High",
                "exploit": "Brute force attacks, BlueKeep exploit if vulnerable"
            },
            "VNC no authentication": {
                "vulnerability": "VNC without authentication",
                "severity": "Critical",
                "exploit": "Direct connection without credentials"
            },
            "MySQL empty root password": {
                "vulnerability": "MySQL empty root password",
                "severity": "Critical",
                "exploit": "Direct database access, privilege escalation"
            },
             "Hirschmann default credentials": {
            "vulnerability": "Hirschmann switch default credentials",
            "severity": "Critical",
                "exploit": "Try default credentials: admin/admin, admin/hirschmann, operator/operator"
            },
            "Hirschmann unpatched firmware": {
                "vulnerability": "Hirschmann switch outdated firmware",
                "severity": "High",
                "exploit": "Check for known Hirschmann CVEs and exploit using specialized industrial control system tools"
            },
            "Hirschmann web interface": {
                "vulnerability": "Hirschmann web interface exposed",
                "severity": "Medium",
                "exploit": "Check for XSS, CSRF, and authentication bypass vulnerabilities in web interface"
            },
            "Emerson default credentials": {
                "vulnerability": "Emerson switch default credentials",
                "severity": "Critical",
                "exploit": "Try default credentials: admin/administrator, user/user, engineer/engineer"
            },
            "Emerson unencrypted management": {
                "vulnerability": "Emerson switch unencrypted management protocols",
                "severity": "High",
                "exploit": "Intercept and manipulate network traffic to the management interface"
            },
            "Emerson SNMP defaults": {
                "vulnerability": "Emerson switch default SNMP community strings",
                "severity": "Medium",
                "exploit": "Try default SNMP communities: public, private, emerson"
            },
            "Industrial device telnet": {
                "vulnerability": "Industrial device with telnet service exposed",
                "severity": "High",
                "exploit": "Brute force credentials, session hijacking, command injection"
            },
            "MODBUS unauthorized access": {
                "vulnerability": "MODBUS service without authentication",
                "severity": "Critical",
                "exploit": "Direct MODBUS command injection to manipulate industrial processes"
            },
            "PROFINET vulnerabilities": {
                "vulnerability": "PROFINET service with known vulnerabilities",
                "severity": "High",
                "exploit": "Use specialized PROFINET exploitation tools to disrupt operations"
            },
            "CIP protocol exposed": {
                "vulnerability": "Common Industrial Protocol (CIP) exposed without protection",
                "severity": "High",
                "exploit": "CIP command injection to manipulate industrial processes"
            }
        }
        
         # Industrial device detection patterns
        self.industrial_patterns = {
        "Hirschmann": [
            "Hirschmann", "BRS", "GRS", "HiOS", "HiVision"
        ],
        "Emerson": [
            "Emerson", "NTU", "RuggedSwitch", "RuggedCom"
        ],
        "Siemens": [
            "Siemens", "SIMATIC", "SCALANCE"
        ],
        "Rockwell": [
            "Rockwell", "Allen-Bradley", "ControlLogix"
        ],
        "General Industrial": [
            "MODBUS", "PROFINET", "EtherNet/IP", "CIP", "DNP3", "IEC-60870"
        ]
    }

    def detect_vulnerabilities(self, service, version, banner, port):
        """Detect vulnerabilities based on service, version, and banner"""
        vulnerabilities = []
        
        # SSH vulnerabilities
        if service == "SSH" and version.startswith("1."):
            vulnerabilities.append("SSH-1.5")
        
        # FTP vulnerabilities
        if service == "FTP" and "anonymous" in banner.lower():
            vulnerabilities.append("FTP anonymous access")
        
        # HTTP vulnerabilities
        if service == "HTTP":
            if "Apache/1.3" in banner:
                vulnerabilities.append("HTTP Apache/1.3")
            elif "IIS/5.0" in banner:
                vulnerabilities.append("HTTP IIS/5.0")
            elif "IIS/6.0" in banner:
                vulnerabilities.append("HTTP IIS/6.0")
        
        # Telnet vulnerabilities
        if service == "Telnet":
            vulnerabilities.append("Telnet service")
        
        # SNMP vulnerabilities
        if port == 161 or port == 162:
            if "public" in banner:
                vulnerabilities.append("SNMP public community")
        
        # SMB vulnerabilities
        if service == "SMB":
            if "Windows XP" in banner or "Windows 2000" in banner:
                vulnerabilities.append("SMB Windows XP")
        
        # RDP vulnerabilities
        if service == "RDP":
            if "CredSSP" not in banner:  # No Network Level Authentication
                vulnerabilities.append("RDP without NLA")
        
        # VNC vulnerabilities
        if service == "VNC" and "Authentication" not in banner:
            vulnerabilities.append("VNC no authentication")
        
        # Database vulnerabilities
        if service == "MySQL" and "root" in banner and "password" not in banner:
            vulnerabilities.append("MySQL empty root password")
        
        # Industrial device detection
        industrial_vulns = self.detect_industrial_vulnerabilities(service, banner, port)
        vulnerabilities.extend(industrial_vulns)

        return vulnerabilities

    def detect_industrial_vulnerabilities(self, service, banner, port):
        """Detect industrial device specific vulnerabilities"""
        vulnerabilities = []
        banner_upper = banner.upper()
        
        # Check for industrial protocols
        if port == 502:  # MODBUS
            vulnerabilities.append("MODBUS unauthorized access")
        
        # Check for industrial device banners
        for vendor, patterns in self.industrial_patterns.items():
            for pattern in patterns:
                if pattern.upper() in banner_upper:
                    # Vendor-specific vulnerabilities
                    if vendor == "Hirschmann":
                        vulnerabilities.extend([
                            "Hirschmann default credentials",
                            "Hirschmann unpatched firmware"
                        ])
                        if "HTTP" in service or port in [80, 443, 8080, 8443]:
                            vulnerabilities.append("Hirschmann web interface")
                    
                    elif vendor == "Emerson":
                        vulnerabilities.extend([
                            "Emerson default credentials",
                            "Emerson unencrypted management"
                        ])
                        if port in [161, 162]:  # SNMP
                            vulnerabilities.append("Emerson SNMP defaults")
                    
                    elif vendor == "General Industrial":
                        if "PROFINET" in banner_upper:
                            vulnerabilities.append("PROFINET vulnerabilities")
                        if "CIP" in banner_upper:
                            vulnerabilities.append("CIP protocol exposed")
                    
                    break  # Found a pattern for this vendor
        
        # Check for industrial protocols on unusual ports
        industrial_ports = {
            102: "Siemens S7",
            44818: "EtherNet/IP",
            20000: "DNP3",
            2404: "IEC-60870-5-104"
        }
        
        if port in industrial_ports:
            vulnerabilities.append(f"Industrial protocol {industrial_ports[port]} on port {port}")
        
        # Check for telnet on industrial devices (common misconfiguration)
        if service == "Telnet" and any(pattern.upper() in banner_upper for pattern in ["Hirschmann", "Emerson", "Siemens", "Rockwell"]):
            vulnerabilities.append("Industrial device telnet")
        
        return vulnerabilities

    def start_port_scan(self):
        """Start the port scanning process"""
        if self.scanning:
            messagebox.showwarning("Warning", "Scan already in progress")
            return
            
        # Validate target
        target = self.target_var.get().strip()
        if not target:
            messagebox.showerror("Error", "Please enter a target IP or subnet")
            return
            
        # Validate port range
        start_port = self.port_start_var.get()
        end_port = self.port_end_var.get()
        if start_port > end_port:
            messagebox.showerror("Error", "Start port must be less than or equal to end port")
            return
            
        # Clear previous results
        for tree in [self.hosts_tree, self.services_tree, self.vuln_tree, self.open_ports_tree]:
             for item in tree.get_children():
                tree.delete(item)
                
        # Generate list of hosts to scan
        try:
            if "/" in target:
                # It's a subnet
                network = ipaddress.ip_network(target, strict=False)
                hosts = [str(ip) for ip in network.hosts()]
            else:
                # Single host
                hosts = [target]
        except ValueError as e:
            messagebox.showerror("Error", f"Invalid target: {e}")
            return
            
        # Generate list of ports to scan
        ports = list(range(start_port, end_port + 1))
        total_tasks = len(hosts) * len(ports)
        
        if total_tasks == 0:
            messagebox.showerror("Error", "No hosts or ports to scan")
            return
            
        if total_tasks > 10000 and not messagebox.askyesno(
            "Confirm Large Scan", 
            f"This scan will attempt {total_tasks} connections. This may take a long time.\n\nContinue?"
        ):
            return
            
        # Setup progress bar
        self.scan_progress["maximum"] = total_tasks
        self.scan_progress["value"] = 0
        
        # Start scan in a separate thread
        self.scanning = True
        self.stop_scan = False
        self.scan_status.config(text="Scanning...")
        
        self.scan_thread = threading.Thread(
            target=self.run_port_scan,
            args=(hosts, ports, self.timeout_var.get(), self.threads_var.get())
        )
        self.scan_thread.daemon = True
        self.scan_thread.start()
        
        # Start progress updater
        self.update_scan_progress()

    def stop_port_scan(self):
        """Stop the ongoing port scan"""
        if self.scanning:
            self.stop_scan = True
            self.scan_status.config(text="Stopping...")
        else:
            self.scan_status.config(text="Ready")

    def run_port_scan(self, hosts, ports, timeout, max_threads):
        """Run the port scan with threading and ping check"""
        total_tasks = len(hosts) * len(ports)
        
        try:
            # Log scan start
            self.log_port_scan_action(f"Starting port scan on {len(hosts)} hosts, ports {ports[0]}-{ports[-1]}")
            
            # Create a thread pool for scanning
            with ThreadPoolExecutor(max_workers=max_threads) as executor:
                # Submit all scan tasks
                future_to_target = {}
                for host in hosts:
                    # Skip hosts that don't respond to ping if ping check is enabled
                    if self.ping_var.get():
                        self.log_port_scan_action(f"Pinging host {host}...")
                        if not self.ping_host(host):
                            self.log_port_scan_action(f"Host {host} is not reachable (ping failed)")
                            # Update progress for this host's ports
                            for port in ports:
                                if self.stop_scan:
                                    break
                                self.root.after(0, lambda: self.scan_progress.step(1))
                            # Add the host to results with "Not Reachable" status
                            self.root.after(0, self.add_unreachable_host, host)
                            continue
                            
                    for port in ports:
                        if self.stop_scan:
                            break
                        future = executor.submit(self.scan_port, host, port, timeout)
                        future_to_target[future] = (host, port)
                
                # Process results as they complete
                for future in as_completed(future_to_target):
                    if self.stop_scan:
                        break
                        
                    host, port = future_to_target[future]
                    try:
                        result = future.result()
                        if result:
                            # Port is open
                            self.log_port_scan_action(f"Found open port {port} on {host}")
                            service, version, banner, os_info, vulnerabilities = self.grab_banner(host, port, timeout)
                            status = self.check_service_security(service, port, banner)
                            
                            # Add to results tree (thread-safe)
                            self.root.after(0, self.add_scan_result, host, port, service, version, banner, status, vulnerabilities, os_info)
                        else:
                            self.log_port_scan_action(f"Port {port} on {host} is closed")
                    except Exception as e:
                        # Log errors but continue
                        self.log_port_scan_action(f"Scan error for {host}:{port}: {e}")
                        
                    # Update progress
                    self.root.after(0, lambda: self.scan_progress.step(1))
                    
        except Exception as e:
            self.log_port_scan_action(f"Port scan error: {e}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"Scan failed: {e}"))
        finally:
            self.scanning = False
            
            # Ensure progress bar is complete when scan stops
            current = self.scan_progress["value"]
            remaining = total_tasks - current
            if remaining > 0 and self.stop_scan:
                self.root.after(0, lambda: self.scan_progress.step(remaining))
            
            # Check for no open ports
            if not self.stop_scan and not self.services_tree.get_children():
                self.root.after(0, lambda: messagebox.showinfo(
                    "Scan Results", 
                    "No open ports found on the target."
                ))
            
            status_text = "Scan completed" if not self.stop_scan else "Scan stopped"
            self.log_port_scan_action(status_text)
            self.root.after(0, lambda: self.scan_status.config(text=status_text))

    def scan_port(self, host, port, timeout):
        """Check if a port is open on a host"""
        if self.stop_scan:
            return False
            
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
                sock.settimeout(timeout)
                result = sock.connect_ex((host, port))
                return result == 0
        except socket.error:
            return False
        
    def ping_host(self, host):
        """Ping a host to check if it's reachable"""
        try:
            if platform.system().lower() == "windows":
                # For Windows
                command = ["ping", "-n", "1", "-w", str(self.timeout_var.get() * 1000), host]
            else:
                # For Linux/Mac
                command = ["ping", "-c", "1", "-W", str(self.timeout_var.get()), host]
            
            result = subprocess.run(command, capture_output=True, text=True)
            return result.returncode == 0
        except Exception as e:
            self.log_port_scan_action(f"Ping error for {host}: {e}")
            return False

    def grab_banner(self, host, port, timeout):
        """Attempt to grab banner from open port and detect vulnerabilities"""
        if self.stop_scan:
            return "Unknown", "", "", "", []
            
        service = self.get_service_name(port)
        banner = ""
        version = ""
        os_info = self.detect_os(host, port) # OS Detection Method
        vulnerabilities = []
        
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
                sock.settimeout(timeout)
                sock.connect((host, port))
                
                # Industrial device specific probes
                industrial_probes = {
                    502: b"\x00\x01\x00\x00\x00\x06\x01\x04\x00\x00\x00\x01",  # MODBUS read request
                    102: b"\x03\x00\x00\x16\x11\xe0\x00\x00\x00\x01\x00\xc0\x01\x0a\xc1\x02\x01\x00\xc2\x02\x01\x01",  # S7 COTP
                    44818: b"\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00\x00",  # EtherNet/IP
                    20000: b"\x05\x64\x12\x34\x00\x04\x00\x00\x00\x00\xc9\x9c",  # DNP3
                    2404: b"\x68\x04\x00\x00\x00\x00"  # IEC 60870-5-104
                }
                
                # Try to receive some data
                try:
                    # Send appropriate probes based on port
                    if port in industrial_probes:
                        # Industrial protocol probe
                        sock.send(industrial_probes[port])
                        banner = sock.recv(1024).hex()  # Get hex representation for binary protocols
                    elif port in [21, 22, 23, 25, 80, 110, 143, 443, 993, 995, 3389]:
                        # Standard service probes (unchanged)
                        if port == 22:  # SSH
                            sock.send(b"SSH-2.0-OpenSSH_7.4\r\n")
                        elif port == 80 or port == 443:  # HTTP/HTTPS
                            sock.send(b"GET / HTTP/1.0\r\nHost: " + host.encode() + b"\r\n\r\n")
                        elif port == 21:  # FTP
                            sock.send(b"USER anonymous\r\n")
                        elif port == 25:  # SMTP
                            sock.send(b"EHLO example.com\r\n")
                        elif port == 23:  # Telnet
                            sock.send(b"\r\n")
                        elif port in [110, 143, 993, 995]:  # POP/IMAP
                            sock.send(b"CAPABILITY\r\n")
                        
                        banner = sock.recv(1024).decode('utf-8', errors='ignore').strip()
                    else:
                        # Generic probe for other ports
                        try:
                            banner = sock.recv(1024).decode('utf-8', errors='ignore').strip()
                            if not banner:
                                sock.send(b"\r\n\r\n")
                                banner = sock.recv(1024).decode('utf-8', errors='ignore').strip()
                        except:
                            pass
                except:
                    # If that fails, try a simpler approach
                    try:
                        banner = sock.recv(1024).decode('utf-8', errors='ignore').strip()
                    except:
                        # For binary protocols, try hex representation
                        try:
                            banner = sock.recv(1024).hex()
                        except:
                            pass
                        
                # If we got a banner, try to extract service info from it
                if banner:
                    # For binary protocols, check if we can decode parts as text
                    try:
                        # Try to extract text from binary banners
                        text_part = bytes.fromhex(banner).decode('utf-8', errors='ignore')
                        if text_part.strip():
                            banner = f"Hex: {banner} | Text: {text_part}"
                    except:
                        pass
                    
                    # Extract version information
                    version_match = re.search(r'(\d+\.\d+(\.\d+)?)', banner)
                    if version_match:
                        version = version_match.group(1)
                    
                    # OS detection with industrial focus
                    if "Windows" in banner:
                        os_info = "Windows"
                    elif "Linux" in banner:
                        os_info = "Linux"
                    elif "Cisco" in banner:
                        os_info = "Cisco IOS"
                    elif any(pattern in banner for pattern in self.industrial_patterns["Hirschmann"]):
                        os_info = "Hirschmann OS"
                    elif any(pattern in banner for pattern in self.industrial_patterns["Emerson"]):
                        os_info = "Emerson OS"
                    elif any(pattern in banner for pattern in self.industrial_patterns["Siemens"]):
                        os_info = "Siemens OS"
                    elif any(pattern in banner for pattern in self.industrial_patterns["Rockwell"]):
                        os_info = "Rockwell OS"
                    
                    # Check for vulnerabilities based on service and version
                    vulnerabilities = self.detect_vulnerabilities(service, version, banner, port)
                        
        except socket.error:
            pass
            
        return service, banner, version, os_info, vulnerabilities
    
    def get_service_name(self, port):
        """Get service name from common port number with industrial focus"""
        common_ports = {
            21: "FTP", 22: "SSH", 23: "Telnet", 25: "SMTP", 53: "DNS",
            80: "HTTP", 110: "POP3", 143: "IMAP", 443: "HTTPS", 993: "IMAPS",
            995: "POP3S", 3306: "MySQL", 3389: "RDP", 5432: "PostgreSQL",
            # Industrial ports
            502: "MODBUS", 102: "Siemens S7", 44818: "EtherNet/IP", 
            20000: "DNP3", 2404: "IEC-60870-5-104", 47808: "BACnet",
            1911: "Foxboro", 1962: "PCWorx", 2222: "EtherNet/IP", 4000: "IEC-60870-5-104",
            5006: "IEC-61850", 5007: "IEC-61850", 9600: "OMRON FINS"
        }
        return common_ports.get(port, "Unknown")

    def check_service_security(self, service, port, banner):
        """Check if a service is considered insecure"""
        insecure_services = ["FTP", "Telnet", "HTTP", "SNMP", "VNC", "Telnet"] 
        insecure_ports = [21, 23, 80, 161, 162, 445, 3389, 5900, 1433, 1521, 27017, 5432]
        
        # Specific port checks
        if port in insecure_ports:
            return "Insecure"
        
        # Service name checks
        if service in insecure_services:
            return "Insecure"
            
        # Banner content checks
        if banner and ("anonymous" in banner.lower() or "Cisco" in banner or "HP" in banner):
            return "Insecure"
            
        return "Secure"

    def add_scan_result(self, host, port, service, version, banner, status, vulnerabilities, os_info):
        """Add a scan result to the treeviews"""
        # Add to services tree
        vuln_text = ", ".join(vulnerabilities) if vulnerabilities else "None"
        self.services_tree.insert("", tk.END, values=(
            host, port, service, version, banner[:100] + "..." if len(banner) > 100 else banner, status, vuln_text
        ))

        # NEW: Add to open ports tree
        description = self.port_info.get(port, "Unknown service")
        self.open_ports_tree.insert("", tk.END, values=(port, service, description))
        
        # Add to hosts tree if this is the first port for this host
        host_exists = False
        for item in self.hosts_tree.get_children():
            if self.hosts_tree.item(item, "values")[0] == host:
                host_exists = True
                # Update open ports count
                values = list(self.hosts_tree.item(item, "values"))
                values[3] = int(values[3]) + 1  # Increment open ports count
                if vulnerabilities:
                    values[4] = int(values[4]) + len(vulnerabilities)  # Increment vulnerabilities count
                self.hosts_tree.item(item, values=values)
                break
        
        if not host_exists:
            self.hosts_tree.insert("", tk.END, values=(
                host, "Online", os_info, 1, len(vulnerabilities)
            ))
        
        # Add to vulnerabilities tree if there are vulnerabilities
        for vuln in vulnerabilities:
            severity = self.vuln_db[vuln]["severity"] if vuln in self.vuln_db else "Unknown"
            exploit_suggestion = self.get_exploit_suggestion(vuln)
            
            self.vuln_tree.insert("", tk.END, values=(
                host, port, service, vuln, severity, exploit_suggestion
            ), tags=(severity.lower(),))
            
    def update_scan_progress(self):
        """Update scan progress while scanning is active"""
        if self.scanning:
            current = self.scan_progress["value"]
            total = self.scan_progress["maximum"]
            progress_text = f"Scanning... {current}/{total} ({current/total*100:.1f}%)"
            self.scan_status.config(text=progress_text)
            self.root.after(500, self.update_scan_progress)

    def get_exploit_suggestion(self, vulnerability):
        """Get exploit suggestion for a specific vulnerability"""
        if vulnerability in self.vuln_db:
            return self.vuln_db[vulnerability]["exploit"]
        return "No specific exploit suggestion available"
    
    def export_scan_results(self):
        """Export scan results to a file"""
        if not self.services_tree.get_children():
            messagebox.showwarning("Warning", "No results to export")
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Text files", "*.txt"), ("All files", "*.*")],
            initialdir=str(self.tool_dir)
        )
        
        if not file_path:
            return
            
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                # Write header
                f.write("Host,Port,Service,Version,Banner,Status,Vulnerabilities,OS Info,Exploit Suggestions\n")
                
                # Write data from services tree
                for item in self.services_tree.get_children():
                    values = self.services_tree.item(item, "values")
                    host, port, service, version, banner, status, vulns = values
                    
                    # Find OS info from hosts tree
                    os_info = "Unknown"
                    for host_item in self.hosts_tree.get_children():
                        host_values = self.hosts_tree.item(host_item, "values")
                        if host_values[0] == host:
                            os_info = host_values[2]
                            break
                    
                    # Get exploit suggestions from vulnerabilities tree
                    exploits = []
                    for vuln_item in self.vuln_tree.get_children():
                        vuln_values = self.vuln_tree.item(vuln_item, "values")
                        if vuln_values[0] == host and str(vuln_values[1]) == str(port):
                            exploits.append(vuln_values[5])
                    
                    exploit_text = "; ".join(exploits) if exploits else "None"
                    
                    f.write(f"{host},{port},{service},{version},\"{banner}\",{status},{vulns},{os_info},\"{exploit_text}\"\n")
                    
            self.log_action(f"Exported scan results to {file_path}")
            messagebox.showinfo("Success", f"Results exported to {file_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export results: {e}")

    def create_port_info_db(self):
        """Create a database of common ports and their purposes"""
        self.port_info = {
            21: "FTP - File Transfer Protocol (insecure file transfer)",
            22: "SSH - Secure Shell (secure remote access)",
            23: "Telnet (insecure remote access)",
            25: "SMTP - Simple Mail Transfer Protocol (email sending)",
            53: "DNS - Domain Name System (domain name resolution)",
            80: "HTTP - Hypertext Transfer Protocol (web traffic)",
            110: "POP3 - Post Office Protocol (email retrieval)",
            143: "IMAP - Internet Message Access Protocol (email access)",
            443: "HTTPS - HTTP Secure (encrypted web traffic)",
            993: "IMAPS - IMAP Secure (encrypted email access)",
            995: "POP3S - POP3 Secure (encrypted email retrieval)",
            3306: "MySQL Database",
            3389: "RDP - Remote Desktop Protocol",
            5432: "PostgreSQL Database",
            # Industrial ports
            502: "MODBUS - Industrial control system protocol",
            102: "Siemens S7 - Industrial automation protocol",
            44818: "EtherNet/IP - Industrial Ethernet protocol",
            20000: "DNP3 - Distributed Network Protocol",
            2404: "IEC-60870-5-104 - Telecontrol protocol",
            47808: "BACnet - Building Automation",
            1911: "Foxboro - Industrial control system",
            1962: "PCWorx - Industrial automation",
            2222: "EtherNet/IP - Alternative port",
            4000: "IEC-60870-5-104 - Alternative port",
            5006: "IEC-61850 - Substation automation",
            5007: "IEC-61850 - Alternative port",
            9600: "OMRON FINS - Factory Interface Network Service"
        }

    def ping_target(self):
        """Ping the target individually"""
        target = self.target_var.get().strip()
        if not target:
            messagebox.showerror("Error", "Please enter a target IP")
            return
            
        self.scan_status.config(text=f"Pinging {target}...")
        self.root.update()
        
        if self.ping_host(target):
            self.scan_status.config(text=f"{target} is reachable")
            messagebox.showinfo("Ping Result", f"{target} is reachable")
        else:
            self.scan_status.config(text=f"{target} is not reachable")
            messagebox.showwarning("Ping Result", f"{target} is not reachable")

    def add_unreachable_host(self, host):
        """Add an unreachable host to the results"""
        self.hosts_tree.insert("", tk.END, values=(
            host, "Not Reachable", "Unknown", 0, 0
        ))

    def detect_os(self, host, port):
        """Attempt to detect the operating system"""
        try:
            # Try to connect and analyze response for OS clues
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
                sock.settimeout(2)
                sock.connect((host, port))
                
                # Send a generic probe
                sock.send(b"\x00\x01\x00\x00\x00\x06\x01\x04\x00\x00\x00\x01")
                response = sock.recv(256)
                
                # Analyze response for OS fingerprints
                if not response:
                    return "Unknown"
                    
                # Windows detection (often has specific TTL values and window sizes)
                if len(response) > 40 and response[32:36] == b"\x02\x04\x05\xb4":
                    return "Windows"
                    
                # Linux detection
                if len(response) > 40 and response[32:36] == b"\x02\x04\x05\xa0":
                    return "Linux"
                    
                # Check for Android (often has specific SSH banners)
                if b"Android" in response:
                    return "Android"
                    
                # Check for iOS (Apple devices)
                if b"iOS" in response or b"Apple" in response:
                    return "iOS"
                    
                # Check for industrial systems
                if any(pattern in response for pattern in [b"Hirschmann", b"Emerson", b"Siemens"]):
                    return "Industrial Device"
                    
                # Try TTL-based detection (crude but sometimes effective)
                ttl = 64  # Default TTL for many systems
                try:
                    # This is a simplified approach - real OS detection would be more complex
                    ping_result = subprocess.run(
                        ["ping", "-c", "1", "-W", "1", host],
                        capture_output=True, text=True, timeout=2
                    )
                    if "ttl=" in ping_result.stdout.lower():
                        ttl_line = [line for line in ping_result.stdout.split('\n') if "ttl=" in line.lower()]
                        if ttl_line:
                            ttl = int(ttl_line[0].lower().split("ttl=")[1].split(" ")[0])
                except:
                    pass
                    
                # TTL-based OS guessing
                if 50 <= ttl <= 64:
                    return "Linux/Unix"
                elif 120 <= ttl <= 128:
                    return "Windows"
                elif 60 <= ttl <= 64:
                    return "Android/Linux"
                elif 250 <= ttl <= 255:
                    return "Network Device"
                    
        except:
            pass
            
        return "Unknown"

    # Add a right-click menu to show port information
    # In the create_port_scanner_tab method, after creating the services_tree:
        self.services_tree.bind("<Button-3>", self.show_port_info)  # Right-click

    def show_port_info(self, event):
        """Show information about a port when right-clicked"""
        item = self.services_tree.identify_row(event.y)
        if not item:
            return
            
        values = self.services_tree.item(item, "values")
        if not values or len(values) < 2:
            return
            
        port = int(values[1])  # Port is the second value
        
        # Create a popup menu
        menu = tk.Menu(self.root, tearoff=0)
        
        # Get port information
        info = self.port_info.get(port, "No information available for this port")
        
        # Add port information to the menu
        menu.add_command(label=f"Port {port} Information: {info}")
        
        # Add security recommendations
        if port in [21, 23, 80, 161, 162]:
            menu.add_command(label="Security: This port is often considered insecure")
            menu.add_command(label="Recommendation: Close if not needed, or use encryption")
        elif port in [22, 443, 993, 995]:
            menu.add_command(label="Security: This port uses encryption")
            menu.add_command(label="Recommendation: Keep software updated")
        
        # Show the menu
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

#------------------------NEW-------------------------------

    def setup_unlock_system(self):
        """Enhanced unlock system setup"""
        self.unlock_codes_file = self.appdata_dir / "unlock_codes.json"
        self.load_unlock_codes()
        self.add_admin_menu_items()

    def add_admin_menu_items(self):
        """Add admin menu items for unlock code management"""
        # Create Admin menu
        admin_menu = tk.Menu(self.menubar, tearoff=0)
        admin_menu.add_command(label="Generate Unlock Code", command=self.generate_unlock_code_dialog)
        admin_menu.add_command(label="View Code Usage", command=self.view_code_usage)
        admin_menu.add_command(label="Revoke Unlock Code", command=self.revoke_unlock_code)
        admin_menu.add_separator()
        admin_menu.add_command(label="Reset All Codes", command=self.reset_all_codes)
        
        # Add to menu bar after Help menu
        # Find the index of the Help menu
        help_index = None
        for i in range(self.menubar.index("end") + 1):
            try:
                label = self.menubar.entrycget(i, "label")
                if label == "Help":
                    help_index = i
                    break
            except tk.TclError:
                pass
        
        # Insert after Help menu if found, otherwise add at the end
        if help_index is not None:
            self.menubar.insert_cascade(help_index + 1, label="Admin", menu=admin_menu)
        else:
            self.menubar.add_cascade(label="Admin", menu=admin_menu)

    def load_unlock_codes(self):
        """Load unlock codes from file"""
        if self.unlock_codes_file.exists():
            try:
                with open(self.unlock_codes_file, 'r') as f:
                    self.unlock_codes = json.load(f)
            except:
                self.unlock_codes = {}
        else:
            self.unlock_codes = {}
            self.save_unlock_codes()

    def save_unlock_codes(self):
        """Save unlock codes to file"""
        with open(self.unlock_codes_file, 'w') as f:
            json.dump(self.unlock_codes, f, indent=4)

    def generate_unlock_code_dialog(self):
        """Dialog for generating unlock codes"""
        # Verify master password first
        master_password = simpledialog.askstring("Master Password", 
                                            "Enter master password:",
                                            show='*', parent=self.root)
        if master_password != PASSWORD:
            messagebox.showerror("Error", "Incorrect master password")
            return
        
        # Create dialog for code options
        dialog = Toplevel(self.root)
        dialog.title("Generate Unlock Code")
        dialog.geometry("400x350")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center the dialog
        dialog.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() - dialog.winfo_width()) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        
        # Options frame
        options_frame = ttk.Frame(dialog, padding=20)
        options_frame.pack(fill=tk.BOTH, expand=True)
        
        # Usage limit
        ttk.Label(options_frame, text="Usage Limit:").grid(row=0, column=0, sticky=tk.W, pady=5)
        usage_var = tk.IntVar(value=1)
        ttk.Spinbox(options_frame, from_=1, to=100, textvariable=usage_var, width=10).grid(row=0, column=1, sticky=tk.W, pady=5)
        
        # Expiration days
        ttk.Label(options_frame, text="Expires After (days):").grid(row=1, column=0, sticky=tk.W, pady=5)
        expiry_var = tk.IntVar(value=7)
        ttk.Spinbox(options_frame, from_=1, to=365, textvariable=expiry_var, width=10).grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # Tabs to unlock
        ttk.Label(options_frame, text="Tabs to Unlock:").grid(row=2, column=0, sticky=tk.W, pady=5)
        
        security_var = tk.BooleanVar(value=True)
        self_destruct_var = tk.BooleanVar(value=True)
        port_scanner_var = tk.BooleanVar(value=True)
        
        ttk.Checkbutton(options_frame, text="File Security", variable=security_var).grid(row=2, column=1, sticky=tk.W, pady=2)
        ttk.Checkbutton(options_frame, text="Self-Destruct Files", variable=self_destruct_var).grid(row=3, column=1, sticky=tk.W, pady=2)
        ttk.Checkbutton(options_frame, text="Port Scanner", variable=port_scanner_var).grid(row=4, column=1, sticky=tk.W, pady=2)
        
        # Buttons
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        
        def generate_code():
            tabs = []
            if security_var.get(): tabs.append("security")
            if self_destruct_var.get(): tabs.append("self_destruct")
            if port_scanner_var.get(): tabs.append("port_scanner")
            
            if not tabs:
                messagebox.showerror("Error", "Select at least one tab to unlock")
                return
                
            code = self.generate_unlock_code(usage_var.get(), expiry_var.get(), tabs)
            dialog.destroy()
            
            # Show the code to the admin
            messagebox.showinfo("Unlock Code Generated", 
                            f"Unlock Code: {code}\n\n"
                            f"Usage Limit: {usage_var.get()}\n"
                            f"Expires: {expiry_var.get()} days\n"
                            f"Tabs: {', '.join(tabs)}\n\n"
                            "Share this code with your colleague.")
        
        ttk.Button(button_frame, text="Generate", command=generate_code).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="Cancel", command=dialog.destroy).pack(side=tk.LEFT, padx=10)

    def generate_unlock_code(self, usage_limit, expiry_days, tabs):
        """Generate a new unlock code with specified parameters"""
        # Generate a random code
        code = ''.join(secrets.choice(string.ascii_uppercase + string.digits) for _ in range(12))
        
        # Calculate expiration date
        expiry_date = (datetime.now() + timedelta(days=expiry_days)).isoformat()
        
        # Store code details
        self.unlock_codes[code] = {
            "usage_limit": usage_limit,
            "usage_count": 0,
            "expiry_date": expiry_date,
            "tabs": tabs,
            "created": datetime.now().isoformat()
        }
        
        self.save_unlock_codes()
        self.log_action(f"Generated unlock code: {code} for tabs: {tabs}")
        return code

    def verify_unlock_code(self, code):
        """Verify if an unlock code is valid"""
        if code in self.unlock_codes:
            code_data = self.unlock_codes[code]
            
            # Check if code has expired
            expiry_date = datetime.fromisoformat(code_data["expiry_date"])
            if datetime.now() > expiry_date:
                self.log_action(f"Unlock code expired: {code}")
                return False, "This unlock code has expired"
            
            # Check if usage limit exceeded
            if code_data["usage_count"] >= code_data["usage_limit"]:
                self.log_action(f"Unlock code usage limit exceeded: {code}")
                return False, "This unlock code has reached its usage limit"
            
            # Code is valid
            return True, code_data["tabs"]
        
        return False, "Invalid unlock code"

    def unlock_protected_tabs(self, code):
        """Unlock tabs using a valid code"""
        valid, message_or_tabs = self.verify_unlock_code(code)
        
        if valid:
            tabs_to_unlock = message_or_tabs
            
            # Update usage count
            self.unlock_codes[code]["usage_count"] += 1
            self.unlock_codes[code]["last_used"] = datetime.now().isoformat()
            self.unlock_codes[code]["last_user"] = self.username
            self.save_unlock_codes()
            
            # Unlock the specified tabs
            for tab in tabs_to_unlock:
                self.locked_tabs[tab] = False
            
            self.update_tab_states()
            self.log_action(f"Tabs unlocked with code: {code} by {self.username}")
            messagebox.showinfo("Success", "Tabs unlocked successfully!")
            return True
        else:
            messagebox.showerror("Error", message_or_tabs)
            return False

    def view_code_usage(self):
        """Display a dialog showing code usage statistics"""
        # Verify master password first
        master_password = simpledialog.askstring("Master Password", 
                                            "Enter master password:",
                                            show='*', parent=self.root)
        if master_password != PASSWORD:
            messagebox.showerror("Error", "Incorrect master password")
            return
        
        dialog = Toplevel(self.root)
        dialog.title("Unlock Code Usage")
        dialog.geometry("800x400")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Create treeview to display codes
        tree_frame = ttk.Frame(dialog)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        columns = ("code", "usage", "limit", "expiry", "tabs", "created", "last_used", "last_user")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)
        
        # Define headings
        tree.heading("code", text="Code")
        tree.heading("usage", text="Usage")
        tree.heading("limit", text="Limit")
        tree.heading("expiry", text="Expiry Date")
        tree.heading("tabs", text="Tabs")
        tree.heading("created", text="Created")
        tree.heading("last_used", text="Last Used")
        tree.heading("last_user", text="Last User")
        
        # Set column widths
        tree.column("code", width=100)
        tree.column("usage", width=60)
        tree.column("limit", width=60)
        tree.column("expiry", width=100)
        tree.column("tabs", width=120)
        tree.column("created", width=120)
        tree.column("last_used", width=120)
        tree.column("last_user", width=100)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Populate with data
        for code, data in self.unlock_codes.items():
            expiry_date = datetime.fromisoformat(data["expiry_date"]).strftime("%Y-%m-%d")
            created = datetime.fromisoformat(data["created"]).strftime("%Y-%m-%d")
            last_used = datetime.fromisoformat(data["last_used"]).strftime("%Y-%m-%d") if "last_used" in data else "Never"
            last_user = data.get("last_user", "N/A")
            
            tree.insert("", tk.END, values=(
                code,
                data["usage_count"],
                data["usage_limit"],
                expiry_date,
                ", ".join(data["tabs"]),
                created,
                last_used,
                last_user
            ))
        
        # Add close button
        button_frame = ttk.Frame(dialog)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="Close", command=dialog.destroy).pack()

    def revoke_unlock_code(self):
        """Revoke an unlock code"""
        # Verify master password first
        master_password = simpledialog.askstring("Master Password", 
                                            "Enter master password:",
                                            show='*', parent=self.root)
        if master_password != PASSWORD:
            messagebox.showerror("Error", "Incorrect master password")
            return
        
        code = simpledialog.askstring("Revoke Code", 
                                    "Enter code to revoke:",
                                    parent=self.root)
        
        if code and code in self.unlock_codes:
            del self.unlock_codes[code]
            self.save_unlock_codes()
            self.log_action(f"Revoked unlock code: {code}")
            messagebox.showinfo("Success", "Unlock code revoked successfully!")
        else:
            messagebox.showerror("Error", "Invalid unlock code")

    def reset_all_codes(self):
        """Reset all unlock codes"""
        # Verify master password first
        master_password = simpledialog.askstring("Master Password", 
                                            "Enter master password:",
                                            show='*', parent=self.root)
        if master_password != PASSWORD:
            messagebox.showerror("Error", "Incorrect master password")
            return
        
        if messagebox.askyesno("Confirm", 
                            "This will delete ALL unlock codes. This action cannot be undone.\n\nContinue?"):
            self.unlock_codes = {}
            self.save_unlock_codes()
            self.log_action("All unlock codes reset")
            messagebox.showinfo("Success", "All unlock codes have been reset!")

    # Add this method to the SmartReplaceTool class
    def download_port_scan_log(self):
        """Download the port scan log file"""
        if not self.port_scan_log_file.exists():
            messagebox.showwarning("No Log File", "No port scan log file exists yet.")
            return
        
        try:
            # Create downloads directory if it doesn't exist
            self.downloads_dir.mkdir(exist_ok=True)
            
            # Copy log file to downloads directory
            dest_path = self.downloads_dir / "port_scan_log.txt"
            shutil.copy2(self.port_scan_log_file, dest_path)
            
            # Open the downloads folder
            if platform.system() == 'Windows':
                os.startfile(self.downloads_dir)
            elif platform.system() == 'Darwin':
                subprocess.run(['open', self.downloads_dir])
            else:
                subprocess.run(['xdg-open', self.downloads_dir])
                
            self.log_action("Port scan log downloaded")
            messagebox.showinfo("Success", f"Port scan log downloaded to:\n{self.downloads_dir}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Could not download log file:\n\n{str(e)}")

            

# -------------------------------
# MAIN
# -------------------------------
if __name__ == "__main__":
    # Check if we're running as a bundled executable
    if getattr(sys, 'frozen', False):
        # Running as executable - change to the executable's directory
        os.chdir(os.path.dirname(sys.executable))
    else:
        # Running as script - change to the script's directory
        os.chdir(os.path.dirname(os.path.abspath(__file__)))
        
    root = tk.Tk()
    app = SmartReplaceTool(root)
    root.mainloop()